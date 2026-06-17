import csv
import hashlib
import random
import datetime as _ist_dt
from io import StringIO
from fastapi import FastAPI, Form, Depends, Cookie, Response, UploadFile, File, HTTPException, status
from fastapi.responses import HTMLResponse, RedirectResponse
from fastapi.exception_handlers import http_exception_handler
from starlette.exceptions import HTTPException as StarletteHTTPException
from sqlalchemy.orm import Session
from sqlalchemy import text

import models
import templates
from database import engine, get_db, SessionLocal

# Force absolute structural table synchronization on startup
models.Base.metadata.create_all(bind=engine)

# -------------------------------------------------------------
# IST (INDIA STANDARD TIME) DISPLAY HELPERS
# -------------------------------------------------------------
# All timestamps are stored in the database as naive UTC datetimes
# (via datetime.utcnow() defaults). For display anywhere in the UI,
# convert to IST (UTC+5:30) before formatting.
IST_OFFSET = _ist_dt.timedelta(hours=5, minutes=30)


def to_ist(dt):
    """Convert a naive UTC datetime to a naive IST datetime for display.
    Returns None if dt is None."""
    if dt is None:
        return None
    return dt + IST_OFFSET


def fmt_ist(dt, fmt="%d-%m-%Y %H:%M"):
    """Format a stored (naive UTC) datetime as IST. Returns '' if dt is None."""
    if dt is None:
        return ""
    return to_ist(dt).strftime(fmt)

# -------------------------------------------------------------
# MULTI-VENDOR LOT / FIFO HELPERS
# -------------------------------------------------------------
# Each Item can have multiple ItemLot rows — one per (vendor, price)
# combination. Item.current_stock is kept as a cached running total
# equal to the sum of all of that item's lot quantities, so existing
# code that reads current_stock for display/low-stock checks keeps
# working unchanged.

_LOT_PRICE_EPS = 0.005  # treat prices within half a paisa as the same lot


def add_to_lot(db: Session, item: "models.Item", vendor: str, price: float, quantity: int, received_at=None):
    """Add `quantity` units to the matching (item, vendor, price) lot,
    creating it if it doesn't exist yet. Also bumps item.current_stock."""
    import datetime as _dt
    if quantity <= 0:
        return
    vendor = (vendor or item.supplier or "Approved Vendor").strip() or "Approved Vendor"
    price = float(price) if price is not None else 0.0

    lot = db.query(models.ItemLot).filter(
        models.ItemLot.item_id == item.id,
        models.ItemLot.vendor == vendor,
        models.ItemLot.price.between(price - _LOT_PRICE_EPS, price + _LOT_PRICE_EPS)
    ).first()

    if lot:
        lot.quantity += quantity
    else:
        db.add(models.ItemLot(
            item_id=item.id, vendor=vendor, price=price,
            quantity=quantity, received_at=received_at or _dt.datetime.utcnow()
        ))

    item.current_stock = (item.current_stock or 0) + quantity


def consume_fifo(db: Session, item: "models.Item", quantity: int):
    """Consume `quantity` units from this item's lots, oldest first.
    Returns the total cost (sum of qty_taken * lot.price) of the units
    consumed. Reduces item.current_stock accordingly. Caller is
    responsible for checking sufficient stock beforehand; if lots run
    out early, the remainder is costed at the item's current `price`."""
    remaining = quantity
    total_cost = 0.0

    lots = db.query(models.ItemLot).filter(
        models.ItemLot.item_id == item.id,
        models.ItemLot.quantity > 0
    ).order_by(models.ItemLot.received_at.asc(), models.ItemLot.id.asc()).all()

    for lot in lots:
        if remaining <= 0:
            break
        take = min(lot.quantity, remaining)
        if take > 0:
            lot.quantity -= take
            total_cost += take * (lot.price or 0.0)
            remaining -= take

    if remaining > 0:
        # Lots didn't cover the full quantity (e.g. legacy stock with no
        # lot records yet) — cost the shortfall at the item's list price.
        total_cost += remaining * (item.price or 0.0)
        remaining = 0

    item.current_stock = max(0, (item.current_stock or 0) - quantity)
    return total_cost


def reconcile_stock_delta(db: Session, item: "models.Item", delta: int):
    """Apply a manual stock adjustment (e.g. from the Edit Item modal) of
    `delta` units (positive or negative) to the lot ledger, keeping
    sum(lot.quantity) == item.current_stock."""
    if delta == 0:
        return
    if delta > 0:
        add_to_lot(db, item, "Manual Adjustment", item.price or 0.0, delta)
    else:
        consume_fifo(db, item, -delta)


def build_txn_logs(db: Session, items):
    """Build per-item transaction logs used by both the Material Flow view
    and the merged Inventory Configuration table.

    Each inward (IN) row carries its OWN vendor and per-unit price taken
    from its GRN record (not the item's moving catalog value), so a later
    inward at a different vendor/price no longer rewrites historical rows.

    Price Diff / Percentage Diff are computed batch-over-batch: each IN is
    compared to the immediately preceding IN of the same item, so they show
    how the newly entered price changed versus the previous purchase.
    Returns {item_id: [entry, ...]} sorted newest-first.
    """
    import datetime as _dt
    transactions = db.query(models.Transaction).all()
    assignments = db.query(models.MaterialAssignment).all()
    grns = db.query(models.GRNRecord).all()
    item_by_id = {i.id: i for i in items}
    _MIN = _dt.datetime.min

    # Legacy fallback: per-unit price by (item_id, minute) from IN transactions
    txn_price_lookup = {}
    for t in transactions:
        if t.type == "IN" and t.quantity and t.quantity > 0 and t.total_value:
            per_unit = round(t.total_value / t.quantity, 2)
            ts_str = fmt_ist(t.timestamp)
            txn_price_lookup.setdefault(t.item_id, {})[ts_str] = per_unit

    # Vendor recovery for legacy GRNs (no stored vendor): the actual vendor
    # for each batch was recorded in item_lots at inward time. We match a GRN
    # to its lot by price so old rows show their ORIGINAL vendor instead of
    # the item's current catalog supplier. Include consumed (qty 0) lots too,
    # since they still hold the vendor history.
    lots_by_item = {}
    for lot in db.query(models.ItemLot).all():
        lots_by_item.setdefault(lot.item_id, []).append(lot)

    def _recover_vendor(item_id, unit_price, when, fallback):
        lots = lots_by_item.get(item_id)
        if not lots:
            return fallback
        # candidates whose price matches this batch's per-unit price
        cands = [l for l in lots if abs((l.price or 0.0) - (unit_price or 0.0)) <= 0.01]
        if not cands:
            return fallback
        # prefer the lot that existed by this GRN's time, closest in time
        prior = [l for l in cands if (l.received_at or _MIN) <= (when or _MIN)]
        pool = prior if prior else cands
        best = min(pool, key=lambda l: abs(((l.received_at or _MIN) - (when or _MIN)).total_seconds()))
        return best.vendor or fallback

    # IN entries from GRNs (keep real datetime for correct chronological sort)
    in_entries = {}
    for g in grns:
        item_obj = item_by_id.get(g.item_id)
        catalog_price = (item_obj.price if item_obj else 0.0) or 0.0
        catalog_vendor = (item_obj.supplier if item_obj else "\u2014") or "\u2014"
        gdt = g.timestamp or _MIN
        grn_date = fmt_ist(g.timestamp)
        unit_price = getattr(g, "unit_price", None)
        if unit_price is None:
            unit_price = txn_price_lookup.get(g.item_id, {}).get(grn_date, catalog_price)
        # Use the GRN's own stored vendor; for legacy rows, recover it from the
        # matching lot by price, and only then fall back to the catalog vendor.
        vendor = getattr(g, "vendor", None)
        if not vendor:
            vendor = _recover_vendor(g.item_id, unit_price, gdt, catalog_vendor)
        person = getattr(g, "received_by", "") or (g.uploader.username if g.uploader else "\u2014")
        entry = {
            "date": grn_date, "type": "IN", "qty": g.quantity,
            "uom": g.uom or "Nos", "person": person or "\u2014", "direction": "IN",
            "vendor": vendor, "price": round(unit_price or 0.0, 2),
            "grn_no": getattr(g, "grn_no", None) or "\u2014",
            "challan_no": getattr(g, "challan_no", None) or "\u2014",
            "serial_no": "\u2014",
            "price_diff": None, "pct_diff": None,
        }
        in_entries.setdefault(g.item_id, []).append((gdt, entry))

    # Compute batch-over-batch diffs per item
    for lst in in_entries.values():
        lst.sort(key=lambda x: x[0])  # oldest first
        prev_price = None
        for _dtk, e in lst:
            if prev_price is not None and prev_price > 0:
                diff = round(e["price"] - prev_price, 2)
                if abs(diff) > 0.01:
                    e["price_diff"] = diff
                    e["pct_diff"] = round((diff / prev_price) * 100, 2)
            prev_price = e["price"]

    # OUT entries from material assignments
    # OUT entries (issues) and RETURN entries from material_assignments.
    # For serial-tracked items, fetch the serial via asset_unit_id link.
    unit_serials = {u.id: (u.serial_no or "\u2014") for u in db.query(models.AssetUnit).all()}
    out_entries = {}
    for a in assignments:
        adt = a.timestamp or _MIN
        is_ret = bool(getattr(a, "is_return", False))
        serial = "\u2014"
        aid = getattr(a, "asset_unit_id", None)
        if aid:
            serial = unit_serials.get(aid, "\u2014")
        if is_ret:
            entry = {
                "date": fmt_ist(a.timestamp),
                "type": "RETURN", "qty": a.quantity, "uom": a.uom or "Nos",
                "person": a.issued_to or "\u2014", "direction": "IN",
                "serial_no": serial,
                "is_return": True,
            }
        else:
            entry = {
                "date": fmt_ist(a.timestamp),
                "type": "OUT", "qty": a.quantity, "uom": a.uom or "Nos",
                "person": a.issued_to or "\u2014", "direction": "OUT",
                "serial_no": serial,
            }
        out_entries.setdefault(a.item_id, []).append((adt, entry))

    logs = {}
    for iid in set(in_entries) | set(out_entries):
        combined = in_entries.get(iid, []) + out_entries.get(iid, [])
        combined.sort(key=lambda x: x[0], reverse=True)  # newest first
        logs[iid] = [e for _dtk, e in combined]
    return logs


app = FastAPI(title="EIPL Enterprise Procurement Framework")

# -------------------------------------------------------------
# ASSET CLASS CATEGORIES
# -------------------------------------------------------------
ASSET_CLASSES = ["Fixed Assets and Equipments", "Consumables", "Tools and Tackles"]
DEFAULT_ASSET_CLASS = "Consumables"
# Only items in TRACKED_CATEGORIES are serial-tracked (one AssetUnit per physical thing).
TRACKED_CATEGORIES = ("Fixed Assets and Equipments",)


def is_tracked(item_or_category) -> bool:
    """True if the item (or category string) is serial-tracked.
    Accepts either an Item instance or a raw category string."""
    cat = item_or_category if isinstance(item_or_category, str) else getattr(item_or_category, "category", None)
    return cat in TRACKED_CATEGORIES


def category_options_html(selected=None):
    """Return <option> tags for the asset class dropdown, marking `selected`."""
    sel = selected if selected in ASSET_CLASSES else None
    opts = ""
    for c in ASSET_CLASSES:
        is_sel = " selected" if (sel == c) else ""
        opts += f'<option value="{c}"{is_sel}>{c}</option>'
    return opts



# Serve static assets (company logo etc.) from ./static folder
import os as _os
from fastapi.staticfiles import StaticFiles
_os.makedirs("static", exist_ok=True)
app.mount("/static", StaticFiles(directory="static"), name="static")

# Logo embedded as base64 — works on Render without a persistent filesystem
LOGO_DATA_URL = "data:image/png;base64,/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAGmAZADASIAAhEBAxEB/8QAHQABAAEEAwEAAAAAAAAAAAAAAAUEBgcIAQIDCf/EAF8QAAEDAwICBQYGCRAHBgUFAAECAwQABREGIRIxBxNBUWEUIjJxgZEIQlKhscEJFRYjYnJzktEXJDM0NTdDU1RWgpOUorLSJTZjdHXC4SZEVYOjs0VXZJXwGEaEpLT/xAAbAQEAAgMBAQAAAAAAAAAAAAAABAUBAgMGB//EADcRAAICAQMCAwUGBQQDAAAAAAABAgMRBBIhBTETQVEiYXGRoRQygbHB8AYVUtHhFiNC8SQzNP/aAAwDAQACEQMRAD8A3LpSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClK4UoJBJwAOZoDmlRj18gpcLUcuTHhzRGRx49Z5D2mvPrr3J/YokaEj5T6y4v8ANTt89AS+RXVxxDaeJxaUDvUcVFi1Snt5l4lufgs4ZT82/wA9d27BaUq4lQ0PK+U8S4T+cTQHd682po4cuMUHuDgJ9wry+6G1H0JC3PxGVq+gVXsxIzIAajstgfJQBXrjxoCLF+hHk3NV6obn+Wn2/gj0kzE/jRHB9VSmPGuceJoCK+6G05wqWG/x21J+kVUR7rbX9mrhFWe4OjPuqsKcjB3FU78CE+MPRI7g/CaBoCoSoKTxJIIPaK5yO+opVgtgPEy0uMrvYdU39BxXU2+4sftS8OqA+JJbS4PeMGgJelRAm3aN+27Yl9A5riOZP5isH5zXvDvECS51KXg29/FOgoX7jQEhSgOezFKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlcKOBzoDmvKRIYjtKdfdQ02nmpZwBUa7dXZLio9oaTJWDhb6jhls+v4x8BXeJaGy8JNwdVOkp3SpwYQj8VHIevnQHn9spk48NqiEtn/vEjKG/WE+kr5hXIsokELukp2crnwHzGh/QHP25qWAArnI76A82GWmWw200htA5JSkAD2V32qkulygWuIqXcZ0aHHQMqdfcCEj2msT6v+EV0f2Qrat78q+yE/Fht4bz+UVge7NdqdPbc8VxbOc7oV/eeDMeRigUDyrUTUvwoNWTCtFis1ttbZ9FbxMhz6k/NWOr50tdI95KvLNXXJCFc246wyn1YQBVrV0HVT5lhEKfU6o9ss38fksMI433m2kjmVqCfpqIl6w0pEz5TqWztEcwua2D9NfOuZPnzV8c2dLkq73n1LPzmqbhTnPCnPqqbH+HP6rPocJdVflE+hjnSX0fNnC9ZWIH/AH1H6a7M9I+gXiA3rGxKJ/8ArUfpr54gDsAoQO0Cun+nK/62afzWf9J9IoeqNNzCBEv9qfJ5BuY2o/MalG3W3E8SFBY70nP0V8yQkA5CQD3gVIW69Xm2qCrdd7hEI3HUSVox7jXKf8OP/jZ9DePVfWJ9KOIUyK0LsPTZ0m2cpDWqJEtsfEmoS+D7SM/PWSdL/CmurKkN6k01GlI+M7BdLavXwqyD7xUG3oWrhzHD+BIh1KmXfg2r27q8ZsOLNa6uVHaeT3LTnHq7qxvo7p06O9SKQym8/ayUvYMXBPUnPcFeiffWS2HmnmUOsuIcbUMpWlQKSPAiqqymyp4nFomwsjNZi8kZ9rJkPe1zlpQP4CSS436gfST7zRu8eTrDd2jqhKOwcJ4mVHwWOXtxUvkcs11W2haSlaQpJGCDuDXM3OUrSoApIIIyCORrmodVrehLLtme6gZyYzmVMq9Xaj2e6vaFdUOviJKaVDl/xTnJfihXJQ+egJKlKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClDVBdLgmHwNNtqflO7NMJO6vE9yR2mgPe4TY8FjrpLgSnOAOZUe4DmT4VGJizLseKfxxYZ3TFScLWP9oR/hHtr3t9uc68TrgtMiZjzSPQZHyUD6+ZqUFAebLLbLaW2kJQhIwlKRgAeArua8Z0uPDiuyZT7bDDSSpxxxQSlAHaSeQrXHpa+EkzGL1q0ChEl0ZSu5vJy2k/7NB9L8Y7eBqRptJbqZba1k43XwpWZszrrHWGnNI24z9RXWPBax5iVqytw9yUDdR9QrXPpD+E5PkrciaJtgiN8hOmpCnD4pb5J9ufVWAb7eLrfrm5crzcJE+Y5up19ZUr1DuHgNqoa9Vo+g01e1b7T+hTX9RnPiHCJXUuo79qWYZd/u8y5PE5BfdKgn1J5AeoVFUpV5GEYLbFYRXuTk8sUpStjApSlAKUpQClKUApSlADvtVz6K1/q/Rz4XYL5KjNg5MdSusZV60KyPdirYpWllcLFtmso2jOUXmLNqujn4Tdslqbha2txtzpwPLYgK2Se9SPST7M1n+x3e23q3t3G0zo86I6MoeYcC0n2jt8K+a1XBojWepdF3ITtO3R6Ion741niadHctB2P0+NUGs6BXP2qHh+nkWVHUpR4s5PoxVPPhR5rBZktpcRzGeYPeD2HxrDXRB0/wBg1Ytm16hDdkvCsJTxL/W76vwVH0T+Cr2E1moKBHOvL30WaeeyxYZcV2wtW6LyQ4em2f8AbJcmQB/DYy6yPwwPSHiN++pdh5t9tDrK0uNrGUqScgivTA7qhn4Um3Orl2pAU2o8T0TOEq71I+Srw5GuJ0JmlU1vmx5sYPx18SScEEYKT2gjsI7qqaAUpSgFKUoBSlKAV5S30Ro7j7ueBtJUrAycCvUnAzVt6vvsGEkQnescW6klSGxvjsBPZnb2ZoClla1S05xtW5bscFHndZwrAVyJSRt2jnV2NLDjSXE8lAEVj+BBE2MzHUygmVJ854E/sSckDv4c5x34q/2EBtlDaeSUgD2UB3pSlAKUpQClKUApSlAKUqhu84QmkBDZekOngYaB3Wr6gOZPYKA63W4GMUR47fXzXv2JrO3ipR7Ejvpa7eIvE+84X5b37M8obnwA7EjsFcWiAYwW/IWHpj+7zuPclPckdgqR5CgONhVs9IWuLBoaxrut+mBpO4ZZRu6+r5KE9p8eQ7ah+mPpOsvR3Y/KJZEm5vgiHBSrCnD8pXyUDtPsFaRa41ZfNZ3929X6YqRIXshA2bZR2IQnsA+ftq36b0qere+XEfz+BB1etjT7MeWXP0vdLeo+kOapp9wwLMlWWbe0rzfBTh+Or5h2CseUpXtKaK6IKFawigsslZLdJ8ilKV1NBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgGKzX0J9PN20mtizanW/dLGCEocJ4n4o8CfTT+Cdx2HsrClKj6nS1amGyxZOlVs6pbos+k9gvNtvtqYutpmszYUhPE080rIUPqPgdxUiRkVoF0QdJ186OruHYilSrU8oGXAWrCVj5SfkrHf29tbv6F1ZZdY6ej3qxSQ/GdGFA7LaX2oWOxQrxHUOm2aOXPMX2Z6HS6uN69GVNxgvNSTcbYkCTj760ThMgdx7ldx99VdtnMzo/XNEjB4VoUMKQoc0kdhFVXOoi6RHo0k3SAnjexh9gcn0j/AJh2H2VXEsl6VTwJTMyK3Ijq4m1jIPd4HuNVFAKUpQClKorhMSygpSpKVhPEpSj5raflK/R20Bxc5ojtrSlxCFJTxKWr0Wk/KP1DtqxHJMaXc0IZR1CnsgvO7uLSc+cr14J8AAO2vC73RV6ddix18EFo8RU4rhU+vPM+PcOwVGR43l99jRmUqK3VJ87PZnJPfyFDGS/rLB8jgQlKUVLckpVxEY83hUEj3fTVyjlVFcEhCIoSMBMhAA7hyqtHKhkUpSgFKUoBSlKAUpXCjigPGdKZiRXJD6+BtsZJ+oeNUFpivuvKuk5GJTqcIbP8A32J9Z5k99eSP9L3XrTvBhrwjuddHNXqT2ePqqbAAoDgkAZqxOmXpHtnR3ptU6TwyLg/lEGGFYLq+89yB2n2c6ndf6ptWjtLS7/d3eGPHT5qAfOdWfRQnvJNaD9Ier7trjVMm/Xdwlxw8LLIPmMNg+ahPgPnOTVt0rpr1c90vur6+4g63VKmOF3ZRas1Dd9U36Te73LVKmSFZUo8kjsSkdiR2CoulK9xCEYRUYrCR55tyeWKUpWxgUpSgFX9bOifU8ryN1/yaOw+QXD1nEtlJGclPafAGrB9uKyPofTOqtUT0XuVdJcKDJSeOS0+QtwJ83hCQduXPlsaga+2dUNykor38knTQjOWHHJeMXSnRnBWmzSFwX5x81RflffSr2EAHwq29RdEE43rFhkMpt608RMlzzmj8nYZUPGrrHRVo9bS2ymYt34zplErB7+7NW+50lSdLCTpyXBNyl25xUduUXeEOJT6JWMZzjGcd1UFF18pN6Wbk/NP80WNldaX+9FJeWCmR0Ly+qyu/wAcOY5JjqKffnNY91TYZ+m7w5bLgEdYkBSVoOUrSeShWS+jrXmqtRayahvNx3IKwpTyG2cBlIGyuLnzwN+dRPTzarsjUSbw8nrLc42hllaRs2QN0q7iTk57asNJqdVXqvB1Mlyska6qqVPiVJ8GNqUpV8VwpSlAKUpQClKUApSlAKUpQCry6JOkO8dHepE3GApT0J0hM2GVYS+j6lDsP1VZtK521Qtg4TWUzaE5Qlui+T6PaL1LaNV6di3yzSg/EkpyD8ZCu1Ch2KB2Iqa2NaIdAnSdK6PNShMlTjtimqCZrA34OwOpHyh2949lbzW2XHnw2ZkR9t+O+2lxpxBylaSMgg92K8F1HQT0dm3/AIvsz0mk1KvhnzI2YPtPNVPbBEJ9X66QOTaux0eHYr31MoIKcg5B7a4dQhaFJWkKSoYIO4I7qibStVumGzuqJaIK4a1dqO1Ge9P0VAJRM0NKpJsrqstoKePGSVeihPyj4fTQHE+WGQUtqTx8PESr0UD5Sv0dtWbqC4JLakeeW+MDh+O64eXF443x8UeJGK25SHZD5hRlFB/ZHnVj9iSBu4v8LHop7NieyrccfWtfFHhvJYdQWbediojOFK3+Mo8zjPYKAp5rkd9htuClhoNqS2pxw4Uta+ZJ5Y27OQ2q5eji1r62Td5KfvhJabJ8PSI8OQ9lRFthqejtoSkKkRXOpQ2GweFwqGVLV3bYwNyBWR7fHbiQ24zQ81tOPX40MI87p+xMHukN/wCIVVjlVJdv2qg9zzZ/viquhkUpSgFKUoBSlKAVF32S7wNQIiuGVKJSlQ/g0D0l+wfORUmpSUglRwAMknsqJsKTLeeu7g/Z/Mjg/FZB2/OO/uoCQhRmosVqOwnhbbSEpFejziWkKWtQShIJUonAA7zXYkAZrA3wuekM2HTSNI2t/huN2bJkKSfOajcj6is7eoGu+molqLVXHuzndaqoOTMKfCN6SnNeatVEt7x+0NtWpuIkcnl8lPH18h4es1iygpX0PT0Q09arh2R5e2x2ScpClKV2OYpSlAKVP6E081qW7rhvXJmAhtvrVKc5qSFAEDszg9tZdsmmejaXLk2SDDjzZUVH39RWtShvjPHnGc91V+r6lXpntab9ceRKp0srVlPBgOrj05qjVUSC3ZLJKkBC3FcDbSOJRKxgge7I7jmqPWVncsWpJtuU062026eoLnNTZ3Sc9u1XhpPWN3s+h1C16cjKaiBSXZ3csnZahzOAce6ttTYrKoyhFSz69jWqO2bUnjBIaBuNw0O8ljUlokxEXSTl2fJd8xACTjYZ3znn3+FT2qbTo7XNvuNzts1py4xI5JeZWRjhBICknmDgjPz1bXSRrizX7Rca2I45FyJadcWG+FDSwPO9faMDberK09d12u3XGPDUUyri15O46s4QyznKj4qPLwHeaq4aOy3/AMjmE8+XZkyV8Yf7X3okv0NXmRbNaxWGkFbVwxHdSBk4O4V7D82azH0mzrPE0fPZu7jYTIZUhls7rW5jzeEeBwc9lWZ0Ls6YtqZk77ewn5KkpTh5vqVsgZzjiO4PeO6sedId5VfNXz5gkF+Ol0txjnzQ2NhjwPP21izTrWa/KylHu/UzC10af1yW+M4351zXs3ElOsF9uO4trJHEE7bc/dXh416JYKo5pXGa5rIFKUoBSlKAUpSgFKUoBSlKAGtkPgi9JyoslOgL0+Sy6Sq1OLPoL5qZ9R5p8cjtFa316Rn3oslqTGdW080sONuIOFIUDkEHvBqJrdJHVVOuX4fE7UXOmakj6ZpORnvqhvcNUyKAyrgktKDjC/krHL2HkfA1Z/QXrxrX2hY1zcWkXGP+t57Y+K6B6WO5Qwoes91XlOldWottFPWcPEoq9FtPylfUO2vntlcqpuEu6PTwmpxUl2ZTR7omTBQ4hIQ9g9alfJkjZXF6j76jJz7hIZYyXFELysZKc8lrHao/FT2czVNJWqFd0hvBTOIPC6ccLoGzi/WOzswK7XSSq0WsTWAHZLq8R+tG7ij6Th9mcDsHrrQ3KScWm5zWn45b4nSXZ7jisjA34Se3vPfVPD42rlJlR2C86tXC064sICVqScKx8lKQfAb1HvIlSW2oaJRU3Ib615RbSC6oqyT3+o7dlVUuSmPcVB3gksRWSxGWThtSkk4LgHMnkB24oCc0dH450hbbS2oUdf3pBXxBTikjiXnG+3+KrtqhskVcW2sodVxvKHG8o/GWrdR99V1AUl4/aJPc4g/3xVXVHev3OcPcUn+8KrKAV1cWEJKlKCUgZJPIV2q3OkeQY+lnkpVgvLS37Ccn5hQHE7WtiiudX17j5B3LKOID2/oqfivokR232lBTbiQpJ7wawSeVZWsUKQuywpduuLjS1sIy2557SiBjlzHLsNDBctKhhd3YZ4btCXHT/KGsuMn1nmn2ipRiQzJZDrDqHW1clIUCD7RQyR2olKeQzbGlELmK4FEc0tjdZ923tqUabQ2hKEJCUpAAA5AVE2weV3ibP5oaPkrPqG6z7VbeypfNAUN9uUSz2iZdbg6GosRlTzyj2JSMmvnn0gammax1hcdRTiQ5LdJQgn9jbGyED1Jx89bK/DO1gbfpqFo+I7h+5q6+UAdwwg7A/jK/w1qaK9d/D+k2Vu+Xd9vgUfU7901WvIUpSvRFWKUpQClKUBK6Us5vl2EEXCLAT1alrekL4UhI5jx9VZPtem2dL6cm3zTuonJ0q3qC5bSFAR3lI3KDjfGD31ieyPxotzZlS2+taZJWW+HPGQNk795xvXrcL1OliQ0haYcR9zrFxI2UM8WAM8Ps7ar9Vp7b54UsR4z7yVTZCuOWuS/NR6zTqfo3nqu8GOxLTKbRDLefOPNRGd9hsfWKx9LfmohMW91KmWAOuSjl1nHuFnv2xjwrzgNGXNjxVrVwKWEjfkCd8VWaiYW09HUt11w9SlA6zHmhKU4SPVnH0VvTRXp34ce2cmtlkrFuZF0xSlTSOCAaUpQE5pqNMuEC52y3whKkvJaUgJQVODhXvwkbDnuTjasm6J6MYrmmXGNT28MznH+NC2nfviGwBhORkDJztVt/B+c4dZyW8+nBX8yk1mUTpCdTG2LS0WVRPKG1JB4gQsJIPZ215fq+surtdVfHnnzLjRUQlBTlz5Fl33o9tF5u6beHXoaIUJgNrbAUpScrTg558gc11Y6HtOIA62bcncc/PSn6BV7JTw6pcV/GQE/3XD+mqbUylMTbLJQ66jNwbZWEuEJKFJVkFPI74qrjr9TxCM2uCW9PTzJxMa9J3R9ZLBpJdztaZPXNPIC1OvcXmE4O2O8isTg1txKjMSo6mJTLb7KvSQ4kKSfWDWGvhA26LDkWZ2JGZYQpt1BDTYQCQQez11bdH6pKclRZy35kLXaRRXiR4XoYtqstdsn3N1SIUdTgQMuLJCUNjvUo7JHrqkbQXHUNpOCtQSD4k4qe1nLWzcHrBEUWbdb3CylpOwcWnZTi/lKJzueQ5V6Gc5blGPcrIpYyz0jpsFoaHlTzN0l8RK0MNcbYHYkLVgd+SAfCredUlby1obDaVKJSgckgnkPVXXspWYV7XnOWYlLPApSldDUUpSgFDQ1J2m3B5KZctCzHKuBptGy5C/kp7gO1XIDxrWU1BZZmMW3gyJ8GrVUrSOskPPyUR7VdSITvWAkKcz5i0gc+Enc8sGtwXFFvDbaVLdUriSlW5KjyWvx7k9nPsrRBtC5S+tWQGv2IFk8IUBuWWT8VA5qc/Tvt90E6na1doSLLZd6yeyTElOgEBBRgcYzvlQxgn6q8j1yjMlevg/0Lvp9mF4bLuetflsR+Lx+cd3HjuS4N0geAOCT21AvXKPd5aZM1D5VFSlLbCGwoKc3ClHO2ARz9VXHqOWi3W1uIw4hlyQQ02VHZIPNR9VWY01GZuwjMS1GIoFTjzieAugDzx4DG49ea8+WbPPyWU7b1rROBeUU8LYT5/VDOVZ7AOe3ZirhsNqtk1xqM2yt2PAJLy3D+yPHkNjggc8+qvFpCZLH2yWp12I0owo8dIwHmz6Oe3BJT7BV12CAm2WxqIndSRlw/KUeZoZK5IwAK5pSgKO9fuY/+L9YqsqkvH7mSPyZqrHIUAq0deIj3NbFnFwbjygeuSlxJ4V9gBI5GruNY91Gy4dZvyjwdWhpsDCwSDlI5cxzoGWVKYdiyXY7yeFxpRSsZ5EVkDouuodguWpw4Wxlbfikncew/TVnarH/aa4j/AOpV9NXDdJv2gujUcS5LbaEpUhiO2kICMY84nck7k0MIyN5vLbeoC+22HDiyLnDU7BkIQVBUc8IWrsBTyOT4VZrWopbTxUxqSSU52TIi8Qx44zUqNTCf5JEmOw3EeUJW44yspylO+CleMZOO2hkl7e9cbDbwi5MMOQ2WytUhpeFA8yFJPMknmKkrTfLZdG+KHLQpWMlCjwrT6waqGZcSW35qkrSobhQ2P1Vjvpgi2LSmkZ2qwlxkxgnDLKhh1SlBISM8jvWYxc5KK7s1lJRWWamdOOqFau6ULzdULK4yHjGi77dU35oI9ZyfbVlCswHT3R/rdKnbDN+1VwVuWMcOT4tnn/RNWVqjo+1JYeJ1yJ5ZFT/DxsqAHiOY91e80msojFUv2WuMPg85fRY2590/NFqUpSrMiClKUApSnbQEtY47/ksuQ2qOA4y4wQ72A8OVDx3FRJ5kfPV56VYK7NESy0rjW8suFGSpQ42wDgcgMHBzucjsqzV+mrfO5376j02bpyOs44iiU0iVJvzDwKU9ShxwkpCsAIO4B5nwqu1rEVFTbwoIHEhfCErCiAOEAHAHv7a56PmwbjOkFp5wsQHlJ6r0kqI4Qr1DO/hmvbpDUpS4CllC1nr+JbQw2SHMebtnbG+a4Sk3qkjolinJatKU7ankYZqptEF+6XSNbo5SHZLobQVHCQT31emmLBpZu1wZ+oXkuty0ErDM0JcYVk44m+ZSRjccs7ipqNC0ezfW1WRwxktutOIkttrfeaT1RUeEEEZKsA5Bqtt6jGOYxT48/IlQ0reG2V+gdHO6T1nBU/cWZT8hp9pxDKSA2QhKgMnc5FX7NbU3rC2ycHhciSGD6wULH0GrEcl2t+5KjvxLvdEi7NcK3mHOtUFR9xnzcEkbDbYVcsZUyK80u1WK/wDVpVuxKcb6vB2PDxrJScdxrzeq8S2anN8te4tqdsI7Y9sk1ZpsS8Bm8QnAtotLZI7UqCxkHxBBrx1qgmyoeTzjzI73sS6nPzE1Z+lp5skxS/JnI0dttCLuyrfqllSgiQnGxSRgKI7N+ysjSWGpcZbDo42nBg78+2oV1fgWp+R3rl4kMeZ7HmfXWLPhEscVjtUjH7HKUj85H/Ssp+2rA6eWOt0IXe1mW0r35H11t0uW3VwfvNdXHNEvgYCSotrS4OaCFD2b1Na7QE6tnuJ9B9aZCT3hxIX9dQmAam9Unro9lncy/bkIUfwm1KbPzAV7ufFkX8Ueej9xohKUpXY5ilKUApSrv0bpRFwjG6XCazFYZ++lp5sq6xsDdWARsTgDJGfVXK66NMd0jeFbm8IjrDp9ySw1cJaFKjuK4WI7asOyVdiR8lO26jyFS8yGoTCmdJisMIAbfcSrzeHsYaCdwjsKuatzv23azOsrsAzG0OSmUjhS5MVwIUQOSW0YGBjtOBirV1HrKTdmPtHaIcZqM4pO7TQRxKHanuHPdWe/aqqN119nC4/ImOuuuPfko75Liy1Khw3mupbQBIksAhptgY4Wm0kA8+fylHwrJ/wRtZiDr+Vp59YYt9zjHydsq81txrKhv3lPFk9prB891pDaYMVQUy2crcH8Kv5XqHIe/tq6OiyCtu5r1G6kCPbCFhS3OrTxHbJUeQAJ/wCtddZRD7LKMvw+PkaUWS8ZNG310lxpzz98lvIWyOEQ2Q6AcJVnzk89yK9L2sybc1cWkpVMfUXzg7Nx8cJBPYDn/wDMVAQeqmwI/WyVog9WXNgcITjZWO0nI37M1WlqUIrkmMlbTM54R0NqSGkPp5JPPI7zjnXhHw8Ho85Lp0/F8svC5a1lxqJhCN0lHW4x5nDtwhOB35q7MCobRjSGtOxmQkIW3xIdH4YUQr56maGRSlKAprt+5kr8ir6KqEboSfCvC5/ubJH+xX9Br1jnLDZ70j6KA71j3UDaEXy5SseeoIBPh1raQPmNZCrHVwlMuS761L40tNzGwhxtIKhlYPDg9mU5oC3tYAjVFxAG5fOPmqbekpn6n+0s2I1LZU6G0qJKXGfNHFwqHZnOxrw1RIt8LU8ySll6RMDgUlLgAaQrAwe9Xf2V06O0LlauTIdJWpCHHVKPMk7Z+ehglbr0fEZXbJue5t8f8w/RVo3G03CA86zJirBaxxlI4kjPLcVm1VRViAdk3OQRnrJZQPEISE/poMGHY77zCuJh5xo96FEfRVLrJp3VunzY71MkvQ+tS7hKgDxJzjfG/PkazRddMWa4EqdhIbWfjteYr5q186VdZ2fQuvndNPR5khpDDbqn0FJKCvJ4SnbO2N6kaaq2yzFK9pcnK2UIx9vsyx7p0OIC+stN3UhQOUh5OCPUpP6KqbWekvTGG5DCb5CRtgr4lgeCxv7xV32XWul7vwiJeI4cP8G8erX7lVcKSCniBBB5EcjU2zXalLZqI5+K/Ujx09T9qt4+BjG6saD1SCbrDdslyV6Sy31agfEgcKvbVmX7oxu8ZtUqxyGL3E5gx1DrAPFOd/Yaz3LixZSOGTHaeB+WgH6ail6YtYc62KhyG5z4mVkV103VZU/dbx6Plf3NbdGp91z69jWF9p1h5TL7S2nEnCkLSUkHxBrpWfukbRT98siltOJkXCMCthxWy1jtQe/PZvzrASkqQsoWkpUk4IOxBHMV6jQ66Grhld13KfUaeVEsM4pSuFcjU4jmStNMtMWu0gJ61RjpceWriAYBcUoBYB3SQMjFY2Vuo+s1mGLGabNp8rb4GI8NoMqYQlSlHq3FZWkZ34vZ2nFYeVniPrNVuglunNkvURwl+/QvTo1bKbdfZSpKorfUtsqewSnzlegcd/f4bb06Um1NizZgpioUw8pBCsh0F0niwdx6jv31K9GSFR9KSpLUUzpD85ttuKUjC8FOCFEgA5yM71J6ksN1l3nT702yNpjtqcR5I9ISEHzVOFIxkgYBJPuAqJK+MdW5N/tI7KtuhJfvkxzqSIxEZtCWW0pW7bm3niPjLUpW58cYqJrIF7hWxq2XW3zLVHj3RqLFciLj8Sk+iCoFWBhRCskHbuqxG4z7jRdQjKN9+IdnhVlpr1OHPkRLq9sjYvRkSK30b290RWA4q28SldUniJ4DuTjNdej5R8okAEjNugK/9Ij6qrNMJx0cwEjstY/9s1RaGT1c5AO3WWSCr2jjB+qvG2SbVuX5/qX0Y4cPgU0xR+6x7f8A+PxP/wDKavlPMeurImIP3VuHvv8AEP8A/VNXt3Vw1faPwOlPmWw3bPtnaUSWFNtz2HH2kKWMocR1igWnB2oI93MVB2KfqMXNvTkN+HCYZQrgVLbLj6AObWMgK4RjBz5ycHvq69JnMScj5FylJ/8AUJ+uvPVNjTcEInRHkxbjFIWy+dhtySrw5+rJ7zXSFyjJ1zWV5e40cG4qUTs1D1KhQUq+wXgPiqt3CD7Qvao7pYjmR0d3ZJAKkNJc270qB/TVaNTwShDTKXZ83hHWMQEF4JVjccfogZ7zVHffunu1nmQRYoTTUthbQbdnffhkc9klPszWKVONsZvCw/gJuLg0ucmttTExXXaPtq+2PLfZPqUErA/xVXz+j/WEHPW2R91I5qYIcHzHNRshmVDsTsKZFkR3RMQ6lLrSk7cCgTuPVXt3dXZtcJJ8lAq5xypIihSuAc8jmualHEUNdm0LdcS20lS1qOEpSMkmry0vplpuE5e7s+mNAYP3yQTtxfIa+UvvUNh2ZO443XxqWZG9dbm8IotL6cMgqmzyhmOyONxTuyGx3q8fwfVnuM+8tm4wy8/1kTTzZyAs8Ls9Q7VdyB3fXiu7ympsJNwujJt9gjnMSBnCnyOS3P8A87e0neztTX6ReZO/3uMjZtpIwABy2+r66roqepnl/wDX+SU3GqODtqS+uXNzqWQGYiBwoQkYHCOQx2Dw9p8KNz9YR1MDaW8nDx7WkH4nrPb4bd9dYqUxGEznEguK/ayD2kfHI7h2d59VUaiVKKlEqUTkkncmrGEEltj2RGlJvl9yqtEF+6XSNbo2OukOBtJPIZ7T4CskSrApuU/pATYabWwynyjzyHet5h0DGM5OAkncZqC0DDZt9kumpp75itpb8liuhvjX1i+ZQPlYzg9nOskWtmbe2osl+OqMwxCStlDuFyGFAFIKjsCSAFYO44TyzVR1DVNT47L8ybpaVt57v8jLOh2YLGm7bbjKLrMceTynW91FkYSlG3apXPFXRAiLv0qPImKcXFZkcERteM8CN1KVjYnICaxN0PzIl8t79hsSA2qK822kre4nMEEqWeHzdyCo7kjPPes+2CIiMyhCMFthsR2yNs8PpK9qvoryN8HCxqXcu6pKUU0LP95uV0idgeS8n1LTv84NS1RK8NaqT2CRDI9ZQv8AQo1LVyOgpSlAeFx/aEj8kr6DXME5hsHvbT9Arrc1BNukqUcANLOfYa62lfWWqIsfGYQf7ooCqrFF4K0x768pKk8dyQASMZAKzWV6xX0iyXnLn1CnVqaSpZCCdhg45UBQ63/1nlH5QQfegVO9E7OZs6QR6LaUA+s5+qoHWR4r4XPlx2Vf+mKvDopY4bPKfxu4/j80D9NDBeR7KitKedZ23e11xx0/0lqNd9R3Zi02tyU8RxYKW0Z3WrGwqwNHaolwprESW8kwlrwrjG7ee492aGTKR5VoP8IyaZ/TXqVwqyGpIYHgEISn6Qa33CgQCORr52dKUgyukvUsknJcukg/3yPqr0H8OxzfKXov1Kvqj/20veW0QDzqZsOobnaHB5LcZ7Cf9i/jH9FWUmoeletnXGxYkslLGTi8ozFpbpIujq0tOmNeB2tpAjy/Ykngc9hzWQ7FqG3XyM6q2vgyGgQ5HeSUOtK7lpO437a1a7vfVy2fVL7bzJuTj5dZ2YnsHEljwz/CI/BV7CKotZ0WE/ar4ZY0dQkuJmfdMS7pIS8i6tLbcSEEBbaUKyQeIAAnKQdgrtrGHThpDyWQdS29rDDysTEJHoLPJfqPb4+up7TGsDIukRVweY4nVhtD7Rw1JQvYkA+ioKCCUHcZPZWRZ0ZmZCeiSWkOsuoKFtr5KBHI1TQts6fqVNrh90T5QhqamsmpIrvGbU9KZZTjiccSgZ8TipbWFjcsN4cjYWY6lKLKlc8A4KT+EnkfYe2qfTDSn9R25pJSCZCCOLlsc+PdXs/FjKrxI9sFDscZ7WZhcktRL88mBEkyrsGB1wQpAbKQkoJJOE8OTsU7+3NY4GgNQqJKkRW8nkp0n6Emso2qI7N1PNENT8fqoLaFIccAdQVrUSFkpO/m8hyGMbVVP6UubqifKEj8aYr/AJWxXmatb9meE0s+pbS0/irLXYgNLMXWy2a3W8xIrhhPqfKi6vhWo57AjPaO3squud5vM2925K0WpktIfWnzlqAykIOeIjfCjXherK5aXo6ZUiGlTqhhRlSAEDOOIkchkjsrvbTdXpSxaZ65TzLe4i3BpwpQT8l1sbZHea0k4zfi8M2ScVtPC+m6RYMovSYbzFxOXUIjlJSWmDw4VxnbzBXtovQ9203bZUqLd4C/K4wUsOwlKKQEk4B4hjnVTZLei73W7u6stsma7HWy00HWRxITwE7pbPDvnmOYqrtNmtsjUV3i28z7dHREYSUNOLa85XHuQee2K0nqHCDrTx2b44ZlVpvc18Cd0wOPQ9uR8u3oSB3koxXGn4DsaSwtaCkN2tiOfxgSSPZUZcYNztdts8IXVpTLc2NHQExAkgA7Eni35VMOW67q3Go3mx3JhtfXmq2eEm9yxJv1JUeWuOxSPROtvSn28KCbqytWPilLJSc+8e+rhztVoWC23J+TeP8AtJPbKLitCihpocRCEb+jt7O6pQ2GYr09UXw/iuNp/wCSl0Y5Scu3uZmtvGUiDtF1kDUF7skWTFjPJuSlgrQpxwpWEkqSnYYHaSfZVXeI1rYbbeuEtV6eLnD1ciQCgeaTs2nCRy7QakEaVtZiPMS/KJq3Xi+ZD7n34LIAylaQCNgOVQGr7fLtNmbT5U1KjCQOBbjIS+jzF7FSRhY8SAfXUiM67bEovH6nJxlCPtInVQ7q7FBXc4tmghIPVwWgCB4uL2HsTUKWNIOrUGIdyvrwP3yUyXX1J/8AMyAPUk+yvV1yyztSwxc5EN+OLUhbbbz6S2l0L3PDnGcEc+6qyfqy3wLuLal62NRUMocDq5BCSCSCEhCSNseHOsJTTwk/Xjj/ALDcccnhanbSsoFq1VcIpUMpYlucQI/FeGceo14dIGq5Ol7Yyp6LFui1upQrjHAkpKSQcb75SakdATIV10sw11jMoR1racSU8QThauHIPenBrGPSGMaXMbAC4khTSvHgfcSD7lprvpaY26nZPyZzum4VZidJGu9I3DP2z0FGUo81srSk/MAax+2yqVKUiK3gEqUATshOeZPYAO2q3T9gud8ecRb4y3ENJK3VgZCEjmccyfAb1fFl0YlksG5ueSxMha4hSS+8cjBcOwHgncDtzXpHPT6LKi+fTOSrUbdRjK+hEWC12+3283a6FXkfIYylctXyEDmEd/arwFSkhxx9CLxqTDEOMMxLeNkNd3mjbi5bdlXBNssL7Yfb+5P3BbDb6Y8KMYwSlpJ7UjtI7+ZrFuqrtIulzd6wlLDTikNNg7AAkZPeT31Gof2ueU/i/T3I62LwInGpb7KvczrHSUMo2aaHJI7KpIEdtYXIk8Qis448bFZPJA8T8wya84UZyXISy2QnIJUtWyUJHNR8AK9bjIbdKI8UKTEZyGweaieaz4n5hgVaqKilCJCbbe5njLkOSn1POYBOwSnYJA5JHgKqbBa5N6vEW1xBl2Q4Eg9iR2qPgBk1Q1kLo1jOWeENQPNsNh91LXlEk4bZYB++K8VKHmgDf3VpqbfBqbj38jamHiTw+xesqzwG7KbfZ7m2t2ChLIJYLqI7yfPU6EjbjKVbKJ2A2rpqSO1abaxZ25ElpC4pcUpLn64kPJwsJR3qPCSonYDbma9oCX7W/cLVBt4DQS060tzZc1aUkIScbYUUpxjAwD3mrdd1Wz1UrUU+BBVfLcepMptJKFOKSoIabySMp3KlDsG3fXmK4WTl6r9X2LeUoxXozKHwfb7Gb13P03Agx4jhtofW22kYakLWkrTnt4UncntzWxkdpDTKG0bJQMCtP/goyXIvS4xEdWVSp8GQ9KUr0skJUlJ8cZUfxvCtw07CoHVKVVfheiJOjnvryRd28y92l3vW6170Z/5alaitQea9bF905A96VD66lariWKUpQERrEzBp2X5EhK3CghQPyPjfNmu2lBJRp+EiUkJWGhgA5wn4ufHGKkJjS3ozjTbgbUpJAUU5x7O2vG1RVw4LUdx3rVIB87GBz5Adw5CgKusQ66VxXpfqWf76qy9WJNRRXZt9XwEJQlPE44v0UAqUd/fsOZoYZSaoUFTIq8+lCYP93H1Veml7jDsWiYsmYsp61S1JQkZUsknYD2VB3ly3Q7ezHeA63yRKAlbOXFgZAyeSMH299Rt9alSFQIjDbjqY8VpvhSM4WpPEduzOedDBSX66ybvcHJD619WVHqmydm09gFeUeCtyDIluAoZbR5qtvOVkDh+n3V3nxUQojKH2HUy3U8Z4zgIHERjHPO3PxqgoZMtaDuibjYWkLID0fDSxnc45H3VoBrYlWs74Vc/tlJ/91VbZ2lLvG+6wVpeZb61soOFcSVDs7e2tb+lvSNxs96lXjzpMKdIceLqUEBpSlE8KvE5OKv8A+H7YQulGT5a4KzqcJSgml2LFpQcqV7EoxSlKAunRjkWZGNvnpJbYd65tQOFIB2UQe8YBH/WswdHOrVXgyLLdVJReYKih3sD6UnHGPrHtrDHR1Ck3HUyYURyOl11lYSHlFKVbZxkA77Z9lS2rYF5skm36ujlMcqd6lC0ryouN5TxKHZkJwU78jVFr9NXqLHU3hvt8f8ljprZ1xU128zJHSrpti5W9co4QlWA45/FL5Id9XxVeBB+LWJ9AwHmtex4cpKWHmFrCy6SkNKAICiR3Egj2b1mTT+t7TfbEysxpEmU8hSZMGOwXVJxsrIHxTnYnvqyUORYuqg+u3SSywowpRlLQz1scgFIVlWSpGySB6SQM71C0Nt1VU6JrsiRfXXOcbIsu3TFvduN3vchd2uLbqHWWS/HeCev4W88eCDt52w5CpmypmxdSzLc7c5U6OiI06kyOEqQtS1DYgDbA7atWKq3wb4/dbbqO0NuF11CUvlbvGyQ2EDKSPR4TgHsNS0DUtvZmy5b8yE/KebaQoMrUE4SFY7CRnNV18JTzjlNIk1uKxngmVK49dBB3DdrJx63f+ldUADXq8ADFqHId71WZd9bSLffH743Z25DCIQZWlEvzgA5njwUg4yoDbOKi4/S1C+3zlzfsslIVFTH4UPJJGFlWdx410h0/UTjmMcrH1NZamqLw35l/R7vBt9/vDclTxdekJUlLbCnDwoaQkk8IONzVDaL6y5rK++SQp8xSmYpw2zwlICVDcLKTzNQVh1lGGo5t4esl+baksBLQRCUvJKsk5G2MAVIabv8Ab1avu13mJl29maGmWTNjLazwgcyRgb55mk9LKtS3R8l8+BG1Saw/MrtZ3aaqLbyzYLmHEXOOtAdShKVkE4TkKO5rhzUmswfN0Ksjxlf9Kua6Qk3FuJh3hDMpqSCBni4DnHtzVbgd1Q1qK1BLYn39Tv4cnJvcWRo+9XFIuy5em7mJDlxcW42wErS2eFPm5Khk+ztqd+30n+bN8/qm/wDPUjBgsw1yltFRMp9T6+LsUQBt4bCqoAd1aW3Vzm2o/mbQhJJLJCfb6Vy+5i+/1Tf+eoXWtxauuiGJ7DbraFy0p4HQApJBUkg48avYYCgTgb1jm4lJ6NGjxJ/dI9v+3UK7aTZKaklhpo53NqLTfkQclbLVwguKcSgeQpB+8Qu0II2Ud/Wrero6Pg05frk6hxLqFRGB+xsDfiX/ABXm/XVNb/t8uyW5TEVTqDFa4T5BGXtwDtU4CfbXoy9q1lwragyUKIAJbt8ROQOzZ6p9uZwcVx5dyPBKMkyv1OJNvTf5EJ1cVx6NFW242AClQWUEjs5YqwekKLJYtl1bmPoedTLkcTqUcAWQWFZ4ezbPz1ec9zU862vRJFtnO9ZwjzmY7fJQPMOnurtrlUlhITEhhhUuU7xOvNtuBZU0DkA5x6HLwrXSzlVOKeG/7YM3RU4t+RB9EzMhizGPHgxZXVAOSArzXfPGSEK79gMHA2qOv/SZd7VMejwdNwYXVKKVKcBUseBxjFVc6+OwOjpy/wASK3GkSbk0FKZHCFpSriB4ezI2OOdQOroVvf1U9eIszqmLotKGh1aiQ8rHEBjGcZB7gTg91TKqoW3SnbHKefmjhOUoQUYMyHNlzb9pe3ti4C33aesSIzaV8PElBHGgH1E+Na+yUqVPeQlJUtTygABkk8R2rJ3TJcXbLqjTqIiz1lsYDyfE8f1hJFWrq+PGs95kS4jqFquA8ph8Jz1LTg4uI/hbkDuwT3VL6XHwo5XafK+Zw1j3vny7kLMWIcdVvZUFOKI8qcB5kckA9w7e8+oVQCgoauorCIDeSQ07a37zeY9vYSSXFAKx2DNZi1Jpm5tyYTDqbe5p2E0nGYoceSr4wCTtk8+I7AZqB6MYLWn7E/qmcl0LQpJQ0hsKcdTy4Eg9vnDJ7M1kfU9wT9rksOERitsPSOtUMNJyMBZz2qIB8AqvM9R1k5ahRh2XH4ltpaIqpuXct2826/XK3vqjRJccZbkNtMuBBkgbJYKviJSnhOe08W1Y81i8wL4bahllq2WQFTjTQAQt5WCU7c/OwnJ3wkmrjlamull09PnPXCRImrBtyXHioFx9JypxCDslCUnnzJxWNbmoxmG7fxEuBXWyVZyS4RsPHhB95NSunaee5uXbyOWqsjjCMmfBUfdf6dre+6oqcdYlKWe8lsmt3RyrST4IrRc6aoawP2OFIUfzQPrrdscqqOv/AP1L4Im9M/8AT+JFaj9GB/v7P01K1Fai3NuT3zmvmyfqqVqkLEUpSgFKUoChvrk1q1yHLc0HZSUZbQe01ji8uXiFGQ/LAEo7kBISIwPI8I24j8rs5DesqkA1F6htSLlDKEpSH0pIQVDY55pV+Cf+vZQGNC3ATZIM2ezKlFfG0A26EJThRPnEgnJzmuki5Wp99byrbLQpYAUUTMZGMY9HltVd5FGRYnYUh2REKX/KHQWS5wJ3SkHB27d+3ao3yKzf+OK/sa/00NTtGZsk6ShkyZkEr2St4pcQO7JGCKqINlmssplMyI5bWSCFtrKVAEggkJIwcd9UaoVo7L2T/wDxF0huNRuLqb5Jj4UQnq21jiHfsdqAkokTEpCY8GB1jR6xC0SljgwMkrJ5gc+ysDdPuppcYPabaZQESF8TskKBDyAc5QnmEk9pxy2rO0RVxvUhNsYvc2QHfTC0KCQntJ35Vrr8KexpsfSoppri6l+Aw42VHOcApUfek1b9ErhZqlu8uUQtfKUaeDFVKUr3J54UpSgKuyXB603iJcmDhyM6lwY7cHce0ZFZX6VbXGuFkXdIAUW5LKZ7OFkpyD99AGcAlKwr+iaw7WXOj+5m59Gc2GoJdlWQqdbQo+m0UnKfUQVp91VPUoOEoXx8nh/iTdK1JSrfmTPQlbm7ToeRenUAOSit4qI/g2wcD1bE1WaZthWWY8MoiSVxkTLjNDSXHluO5UlCSrISMZJ9lSElgWzotERocJTbUMpHisAfSqpDS7SUyrwtI2E0Mp/FabQgD35rzV18puy31ZawqUdsPRHYWWak5Rqa7BX4QZUPdwVGaHlSpV5vXljwdfZLTC1hATxcCnQDgbDbFXZVlaLkR42odUOSH2mUeVDznFhI9JzvrhVOVlU89+PzOk1GM4nXpcjh+1MpV8ZqS3n1t8X/AC1b6ehi3qQFC/yyCMj7wj9NTeuLrDurEdiD1r7SRJKpCUENZ8mc80KPpH1bbVPWu1yxAhvRr9ckhTLaih0odTgpBx5ycj31KhqbtPp4qMtpylVXbY21ksz9TmQ3KTb4+uLqh0M9YlvCsJQDw9isDfbFXDpHRK7Q8qRdb3MvDoP3pLy1BtHjwknJ8TV4BIyVYGSME43rnaotnUb7I7W/yOsNLXF5SApSlQSQKUpQycGrL1To20u3WJfeBLLEVRXMjJb4m5CfFOQARnOavWuFAFJBAIOxB7a7U3TqlmLNJwU1hlju2XQc1r9bW+M87gYbbfUyfUDkAeqrA6QrLaIamWrZFu1rlBXnh1SnmnEnlhSVHBFXvruwQrVFcvjTs1EVohT8dhAXsTjiGSMAVUy4TM3RkduHb3r9EdSFsupdbDzYVzKeIYBHLAq4016q22KTaf7+BBtr35i1hmDURZpWQm48AB3UVuJx48qvx+32u7SWri7rK2sTupw+0p1fVKc4OAOAEDhJ5kY58qpLzJ0jFYlteS6hbkxXgytp5baSCc7jbw+eoFi46d4eFb19QM9vVrA+cGruW65boprHuRXrFfDefxMiahjQXNOwLNb50SZHgpMlxxL6eBbmcJTgEq2znl8/KMh+W3ayW6OqJBuEy2Tes6xDikAI3PAEcI4vNz2jJA7atS7M6YYmONM6gmO+bgqFtSQDjlniB29tdrLbLNJkNx4GpGkyXD5vlEd1vG2TuNhj11wjplCvu/XlM6O3dLGPqSOv4F2kagdu1zhTZDTDTbaAqOQXF44uA8ORhOdyD2Y7ah9Rqk3TSlsvElKy/FdcgyFKRw7em37MFQHqq6YOi9U9UiZaLrFltlWUPRpjnArB7MGpG7sarFuWi/peShPoSUzeDflhYUAlQ37d/GkNTCGyMWnj99hKly3Np8mHs1M6Lszl8vzERIHAFBThPLHjVwxNBXSSyqVboq5TfHsoFp1J7wcKA91XVoW0zbOpbMnRsiPJddbPlDqitnIUN8JyU4G4zkZqXqdfBVvY+ThTppOS3diWYiKsFqevE6Yp19t9Qt7KmlJS2XUhKULTv2pCjjtFQF1ttzuNrXZEPSA8ZDL0+QUEqkKXnCE/i8SVHsHEc8qkmNQzbrrNTEOQftcw91AL2FhTbZKnnSDzJOwPZtURqiRdLc/LuK74ufcrupcKAy2FIbjtEjjVg9wwn15PZVNTGzet2Nz5J85R28di3tZXdE+7LlB5yRb7QgQoBdVxF50DdZPby4j6kjtqyyVKJUolSicknmTVXdX21rbiRlZixQUNn5ZPpL9p+YCqOvS6epVwSKm2e+WTOfwLIhe6TbjLxtGtS9/FbiR9VbiJ5VrJ8Bu2/etT3dSdipiKg+oKWr6U1s4OVeL61PfrJe7H5F/0+OKERV6866WhrvkKWfUltX6RUrUTK++6nho7GYzrntUUpH11LVVE0UpSgFKUoBQjNKUBZWtLUuKu4XuNxOl6MWXm/kggDiHgMcvbWOBWenUJcbU2tIUlQwQeRFWBftBOdap20PI4Sc9S4ccPqPd66GCxa9YUWRNlIjRWlOOuHCUiroh6Buzqx5S/Gjp7cKKz7hVyW+yRNO3G3OR+NXXKWw84s7qKhlPq3Tj20BX6TsLNlgcAIXJXgvOY5nuHgK16+HDZyHNN39CdiHYbhx6lp/5q2gByKxb8KGwG/dD11LbZU/byic0AN/MPnf3Sqp3TbvB1UJe/Hz4I+rhvpkjRmlKV9CPMClKUAqa0feVWeZLBUQzMhuxnP6STwn2HFQtctoW64GmklbijwpSNySdgK53QjODjLsbQk4yTRs5qcBdhgRhykS4bXsK0n/lr30eeO0uyf5RMkPe91X6Kt6ZOiT42n7LeIrkZ/wAoa8pZmNFKTwNKOQr0VAqA5Gu7t2a0rNchQIT9wtrraXGG477ZTHXk8aMqVsk7EDs3rwrplKHhrvnJ6JWJS3PsXuOY9dYqtd505Avl4k3SMibIW+fJ20R+uWSHHMhPYPi+8VKy9WXu4srjxI9nswcSUmTOujS1IB7QhB3NSujI2l7SxHgW+4QJk/hwt9K0rdcPMnbOB4Vmqv7NCXiJtvyX6sxKXiyWzyI2UnVGqFMrFoYslvZS4UGYoqeXxtqR6Cdhso7E1dtmcU3p6GuQ2pkoio40K5pwkbHx2qvqwel7UiLUxCtzUuM0+66l51DqlDiaSeXmg8z9Fc63LVzjVFYRvJKmLm2RrvShLkSFqtdtjORMAoW4HlKz2pVwpwCPrrqvpJvKElS7ZBSkDJJbkbD82rIbbtUyAqVIglTbRSwwtuY4E8Kd1E+aCcZAHaVKArzVEsm6ftU8k+cMLuC8JI3wTjkkbrPIHYb1frQaZcbP38yt+0W/1F7p6T7u5HL0e1wZCcZSEJf87wB4cV5udK89khEiHbWHCkEoWHwR/dqzREseQEWiSRxbJ8tWFHI80Yxso+lj4qdzXXySwkAm1urGEniTMWcjOMgYyeI7IHM7nlWy0Ol84fv5mPtF39RkHTPSbIu9/iWxUaF+uJCWgW+s3SUklQKgMYwBg881k4Vro/LiWC6pat8TydlYEmO85NWpPFwqSCQkEFSSVJ22yKzfoi9s3/TcWch9t50J6t8tk4DiRhXP3+2qjqujjWo2VRxFk3R3ubcZvknKUpVMTjqtCVoUhaQpKhggjII7jVsWC3faLU8i1QElFpkRjLQ0dwy7xhKko7knnirprp1aet63hHHw8OfDOa6wtcU4+TNZRTaZaXSHoaHqqP1rTqYdwTjD3DkOAcgsduM7HsrB2qtJX3TTn+k4ZDJVwokNnibUfX2HwNbQ1S3SBEuUB6DOZS/HeTwrQobEfp8astD1e3TYg+YkTUaKF3tLhmphJJJJyTXvBcLUjrAcEIXj80irn6Q9ETtKy1PICpFscV96fA3T3JX3Hx7atJOSdu4n2Yr2dV1d9anB5TKGdcq5bZLk2I6FuEdHdvGQPPd2z+GapOnVla9ELeGOFt5HF7SAK8Oh2RbnNFQmJDqOtZ69akLRsEhZOeIjs8DVR0pPwLpo56HFfCiuQwnKPE7H1cq8dtceo5x/y/UvMp6XHuMUQEGRpW3RowdZlLnrQl4PqSjh2KgRnGQCDmsp3qRDeYcixpxuK7qlKUKbeV94ZGAl3h7SHQN+4+FYn03cbObG7abw49HxKRLjSG2Q4EKCeFSVJ54Ixy7qvKxQVyrl9vE3SM6pLa248mU6lgq87fDWxOOI8z3Vca2pbsy4w3+JC08+MLzJew24Q7Su4KQywZSAw2pQ5N7uPLP9PIPgisf6tuxkrVcklSPKUGPAbPNqKCQXD4rOfeqrz1lOYRZpEIOqXBRJdLzzZx94UQsNJ/DWpRG3JIJrFFymOz5rkp4JSVYCUJ9FCRslI8AMCt+nUucnZI11ViilFFOKUr0iR3ZctmJHSVvPrS02kdqlHA+c1dt4WSv7m6fwRrObZ0OxJS0cLlykuyie9OeBPzJ+eswfFqJ0baG7Bpa12VkAIgxG2Nu0pSAT781LKICCVEAAZJr5tqbfFulP1bPWUw2QUSKt/wB+1Bcn+YaS1HSfUCo/4hUtUVpcFdsMtXpS3Vvn1KPm/MBUrXE6ClKUApSlAKUpQCmBSlAMCo3UbDj1peU0Mus4eb/GQeIfRipKuFeiaA84byJMZuQ2codSFp9RGa87hFZmwX4clHGw+2ppxJ7UqGCPcaodOkxxJtajvEdPV5/i1ecn6x7KlyM0947nzf1tYn9M6uutgkAhcGUtkE/GSD5p9qSDUPWxHw09ImJf7drGK195mo8klkDk6gZQT605H9GsE6Wt7N11HAt0l0NMPvpS6sq4cJ5nc8thX0LR6tXaZWv05+KPL30uFzge1m0zfLs0iRDtspcZRx14aUUfMMn2Velo6KvK45VKvL0VYGVFcBaGx/SWRWQ7pK+56yMtWOY5cJMp1EKC068lbbazyPmgYAAJx4Va8232GbcFxbtO1Fqy4tHEhEPPUNq7U4GEp9Wc1US6nddzF7V8OfqTlpK4cPllCjTXRtZm1Ivl7izHE8uoeWF58QkkVIR9Q6JFsNutEBhlg4xKkvojlJByFBW7hIIyNuyp2waO0wsZc0V5GAMgy1haifUFH56qLy5pXTr7MSNYYb9xfGWYseO2FqA+MVHZKfEmoU9VGyW1uUn8Vj6HeNLgspJIt6639U6BHhR7ixqdKJDTq2m4TnEsJVkp4gngUD3kCp+DDkPRg/8AcHZ4QPxZLyAoDvISg499Usk66uPAiHMsdpYVzZjPhbw8OLBTn1Coy66XtbHC/qvU85Dyj5jK5vXdZ6kcIJ9QFatQa25x8Mt/TCNluXP+PzJabcYtvB8pRoqFjsW9xEexKKntJTGLjbfLmV211ClqQlyE0pKCBsR5wyd6tW1WO3Etrsmi2nFJVxInXJsR0+sIA4j7hV+W9t9qG2iSWC8BlwsI4EE+A7qiapwjHC7nalSby+x6uuIabU44oJQhJUpR5ADcmtcdR3q96q1HMlW8POR3XeqZQkDARySN+/n7ayr02XtVs0g5DjuIS/OPVnzwFBv4xA5nsG3fWHdGt+TuP3kpaPkwKGC56AdUD557whOVY7Twjtq16Lp/Dqle1y+EQtfbumq/LzJiUUsoj2+OClmEkNoLYypSyTlzxWpRIQOwecapfNCc/e0pCRuBxICUnbA+MhKuQ/hF78hVPGliQFucKup6wttrcXwqeUR52/YVfGV8VOw511Rcohccw/jqzxdYlOOIgY4kDv8AioHJIyo71bquS4wQtyKsjmlSflJUlxftWFK/vOK9SBXPnFW3WKUVdnmrKlD+6tSeXY2jxNUSZ7aIrTz7Yjh04aRji8wHY4+Qnnj469zsK9IcpmY2stjYcQKXVc08yVHuPpLV27JFZcJJZwFJMqp0RV3sio7CQ5Jhjr44bBwpo7LSkdidgUjmeFSu2p3ob1HMtmpvtFdHFJZmJCWkrwOBwDzeXeNvdULBkqaktSWiXlBYUMgZcJHLHIKUnbHJtvxNRlxt5s9yVMjNR34igmRDWZHCQgnKSACFHB25dlc51q6uVM+z7fE2UnCSsXkbOilQmiL63qPTUW6IwlxaeF9A+I4NlD6/bU3XiLISrk4S7o9BGSklJClKVoZFKVxk55UB5TIzEyM5FksodZdSULQsZCge+tUrpHVAusyGQUlh5bRBHYCR9FbZmrL6R9FRtRQ1SYjDTd1QMJdwAXE8ilXZ6ieVXPR9fHSzcZ9mQddpnbFOPdED0VwmHtEQ4y33A5ORMRhSyQ2kjGUjkM4yfVXhfYyLHpFhu48SFMOthbiRxZ4FnG3icb91VmltN6htMSNCnW7ymPHQ6Gyy62laSokgjKt+Zz7KqdYWa6XrSgsyYVwD/WoWX3i0vi4c7YSod9SXbF6nO7MW8nLY/C7c4MCgZAB54xV4a/t8mTeUiIwXw0FNqCMEpOc4I5jYiqsdF+ouMfeXAM8+q/QqvHXultQK1DcLm5bFtRHXiW3XXEIChgAc1eFXv2qmy2Oya4yVyqnCD3RIvU7r7VptVuK1JbQ0FONdzgSkHPiAat+q+6R3IkWHGe6vrB1iyEOJXgKIxukkdlUFS6IqMeDhN5fIrKHwX9MnUnS3b3HG+OJawZz222U7IH5xHurF5rcb4H2kftJ0fOagktcMu9uBxGRuGEZCPeeJXtFQur6nwNNLHd8IkaGrxLl6LkziBio3Ujy27U600fvsghhv8ZZx8wyfZUnUQ/8ArzUbLI3bhILy/wAdWyR7BxGvBHpSTitIYjtstjCG0hKR4AYr0oBgUoBSlKAUpSgFKUoBSlKAUNKUBD3UeRXWJcRs0v8AW0jwCj5ivYrb21MA5rwnxm5kR2K8MtuoKT4eNUlgkuPRCzJP66jK6l/xI5K9RGDQEP0qaSY1toW6aee4UrkN8UdZ/g3k7oV7/mJrQOAp/T+qGjPiYegS+GQw4nO6VYUkg+2vpIdwa1M+GHoBVtvjWuLcziJcFBqcEjZD+PNWfBQGPWPGr7oerUZvTz7S/MrOo05Stj3RSX7UelJUC3T7PcrelVvmolGOnDa1IIKVgJwNwFZ9lUGhNb6b09p/7U3CStMpmS9x9UyVhzKyQvI55GKwrz51yAByq6/k1Tr8OUm0QPt8925I2Uteu9P3B8NsuPtNkZL0hAaQn2qOfcKgrFJjz5d4uqk6fmrlTloT5bKCVBpvzUBIII4eZ8c1gfhGc4GfVV+9FesLZYg9bL5Ebdgur6xDpZCy0rGDkEZKT4cqiX9HWnrlKnLfodq9c7JJT4MnXOSzbLG/Pj2GztS+JDMRUdbbgU8tQSnkkY3OfZXtbtM3CAtTkWZCRLc3kXB6OX5Lq+05JASM8gNgKi9VXvTT1ot021TrepiPc4776GClJCAogkp2O2e6rtOobAE8Zvdt4TvnypH6apZ+LCC2xfOfIsI7JSeX2KaJYZSZrcqdqG6zVNq4g0VJbaJ8UpAyPAmpzkN6s68a8tMWSlEa8WPqQPPcdkLUr1BKEn5zUPqHpO089Z5MKDIkTJr7ZaR5O0ptJKttlK5c+6tFo9Tc1mL+Rnx6YJ8mPukrUrd61a9KgqeV1B8mYCkoUgpB3KQe0neobVMhwOx7WpSD5EjhdKEhIU8rdfLbbZP9GqoXyHFkLWiDNL6OBPC9KQpIU3sjiAQM8J35743q21rKllTi8rUSVEnck8zXs9PSoKMcYUShts3NvPLO7jri0NoccKkNghAJ2SCcn565bZecbcdbZcW236a0pJSn1nkKu+w6chxYDeorreVQLfnCeOLh2R3paSonP42MV7X7VsWZ+tLXdFWu2J2TGahA8f4Thz5yqPVbpba1n1fkZVOFmTwWOpfFgqc4sAAZVyAr2iNTHmnxFbecbCQXurSSAkHPnY5D11eWl7dKv7zvkt/W1Fjp45El2ChDTSfFWefhVRqXV0FNuFp01c5MeO2kh10RRxzV96lZHm9wxWJatufhwjl+fu+IVOI7pMsuLcnWYoYSELTnHnbjgJypOB2KPM8yBjlU/c203bTynEK6x+ES82o44lNq9NJ35j0wkeimrakuypknL5W68cJxw+d6sAV6QJ0iDMYeSSeoWT1SuRB2UkjxGQa7TqziUe6NIzxw+xkHoL1DFt93NkdceAuB83iKerDo5Y7dxt7KziK1jF7tKZCn2o1wbWQgJKXmcoCCCnB4NiMDfme2r5R0yXF17q42n4y88uskkKPr2ArzvU+mW6i3xKo9+5Z6TVwrhtmzMlKxcx0gazfR1jGiEuo+Uh5RFcudIGtGxlzQ6UjxfVVX/KtR7vmiZ9sq78/IyhSsWM9JGqlPtoc0clCFKAUpLi1EDPPHbVfP13d2yhtvTz6+N5DZC2XEHhJ3UCCRt41rLpeoi8NL5oytXU/MyJmvFD4U+Wi08kgZypO3vrznInqBTDVHSDtlwKyPVivHgufXqcxCOU8OOJeOdQlHjk7tkh2VBl2zLnqK7U75QFjiWqAo7nkScfPUhi49WUgREnGAUlRx7xXu0lQKlrQhK1YyUk745VmLUe4fJZmvtR6jslyjm22yMu3pTh52SsJS4s8gFZ2x89Yh1JD1Nfbm7criht1bhyAH08CB2BIJ2ArLvS/PfjaKlFKG0OpdZIJSlxPCV4+MMZ9lYmgKXIQ8l16M0rP3rKoyU457kj6K9N0lKNPiRik+2Sp1rzPa2y250RyE/wBS6WVKwDlpxK0+8Hn4V4VP3Z5cq3TUPORl+RvtFtTaGgVcYIUOJsAKGwqANeiqm5Lkq5rDLk6MtKSda63tuno4UEyHMyFj+DZTutXu29ZFfQq2xI8C3x4URpLUeO2lppA5JSkYA9wrCfwR+j9Wn9LK1VcmOG43hILKVDdqMN0+oqPnerhrO2BXiutaz7Rftj92PH4+Zf6DT+HXufdnlJebjx1vOqCW20lSj3ACqDTjLgiKmvjD8xZfWDzSD6KfYnFeV5JmzI9oScoX9+k47G0nZP8ASOB6gamE+iKqCec0pSgFKUoBSlKAUpSgFKUoBSlKAVC3PNuuLd1SD1CwGZeOxOfNX7CcHwNTVdH2kPMracQFoWClSTyIPOgOyd0iorVtit+ptPzbFdWQ7DmMltwdozyUO4g4I8RXFmccivqs8lRUtgcUdZ/hGuz2p5H2GpY71lNxeUYaTWGfOjpD0nctE6um6fuaSVsKy07jCXmj6Kx6x7jkVb9bzfCF6MWtf6W62GhCL7ASpcJw7dYOZZUe49ncfbWjsqO/ElOxZTK2X2VlDjaxhSFA4II7CDXvOma9aurn7y7/ANzzer0zon7n2POlKVZkQ4xvmvaJCkzHuqiRXZDmM8LSCo49QryrlK1IUFIUpKh2pODWHnyMokmWLlbXG3W7dJbfTkL62NxoI7BwlJr1dnyDHeXPbS2+cFlCrc2Er387JwCMe2o1lctxfCwuStR3w2pRJ91T2kbHeLzeQwm1KmgJ89U1TiGmh8onIPsqNa4wW6eODrDMniJHQWVXmSiMp6LFXnDSERlZcJ7AEJJPtq/LXBt2hkrbvl7iv3HHGxbwlRZST8Z0hJP9GvO4agsmkZBj6UsrEmbw8EqcUuBGe0NZJI9eas9+7wXXVuv6ZiLcWSpSlSXyVHvJKqhN2aldsQ/DJ3SjT75fiS9zvd0uElT7+tIW581CWXAhsdyRwbCpPTVovF1bcnr1ZHbtMY5lSUtkcIHMDiQATXXRtjtV2iPXi8WGJbLJHBLkkyXuJZ+SgFW/r9lUmptbwLjGRZ4tgbTZoyv1uz5QtviA+MoJ2J7fbWjbk/CpXbv24/ybJJLfN/mVuuNWF2MmJpy9RY9uYAAaQlXXST2qWSjB9VWpZJd1k35m4iI5c3GCOIGP1oCfxRjl2VDuqSt1a0IDaSSQgEkJHdk0QtaCShakEjGUqI+irCrSxrr2ojTulOWWXBqW4Xe6XRVyk2x6JMSUhpUaOWgkJ7TgZKuW+aiAzOkSkR1Mr8oWouAuJ4VEnckk9m2d/GvBL8hPovvJ9ThH11VWm7TrZcUz460LfSCAX0B0Y9Sq3VcoQxFL3Gu5SeWXpqLS+qrbpz7bP+QOtMYDzaYTYUkbeeDw+cncb7eqrTiJlz0KSqEEsK2W8xb+Mj3D6KrrvrnVN2hPwp11U5HfGHGw2lIIznGwqAakyWUKbakvtoUclKHCkH2A1w09Vyg1ZjPuOls63JbM4JtyzWthOBqF0KI9EW19J9W+KoXIFuCiDe2wocwuK6k/RVKmdOTymyf61X6a8FqWtRWtRUonJJOSa7xqku7/AH8jm5x8kTlgskW43qJCTe4QDzgSSW3OXqKQD7SKve8aaY0/EamQZFplvtPBf7U3AG/Y4dtu6sbWa5S7TcG58Ith9vPAXGwsDIxnB2zU8vpA1QsEGVEwezyJr/LUPVae+c1sa2kim2qMXuXJeE3Vmt27nKUzpRbjXXK6pRivZ4c7bg4NdmekDpBQFBzSZcz6OIbycfpqzZHSHrN9OFX6Qj8klKPoFRkrU2o5QIfvtyWDzBkKA+Y1wj03P3q4/U3erx2k/oZGh636SX3Q19yvXZPLyJ1v587VMRr1rh25W0yNIy7e0JSPKnUOqcSWzsRwknHPOfCsOs6gvzKuJq9XFJ8JK/01Vo1jqtHo6guI/wDOzWtnSs/djFfMzHWY7yf0M39Jaiu1tW3PWCQ9xucTQWEoBAGxGPSI+erNYs1i64spdgqUjHKK1zO3Io7fGrGGuNXj/wDcU/2rH6K5GudXA5F+lZznJ4c591c6emaiqGxSX7/A3s1dU5bmjvrdlu1yHbMyviSl/rSChAKTgg7pAyN9geWKuj4O/Ry5r3WSHJrKvtHbVJdmqxs4eaWh4q7fAHvqz9L2O+a51cxaoAcl3Gc7lbjhJCR8ZxZ7Egbmt9ejTR1s0NpOLYLYkFLQ4nniMKfdPpLV6/mGBWep637HQqov23+8mNJR49m9r2UXEwhDbSG20BCUgBKUjAA7AK6TpLUOK5JfVwttpKlHwr25VDPn7bXYMjeFCWFO9zjo5J8QnmfHFeOL497DHdS27NlJ4ZUtXWLSfiJ+Kj2D5yak6UoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAoLxCXKaQ4wsNy2FcbDh5BXcfAjY13tM5E6L1nCWnUKKHWid21jmD/+cqrCMjFRF1jPRpQukFsrcCcSGR/DIHd+EOzv5UBLKGRWvvwneh1V/Ze1jpiLm7NIzNjNj9toA9NI/jAPzh41nyFKZmR0SI6wtpYylQ+j117K9E1302pnprFZB8nK6mNsdsj5kYIJBBBzjBpW1Pwjeg/7Zqkau0ZFHlxy5Ot7YwH+0uNj5fent7N+eq6kqQsoWlSVJJCkkYII5give6LW16uvdHv5o83qNPKmWGcVIWiy3K7dYYDCXQ3jiKnUIAz+MRmo+qi3SvIZzUoR48hTR4kofRxoz4jtqTZu2vb3OMcZ5L+0T0f39uYZlzkv2WIlPnKZeAdfHPhRg9vfUjq6drucEQbFaLpb7axsjhc4nncfGWrOfZWPtTaju2opaZFzk8QQOFtpA4W2x+CkcqiQojkoj1Gq9aS2ySsuab9McIlO+EY7YZwXp1HSaj+D1D85q69C2jWL3WXbUtzvEaBG87yUZLz5HYEgZx85+esWWu4uwZzUrhEkNq4g06tXAo9mQCMjwqV1RrO/agltvypZjpaTwttRlKbQnvOxyT66Xaa+fsRUUn54M121x9p5b9Ce13rPVrlx4ksTbJA3THjrZ4OIDtORufoqyrlcJtykeUTpC33QkJClADbu2rykSZMkpMmS++UjCetcKserJ2ryqXp9PCmKSSOFtspttsUpSpByFKUoBSlKAUpSgFKUoBSlKAUpSgFVVpt067XOPbLbFclTJLgbZZbGVLUeyu1ltdxvV0j2u1RHZk2SsIaZaTlSj+jvPIVuj0B9EEHo/t4uVx6qXqKQjDrwGUx0nm239au31VXdQ6jDRw9ZPsiVptLK+XuKzoG6LovR3p8rkBuRfZiQZsgDIT2hpB+SPnO/dWThtzonYGqK7T0QmUkILr7h4WWU83Fd3q7SewV4S22d03Oby2ejrrjXFRj2PG8y3gtFvgkeWSAcKxkNI7Vn1dneaq7dEZhxG4zKSEIHMncntJ7yTvXhZ4KoyVyJCw7MfIU84OXgkdyR2VIVzNxSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoCGmMP2yQu4Qmy5HcPFKjp55/jEDv7x2+upOJIZkx23mHEuNrGUqHI17EZqHkxJFukrm2xsrbWeKRF5Bf4SO5XhyNATGBWD+njoLg6wD1+02GYF/wVOI9FmZ+N8lf4Xb299Znt82POjh+O5xJJwQRgpPaCOw+FVBAPOu1Gos081Ot4ZztqjbHbJHzUvVruNkuj9ru0J+FNYVwusvJwpJ+seI2NUdfQLpQ6NdM9IFu8nvEUoltpxHnMgB5n29qfwTtWn/AEq9EWq9APrflRzPtGfMuEdBKAOwLHNB9e3jXsun9Xq1WIy4l++xQ6nQzp5XKMe0oKVcEEUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUGSQB20ANXBoLRuoNb3tFqsEJT7mxddVs0wn5S1dg+c9lZH6IegLUOrFM3PUQeslmVhQC04kPj8FJ9EH5SvYK210bpax6TsrdpsNuahRUbkJHnOK+UpXNSvE1RdQ63XRmFXMvoix02glY90+EWr0OdFFj6Orbljhm3h5OJM9afOP4KB8VHhzPbWRRskUIGc1Q3O4twwltKVPyXdmWEekvx8B3k15C22ds3Oby2XkIRrjtiuDvdJ7MFkOLClrWeFppG6nFdwFU9rgvdcq4TylcxxOABullPyE/We2ltt7iXzOnrD0xQxkeg0n5KPrPM1KAVzNwNhSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUIyMUpQEXcLavygzrc4liXjzwR97eHcsfWNxXpbroiQ6Yr7ZjTEDzmFnc+KT8YeIqQqkuMCPOaCH0ZKTlC0nC0HvSRuDQFXzFeT7Db7amnkJcbWkpUhSQUqB5gg86i/KLhazwzUqmxB/3htP3xA/DSOfrHuqTiyo8plL0Z1DrauSkHIoDCHSh8HLTd+Ltw0u6mw3BRKiyE8UVw/i80etO3hWtGvOjnWGiXlJv1ndbj5wmW198YV6ljYeo4NfQw7ivJ9lp9lbLzSHG1jCkLSFJUPEHY1b6TrWo0/sy9pe/wDuQL9BXZyuGfMzNK3d1z8H7QOpFOSIsNyxzF5PWwCEoJ7y2fN92Kwlq/4NGtrWpbtilQb4wNwkK6h7H4qtj7DXotP1vS3cSe1+/wDuVdvT7odlkwfSprUOlNTaecLd7sFygEfGejqCT6lDY++oQEHkQfVVrCcZrMXkhuLj3RzSlK2MClKUApSlAKUByrhG5PIdtXTpjo81vqVSftNpi5SEK5OqZLbf5y8Cuc7YVrM2kbRhKXEUWtTNbCaO+C9fpakPaovcW2tc1MRB1zvq4jhI+es46F6HNB6QKHoFmblzUf8Ae5p65zPeM7J9gFVOo67pquIe0/d2+ZNq6dbPmXBqh0d9DGuNZqbfZtxtluXuZk5JQkjvSn0lewY8a2e6Lug/R+ii3MWybvd0b+WS0AhB/wBmjkn17nxrKSU4Fc8q85q+rajU8N4Xoi1o0VdXPdgJxTFUtxuMSC2FSHeFStkIAytZ7kpG5qh6m43UfrkrgQz/AAKVffXB+EoeiPAb+NVhMO0u5uPvKh2pCHn0nDjqv2Jn1ntP4IqptltbhlTynFPynP2V9fpK8PAeAqpiR2YrKWGGkNNpGEpSMAV60ApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAMCoyXaGlPmVCcXCknm40PNX+Mnkr6ak6UBEfbGdB825xCpsf8AeYwKk/0k+kn5xUhDlxpjIdjPtvI+UhWa9yM1HS7NBfc68IVHkfxzCurX7SOftoCRyO+upGaiQ1e4h+9PsXBsfFeHVufnDY+6u327QzgT4UuGe1Sm+NH5ycigJJxptxstuNpWgjdKhkH2Val+6NNBXtRVctJWl1Z5rTHDavenBq5os+FLGYspl78RYJqpBraE5QeYvBrKEZd0YdufwcejSYSWIVygk/yeYrHuVmrel/BZ0stRMXUd6YHYFJbX9QrYOlS49R1Ue1jOD0lL7xRrcv4Kdsz5mspwHjCQfrrlr4KdpH7LrG4KH4MNA+kmtkKV0/m2s/r/ACNfsNH9JgGD8FvRzZzLvl8kjuCm2/oTVy2n4PXRhBUFOWaTOI7ZUtagfWBgVlmlcp9Q1U+9j+ZvHS0x7RRbli0Po+xBP2o0zaoak8ltxU8X5xGfnq4QkYwBXYnFUMu7W6KeF+YwhXyePKvcN6iSlKTy3k7qKXYreVc5FRJuz8gYt9skvDscdHUo96tz7BXAh3WX+3bgmMg824icH2rO/uArBkq7hcocHHlEhKFK9FA3Wr1JG5qjL93uG0dn7XMH+FeSFOkeCOQ9vuqrg2uFCJVHYSlZ9Jw+ctXrUd6rQKAoLda4sRZeCVuyFem+6eJxXt7B4Cq8ADlSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQClKUApSlAKUpQCmKUoCil2m3SlcT8JhavlcACveN6pxZGm/wBqzZ8YdgRIKh7lZqVpQEV5BdkfsV6Uodz0dKvnGKdXqBPKTbnPxmVp+hVStKAi/wDtCOy1n+k4Pqp/2hP/AIWn2uGpSlARXU39Z86bAbH4EdRPzqp9rrk5+zXx/HcyyhH1GpWlARRsURZ/XTsuX+WkKI9wIFVkOBDibRojDP4iADVTSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpQ57KAUryeksMEB59psnlxrAz768/L4P8sj/ANamgKmlU3l8H+WR/wCtTTy+D/LI/wDWpoCppVN5fB/lkf8ArU08vg/yyP8A1qaAqaVS/bCEVcImR8+DgzVSFA4wcjvoDmlK8X5UZlQS9IabV3KWAaA9qVS+Xwv5bH/rE16NSWHs9W82s4zhKgce6gPalcJORviijgZoDmleL0lhkgPPttk/KUB9NdPL4X8sj/1qf00BU0qm8vg/yyP/AFqf008vhHlMj/1qf00BU0rwElhXovtHPLzxXqFbZyDQHaldcnHbtXagFK4UcCvF+UywE9a823kZHGsDPvoD3pVL9sIX8sjf1oru1LjOr4GpDK1dyVgmgPeldeI4ydu/NdqAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFdTni9ldq4PMeugNPfshhIvGjx/sZfb4tVqrnw+etqvsh37saP/Iy/parVSsMHOd8YPvrji8PnqqtFsuN4nCDaoT86UoKUhphJWspAyeXhU5+p9rrP+p97A7MxFCsAtrPh89M+Hz1cv6n2uf5oXr+yqp+p9rn+aF6/sqqAiNPq/0/bT5wImMkb7fsia+qFhA+00E8/wBbN7/0RXzTsWgdbt323OL0leUITLaKlKjKCQAtJNfS2yJU3aYSVjBDDYKSMEHhFbArq0Q+HZgdMrXZ/o1jPvXW99aH/Dt/fka/4Yz9K6AwCTtuTn11sh8ArB1/c+z7ynbu812tb1dtbIfAJ/fBun5FP+F2sZBu6nON65PKlKyDTn7IUVfbzSSckAx5eN/FmtWM9w+etp/shf7vaR/IS/parVYcqwDnOez56ePL31XWOz3a+TvILNbJVxlFBX1MZsrXwjmcDsGRU850adITbZUvROoEpHxlRF0BBWu9XW1vpdt852MtG6VIxsfaKyxoL4SfSTplxCJ05F9ijALc1OVJTn4pRw4PrzWIblbrhbJRiXGE/FkcO7TyOFW/biqXmBg+oA7mmQfSnob6V9M9JtpL1oc8mnMpHXwXj57ee75Q57jI2rIuQB6q+W/Rpq+bofWdv1FBecQmO6kvIbVgONZ85GPEZHtr6caduke9WSHdIqgpuUyl0YOcZAOPZnFASB3xitNPshJKdVaSO4JhSc4PLz263LFaafZDP9atJf7lJ/xt1kGruVfKNZr+BRk9PcHcn/R8o/MmsKVmz4E37/kH/h8r6E1jIN/wnAwNh2YrvXArmsgUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFcHmPXXNcHmPXQGnv2Q792NH/kZf0tVqpW1f2Q792NH/kZf0tVqpWGDM/wLU8Xwg7UFJBSYMvng/wdfQMMtfxTf5or5RWK8Xaw3FNyslymWycgKSmRFeLbiQoYIChuAe0dtXL+qr0nZ/fD1Rnt/wBJu/prAPpx1LX8Uj80U6lr+KR+aK+Y/wCqr0m//MPVH/3R39NP1Vek3/5h6o/+6O/poD6bKabG/VtgAfJrunOU7jGK1C+BNrLV+pde3qLqHVF5uzDVuQttqZMW6hKi5jICicHFbepGPHfmedbA71of8O39+Rr/AIYz9K63wrQ/4dv78jX/AAxn6V1hgwCrtrZD4BP74N0/Ip/wu1rertrZD4BP74N0/Ip/wu0Bu9SlKyDTn7IX+72kfyEv6Wq1WHKtqfshf7vaR/IS/parVYcq1BnP4D4CunVlKgCDa5OQfW3W/PVNnm2j80VoJ8CJxDXTm0txQSkWuTkn1t1vou4Qm0lS5LSUgZJKtqygWt0q6CsOt9IzbTcrey44ppZjuhACmnceaoEdoOOfdXzMuccw7lLhHJ6h9xrJ5nhUU5+avod0y9NOj9Gacmhm8xZt4U0pEaHHdC3OsI80qAOUpzjf11875slUydIlLxxvvLdUM53UoqPzmsA8CARg19DPgdTnbh0D2h59SlrEiS3knsS6oD5q+eaiEpJJxtt419Ifgy6fkaZ6G7Na5KOBxXWSMeDiuMfMqsoGTU9vrrTT7IZ/rVpL/cpP+Nuty07794rTT7IZ/rVpL/cpP+Nusg1drNnwJv3/ACD/AMPlfQmsJ1mz4E37/kH/AIfK+hNag+gIrmuBXNbAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFcHmPXXNcHmPXQGnv2Q792NH/AJGX9LVaqVtX9kO/djR/5GX9LVaqVhgq7TbZ92mphW2I7KkqSVJabGVKA5n2VNfcDrPs05cD/wCVWQ/gV5//AFA2rB/7hM38er7K+giAcecTmgPlx9wOs/5t3D+qp9wOs/5t3D+qr6k++nvoDTj4Demr9ZekK9SLvaZUNpduQltbqMAq6zce6txRz8aHnjfeie73Csg7Vof8O39+Rr/hjP0rrfCtD/h2/vyNf8MZ+ldYYMAq7a2Q+AT++DdPyKf8Lta3q7a2Q+AT++DdPyKf8LtAbvUpSsg05+yF/u9pH8hL+lqtVhyran7IX+72kfyEv6Wq1WHKtQd2nHG1cbTjjasc0KKT6sivby6bgny2WAewvr3+epDR+mL1q28iz2CEqZNLanerSrB4U4yfnFXsOgPpWJwNLuAeLgoDGS1rXu4tatuZUT9O9cHIGxIA3G3LxrMVq+DV0sznUg2OKw0T5zjs1CSP6J3NZe6OPgkxo7rcrWt1TLGcmNFy3j1qBIPurOAYa+DX0UXPpA1hHlvR1t2GC8lyW+oYS5wnPVjt3xjI7+dfQyKy1GjNsMAIZbQlCB8lIGAKotPWe2WC2M2yz2+PAhtDCGWEBCU+wdvjUpWQcJ8K00+yGf61aS/3KT/jbrczAzmtM/shn+tWkv8AcpP+NusMGrtZs+BN+/5B/wCHyvoTWE6zZ8Cb9/yDt/8ADpX0JrAPoCK5rzBOBsrb2V3z4Gtgc0rqSc8jQHPf7qA7UpSgFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAVweY9dc1weY9dAae/ZDv3Y0f+Rl/S1WqlbU/ZD1oTedH8S0p+8y+ZA7Wq1T61r+Nb/PFYBK6av1301d0XexzVw5zaVpQ8gAkBQwRv3iruHTP0nDlq2Z+Yj9FY861r+Nb/ADxTrWv41v8APFAZD/Vo6T/53TPzEfop+rR0n/zumfmI/RWPOta/jW/zxTrWv41v88UBkiH0zdJi5jCVasmKSXUggpRuMjwr6J6NedlaSs8p9wuPPQWHFqI3JU2kmvlVAda8uj/fW/2VHxx8oV9UdB76HsOP/DI3/tJrIJqtD/h2/vyNf8MZ+ldb4Vod8O9aE9MrQUtCT9rGNioDtXWGDASu2tkPgE/vg3T8in/C7WtinWt/vrf54rZH4A60K6QroErQr7ynkoH4rtAbwUpSsg05+yF/u9pH8hL+lqtVhyraj7IatCb9pHiWlP3iXzIHa1Wqodax+yt/nisAzt8B3P6u7OP/AAuT9Ldb8gHv+etBPgOLQrp3ZCXEKP2rk8lA9rdb+CgGDnn7K4KSRiu1KyDqU+NdqUoBWmf2Qz/WrSX+5Sf8bdbmVpl9kOUhOqtI8S0p/WUnmQPjt0Bq9VZZrrcbNPTOtcx6HJSkoDrSsKAPPeqDrWv41v8APFOta/jW/wA8VqC7f1Rtc/zpunj+uFfpp+qNrn+dN0/tCv01aXWtfxrf54p1rX8a3+eKyC7f1Rtc/wA6LoR/vCv01nb4FerNSXvpOkxLrepsxhMQqDbzpUM4V31q51rX8a3+eK2H+AatCuliWErQo+RnkoHsXQG9VKUrIFKUoBSlKAUpSgFKUoBSlKAUpSgFKUoBSlKAUpSgKG6We03RSFXK1wppbyEGRHQ5w5544gccqo/uR0r/ADas39ga/wAtKUA+5HSv82rN/YGv8tPuR0r/ADas39ga/wAtKUA+5HSv82rN/YGv8tPuR0r/ADas39ga/wAtKUAGktLAgjTdmBHL9Ytf5amG0IabS22hKEJASlKRgADkAKUoDtUdcbDZLk/5RcLPb5b2OHjfiocVjuyoE0pQFN9yOlf5tWb+wNf5aqrbY7NbXS7brTAhuK2KmIyGyfakClKAkKUpQFDc7NaLopCrla4M1TYIQZEdDhTnnjiBxyFUf3I6V/m1Zv7A1/lpSgKi32CxW6T5TAs1uiPAFPWMRUIVg8xkAGpKlKAUpSgFKUoBVBdLNaLotC7la4M1TYIQZEdDhSDzxxA4pSgKT7kdK/zas39ga/y0+5HSv82rN/YGv8tKUA+5HSv82rN/YGv8tPuR0r/Nqzf2Br/LSlAPuR0r/Nqzf2Br/LVTbbFZLa8X7faLfDdIwVsRkNqx60gUpQEjSlKAUpSgFKUoBSlKAUpSgP/Z"


# -------------------------------------------------------------
# 1. FIXED EXPLICIT ROUTE FOR FAVICON TO BYPASS 404 EXCEPTIONS
# -------------------------------------------------------------
@app.get("/favicon.ico", include_in_schema=False)
def favicon():
    return Response(status_code=204)


# -------------------------------------------------------------
# 2. FULLY FIXED ASYNC EXCEPTION HANDLER
# -------------------------------------------------------------
@app.exception_handler(StarletteHTTPException)
async def custom_http_exception_handler(request, exc):
    if exc.status_code == 404 and not request.cookies.get("session_user"):
        return RedirectResponse(url="/login", status_code=303)
    return await http_exception_handler(request, exc)


def hash_password(password: str) -> str:
    """Standardized SHA-256 password hashing for system security."""
    return hashlib.sha256(password.encode('utf-8')).hexdigest()


def run_structural_database_migrations():
    """Self-healing migration engine to clean up schema columns in existing databases."""
    # --- Phase 1: Inspect ALL tables while the connection is still open ---
    existing_cols_emp = []
    existing_cols_proc = []
    existing_cols_ma = []
    existing_cols_items = []

    try:
        from sqlalchemy import inspect as _inspect
        _insp = _inspect(engine)
        existing_cols_emp   = [col['name'] for col in _insp.get_columns('employees')]
        existing_cols_proc  = [col['name'] for col in _insp.get_columns('procurement_requests')]
        existing_cols_ma    = [col['name'] for col in _insp.get_columns('material_assignments')]
        existing_cols_items = [col['name'] for col in _insp.get_columns('items')]
    except Exception as e:
        print(f"[Migration] Schema inspection error: {e}")

    # --- Phase 2: Apply any missing ALTER TABLE statements ---
    try:
        with engine.connect() as _conn:
            if "location" not in existing_cols_emp:
                _conn.execute(text("ALTER TABLE employees ADD COLUMN location TEXT DEFAULT 'Not Specified'"))
            if "contact" not in existing_cols_emp:
                _conn.execute(text("ALTER TABLE employees ADD COLUMN contact TEXT DEFAULT 'Not Specified'"))
            if "department" not in existing_cols_proc:
                _conn.execute(text("ALTER TABLE procurement_requests ADD COLUMN department TEXT DEFAULT 'Operations'"))
            if "department" not in existing_cols_ma:
                _conn.execute(text("ALTER TABLE material_assignments ADD COLUMN department TEXT DEFAULT 'General Operations'"))
            if "mis_filename" not in existing_cols_ma:
                _conn.execute(text("ALTER TABLE material_assignments ADD COLUMN mis_filename TEXT"))
            if "mis_uploaded_by_id" not in existing_cols_ma:
                _conn.execute(text("ALTER TABLE material_assignments ADD COLUMN mis_uploaded_by_id INTEGER"))
            if "mis_upload_timestamp" not in existing_cols_ma:
                _conn.execute(text("ALTER TABLE material_assignments ADD COLUMN mis_upload_timestamp DATETIME"))
            if "uom" not in existing_cols_items:
                _conn.execute(text("ALTER TABLE items ADD COLUMN uom TEXT"))
            _conn.commit()
    except Exception as e:
        print(f"[Migration Adaptive Notice] Columns pre-aligned: {e}")

    # Ensure procurement_messages table exists (new feature)
    # Uses engine directly so DDL works on both SQLite and PostgreSQL
    try:
        with engine.connect() as _ddl_conn:
            if engine.dialect.name == "postgresql":
                _ddl_conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS procurement_messages (
                        id SERIAL PRIMARY KEY,
                        request_id INTEGER REFERENCES procurement_requests(id),
                        sender_id INTEGER REFERENCES users(id),
                        message TEXT,
                        timestamp TIMESTAMP
                    )
                """))
            else:
                _ddl_conn.execute(text("""
                    CREATE TABLE IF NOT EXISTS procurement_messages (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        request_id INTEGER REFERENCES procurement_requests(id),
                        sender_id INTEGER REFERENCES users(id),
                        message TEXT,
                        timestamp DATETIME
                    )
                """))
            # Add received_by to grn_records if missing
            try:
                _ddl_conn.execute(text("ALTER TABLE grn_records ADD COLUMN received_by TEXT"))
            except Exception:
                pass  # Column already exists
            # Per-batch vendor + price so the transaction log keeps each
            # inward's own vendor/price instead of the moving catalog value
            try:
                _ddl_conn.execute(text("ALTER TABLE grn_records ADD COLUMN vendor TEXT"))
            except Exception:
                pass
            try:
                _ddl_conn.execute(text("ALTER TABLE grn_records ADD COLUMN unit_price DOUBLE PRECISION"
                                       if engine.dialect.name == "postgresql"
                                       else "ALTER TABLE grn_records ADD COLUMN unit_price REAL"))
            except Exception:
                pass
            # GRN / Challan reference numbers captured at inward time
            try:
                _ddl_conn.execute(text("ALTER TABLE grn_records ADD COLUMN grn_no TEXT"))
            except Exception:
                pass
            try:
                _ddl_conn.execute(text("ALTER TABLE grn_records ADD COLUMN challan_no TEXT"))
            except Exception:
                pass
            # Admin-entered procurement pricing (vendor + rate) that gates ordering
            try:
                _ddl_conn.execute(text("ALTER TABLE procurement_requests ADD COLUMN vendor TEXT"))
            except Exception:
                pass
            try:
                _ddl_conn.execute(text("ALTER TABLE procurement_requests ADD COLUMN unit_price DOUBLE PRECISION"
                                       if engine.dialect.name == "postgresql"
                                       else "ALTER TABLE procurement_requests ADD COLUMN unit_price REAL"))
            except Exception:
                pass

            # --- Store uploaded file bytes in the shared DB so they are
            #     available across all devices (local disk is per-machine
            #     and ephemeral on Render). BYTEA on Postgres, BLOB on SQLite. ---
            _blob_type = "BYTEA" if engine.dialect.name == "postgresql" else "BLOB"
            try:
                _ddl_conn.execute(text(f"ALTER TABLE grn_records ADD COLUMN grn_filedata {_blob_type}"))
            except Exception:
                pass  # Column already exists
            try:
                _ddl_conn.execute(text(f"ALTER TABLE material_assignments ADD COLUMN mis_filedata {_blob_type}"))
            except Exception:
                pass  # Column already exists
            # --- Return-to-Store columns on material_assignments (same table) ---
            _bool_default = "BOOLEAN DEFAULT FALSE" if engine.dialect.name == "postgresql" else "BOOLEAN DEFAULT 0"
            try:
                _ddl_conn.execute(text(f"ALTER TABLE material_assignments ADD COLUMN is_return {_bool_default}"))
            except Exception:
                pass
            try:
                _ddl_conn.execute(text("ALTER TABLE material_assignments ADD COLUMN return_filename TEXT"))
            except Exception:
                pass
            try:
                _ddl_conn.execute(text(f"ALTER TABLE material_assignments ADD COLUMN return_filedata {_blob_type}"))
            except Exception:
                pass
            # --- ASSET CLASS CATEGORY columns ---
            try:
                _ddl_conn.execute(text(f"ALTER TABLE items ADD COLUMN category TEXT DEFAULT '{DEFAULT_ASSET_CLASS}'"))
            except Exception:
                pass
            try:
                _ddl_conn.execute(text("ALTER TABLE procurement_requests ADD COLUMN category TEXT"))
            except Exception:
                pass
            try:
                _ddl_conn.execute(text("ALTER TABLE procurement_requests ADD COLUMN uom TEXT DEFAULT 'Nos'"))
            except Exception:
                pass
            # --- Fixed Assets and Equipments per-unit tracking ---
            try:
                _ddl_conn.execute(text("ALTER TABLE material_assignments ADD COLUMN asset_unit_id INTEGER"))
            except Exception:
                pass
            _ddl_conn.commit()
    except Exception as e:
        print(f"[Migration] procurement_messages: {e}")


run_structural_database_migrations()


def seed_system_data():
    """Populates basic master data if tables are freshly generated."""
    db = SessionLocal()
    try:
        if db.query(models.User).count() == 0:
            db.add_all([
                models.User(
                    username="admin",
                    hashed_password=hash_password("password123"),
                    role="Admin",
                    full_name="Rajesh Kumar Sharma",
                    designation="Project General Manager",
                    workstation_location="Corporate HQ - New Delhi"
                ),
                models.User(
                    username="staff",
                    hashed_password=hash_password("password123"),
                    role="Staff",
                    full_name="Amit Kumar Verma",
                    designation="Store Logistics Officer",
                    workstation_location="Okhla Logistics Hub"
                )
            ])
            db.commit()
        if db.query(models.Employee).count() == 0:
            db.add_all([
                models.Employee(name="Piyush Bhatia", role_title="Sr. Manager- Commercial", contact="9926306363", location="Corporate HQ"),
                models.Employee(name="Biswajit Pradhan", role_title="Engineer", contact="9348653235", location="Udaipur Site Base")
            ])
            db.commit()
    finally:
        db.close()


seed_system_data()


def get_current_user(session_user: str = Cookie(None), db: Session = Depends(get_db)):
    """Cookie-based identity validation session guard."""
    if not session_user:
        return None
    return db.query(models.User).filter(models.User.username == session_user).first()


@app.get("/login", response_class=HTMLResponse)
def login_page():
    return templates.LOGIN_HTML


@app.post("/login")
def do_login(response: Response, username: str = Form(...), password: str = Form(...), db: Session = Depends(get_db)):
    user = db.query(models.User).filter(
        models.User.username == username.strip().lower()).first()
    if not user or user.hashed_password != hash_password(password):
        return HTMLResponse("<script>alert('Invalid credentials!'); window.location='/login';</script>")

    res = RedirectResponse(url="/", status_code=303)
    res.set_cookie(key="session_user", value=user.username, samesite="lax", httponly=True)
    return res


@app.get("/logout")
def do_logout():
    res = RedirectResponse(url="/login", status_code=303)
    res.delete_cookie("session_user")
    return res


@app.post("/transaction")
async def create_transaction(
    item_id: str = Form(...),              # "NEW_INWARD_ITEM" or existing item id
    quantity: int = Form(...),
    uom: str = Form(...),
    grn_no: str = Form(None),
    challan_no: str = Form(None),
    new_item_name: str = Form(None),
    new_item_code: str = Form(None),
    new_item_category: str = Form(None),       # used when creating a new item
    serial_numbers: str = Form(None),          # newline/comma-separated for tracked items
    site: str = Form(None),
    grn_file: UploadFile = File(None),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)

    if not grn_file or not grn_file.filename:
        return HTMLResponse("<script>alert('GRN Upload is mandatory before recording an Inward Transaction.'); window.history.back();</script>")

    import os, datetime as _dt
    grn_dir = "grn_uploads"
    os.makedirs(grn_dir, exist_ok=True)
    timestamp_str = _dt.datetime.utcnow().strftime("%Y%m%d_%H%M%S")

    placed_orders = []

    if item_id == "NEW_INWARD_ITEM":
        if not new_item_name or not new_item_code:
            return HTMLResponse("<script>alert('Item Name and Item Code are required for a new item.'); window.history.back();</script>")
        clean_code = new_item_code.strip().upper()
        existing = db.query(models.Item).filter(models.Item.item_code == clean_code).first()
        if existing:
            return HTMLResponse(f"<script>alert('Item Code {clean_code} already exists in catalog. Please use Search & Select for existing items.'); window.history.back();</script>")
        cat = (new_item_category or "").strip()
        if cat not in ASSET_CLASSES:
            cat = DEFAULT_ASSET_CLASS
        new_item = models.Item(
            name=new_item_name.strip(),
            item_code=clean_code,
            category=cat,
            current_stock=0,
            price=0.0,
            supplier="Approved Vendor",
            storage_site=site.strip() if site else "Store Yard",
            uom=uom,
            minimum_stock=0
        )
        db.add(new_item)
        db.flush()
        item = new_item
        lot_vendor = item.supplier or "Approved Vendor"
        lot_price = item.price or 0.0
    else:
        item = db.query(models.Item).filter(models.Item.id == int(item_id)).first()
        if not item:
            return HTMLResponse("<script>alert('Item not found.'); window.location='/';</script>")
        if not item.uom:
            item.uom = uom
        uom = item.uom

        placed_orders = db.query(models.ProcurementRequest).filter(
            models.ProcurementRequest.item_id == item.id,
            models.ProcurementRequest.status == "Order Placed"
        ).order_by(models.ProcurementRequest.timestamp.asc()).all()
        source_order = placed_orders[0] if placed_orders else None

        if source_order and source_order.vendor:
            lot_vendor = source_order.vendor
        else:
            lot_vendor = item.supplier or "Approved Vendor"
        if source_order and source_order.unit_price is not None:
            lot_price = float(source_order.unit_price)
        else:
            lot_price = item.price or 0.0

        item.supplier = lot_vendor
        item.price = lot_price

    # --- SERIAL-TRACKED PATH (Fixed Assets and Equipments) ---
    serial_list = []
    if is_tracked(item):
        raw = (serial_numbers or "").replace(",", "\n")
        serial_list = [s.strip() for s in raw.split("\n") if s.strip()]
        if not serial_list:
            return HTMLResponse("<script>alert('Serial number(s) are required for Fixed Assets and Equipments. Enter one serial per line.'); window.history.back();</script>")
        if quantity != len(serial_list):
            return HTMLResponse(f"<script>alert('Quantity ({quantity}) must equal the number of serials provided ({len(serial_list)}).'); window.history.back();</script>")
        if len(set(serial_list)) != len(serial_list):
            return HTMLResponse("<script>alert('Duplicate serials within this entry. Each serial must be unique.'); window.history.back();</script>")
        existing_serials = {row[0] for row in db.query(models.AssetUnit.serial_no).filter(
            models.AssetUnit.serial_no.in_(serial_list)
        ).all()}
        dup = [s for s in serial_list if s in existing_serials]
        if dup:
            return HTMLResponse(f"<script>alert('These serial numbers already exist: {', '.join(dup[:5])}'); window.history.back();</script>")
        now = _dt.datetime.utcnow()
        for s in serial_list:
            db.add(models.AssetUnit(item_id=item.id, serial_no=s, status="In Stock", acquired_at=now))
        item.current_stock = (item.current_stock or 0) + len(serial_list)
    else:
        add_to_lot(db, item, lot_vendor, lot_price, quantity)

    item_unit_price = lot_price

    received_by_name = current_user.full_name or current_user.username

    safe_filename = f"GRN_{item.item_code}_{timestamp_str}_{grn_file.filename.replace(' ', '_')}"
    grn_path = os.path.join(grn_dir, safe_filename)
    contents = await grn_file.read()
    with open(grn_path, "wb") as f:
        f.write(contents)

    db.add(models.Transaction(
        item_id=item.id,
        type="IN",
        quantity=quantity,
        user_id=current_user.id,
        total_value=float(quantity * item_unit_price)
    ))

    db.add(models.GRNRecord(
        item_id=item.id,
        quantity=quantity,
        uom=uom,
        received_by=received_by_name,
        vendor=lot_vendor,
        unit_price=item_unit_price,
        grn_no=(grn_no.strip() if grn_no else None),
        challan_no=(challan_no.strip() if challan_no else None),
        grn_filename=safe_filename,
        grn_filedata=contents,
        uploaded_by_id=current_user.id
    ))

    # Mark the placed order(s) for this item as Delivered
    for ord_req in placed_orders:
        ord_req.status = "Delivered"

    db.commit()
    return RedirectResponse(url="/", status_code=303)


@app.get("/inward/bulk-template")
def inward_bulk_template(current_user: models.User = Depends(get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    from fastapi.responses import StreamingResponse
    import io
    csv_content = "item_code_or_NEW,item_name,quantity,uom,vendor,price,site\n"
    csv_content += "EIPL-ST-01,,50,Nos,Tata Steel,350.00,Udaipur Yard\n"
    csv_content += "NEW,Copper Cable 4mm,100,Mtr,Polycab,125.50,Store Room A\n"
    return StreamingResponse(
        io.StringIO(csv_content),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=EIPL_Inward_Transaction_Template.csv"}
    )


@app.post("/inward/bulk-import")
async def inward_bulk_import(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    try:
        contents = await file.read()
        try:
            decoded = contents.decode('utf-8-sig')
        except UnicodeDecodeError:
            decoded = contents.decode('latin-1')
        lines = [l for l in decoded.splitlines() if l.strip() and not l.strip().startswith('#')]

        try:
            sample_data = "\n".join(lines[:5])
            dialect = csv.Sniffer().sniff(sample_data)
            if dialect.delimiter not in [',', ';', '\t', '|']:
                dialect.delimiter = ','
            reader = csv.DictReader(lines, dialect=dialect)
        except Exception:
            reader = csv.DictReader(lines)

        added_count = 0
        updated_count = 0

        for row in reader:
            if not row:
                continue
            clean_row = {str(k).strip().lower(): str(v).strip() for k, v in row.items() if k is not None}

            item_code_raw = clean_row.get("item_code_or_new") or clean_row.get("item_code") or ""
            item_name = clean_row.get("item_name") or clean_row.get("name") or ""
            qty_raw = clean_row.get("quantity") or clean_row.get("qty") or "0"
            uom = clean_row.get("uom") or "Nos"
            vendor = clean_row.get("vendor") or clean_row.get("supplier") or "Approved Vendor"
            price_raw = clean_row.get("price") or clean_row.get("rate") or "0.0"
            site = clean_row.get("site") or clean_row.get("storage_site") or "Store Yard"

            try:
                quantity = int(float(qty_raw))
                price = float(price_raw)
            except ValueError:
                continue

            if quantity <= 0:
                continue

            is_new = item_code_raw.strip().upper() == "NEW"

            if is_new:
                if not item_name:
                    continue
                import random, string as _string
                gen_code = "EIPL-" + ''.join(random.choices(_string.ascii_uppercase, k=2)) + "-" + ''.join(random.choices(_string.digits, k=2))
                new_item = models.Item(
                    name=item_name, item_code=gen_code, current_stock=0,
                    price=price, supplier=vendor, storage_site=site, uom=uom, minimum_stock=0
                )
                db.add(new_item)
                db.flush()
                add_to_lot(db, new_item, vendor, price, quantity)
                db.add(models.Transaction(item_id=new_item.id, type="IN", quantity=quantity, user_id=current_user.id, total_value=float(quantity * price)))
                db.add(models.GRNRecord(item_id=new_item.id, quantity=quantity, uom=uom,
                    received_by=current_user.full_name or current_user.username,
                    vendor=vendor, unit_price=price,
                    grn_filename="BULK_CSV_IMPORT", uploaded_by_id=current_user.id))
                added_count += 1
            else:
                clean_code = item_code_raw.strip().upper()
                if not clean_code:
                    continue
                item = db.query(models.Item).filter(models.Item.item_code == clean_code).first()
                if not item:
                    if not item_name:
                        continue
                    item = models.Item(name=item_name, item_code=clean_code, current_stock=0,
                        price=price, supplier=vendor, storage_site=site, uom=uom, minimum_stock=0)
                    db.add(item)
                    db.flush()
                    added_count += 1
                else:
                    if price > 0:
                        item.price = price
                    if vendor and vendor != "Approved Vendor":
                        item.supplier = vendor
                    if not item.uom:
                        item.uom = uom
                    uom = item.uom  # Use locked uom
                    updated_count += 1

                lot_vendor = vendor if (vendor and vendor != "Approved Vendor") else item.supplier
                lot_price = price if price > 0 else (item.price or 0.0)
                add_to_lot(db, item, lot_vendor, lot_price, quantity)
                db.add(models.Transaction(item_id=item.id, type="IN", quantity=quantity, user_id=current_user.id, total_value=float(quantity * item.price)))
                db.add(models.GRNRecord(item_id=item.id, quantity=quantity, uom=uom,
                    received_by=current_user.full_name or current_user.username,
                    vendor=lot_vendor, unit_price=lot_price,
                    grn_filename="BULK_CSV_IMPORT", uploaded_by_id=current_user.id))

        db.commit()
        return HTMLResponse(f"<script>alert('Bulk Inward Import Complete!\\nNew Items Added: {added_count}\\nExisting Items Updated: {updated_count}'); window.location='/';</script>")
    except Exception as e:
        db.rollback()
        return HTMLResponse(f"<script>alert('Import error: {str(e)}'); window.location='/';</script>")


@app.get("/grn/list", response_class=HTMLResponse)
def grn_list(current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    grns = db.query(models.GRNRecord).order_by(models.GRNRecord.timestamp.desc()).all()
    rows = ""
    for g in grns:
        ts = fmt_ist(g.timestamp)
        ts_iso = fmt_ist(g.timestamp, "%Y-%m-%d")
        item_name = g.item.name if g.item else "Unknown Item"
        item_code = g.item.item_code if g.item else "—"
        uploader = g.uploader.username if g.uploader else "System"
        vendor = getattr(g, 'vendor', '') or '—'
        grn_no = getattr(g, 'grn_no', '') or '—'
        challan_no = getattr(g, 'challan_no', '') or '—'
        rows += f"""
        <tr class="border-b hover:bg-slate-50 text-xs grn-row"
            data-date="{ts_iso}" data-item="{item_name.lower()}" data-code="{item_code.lower()}"
            data-vendor="{vendor.lower()}" data-uploader="{uploader.lower()}" data-grn="{grn_no.lower()}">
            <td class="p-3 text-slate-500 font-mono whitespace-nowrap">{ts}</td>
            <td class="p-3 font-semibold text-slate-800">{item_name}</td>
            <td class="p-3 font-mono text-slate-500">{item_code}</td>
            <td class="p-3 text-center font-mono font-bold text-slate-700">{g.quantity} {g.uom}</td>
            <td class="p-3 text-slate-600">{vendor}</td>
            <td class="p-3 font-mono text-slate-500 text-center">{grn_no}</td>
            <td class="p-3 font-mono text-slate-500 text-center">{challan_no}</td>
            <td class="p-3 text-slate-500">{uploader}</td>
            <td class="p-3 text-right">
                <a href="/grn/download/{g.id}" class="bg-indigo-600 hover:bg-indigo-700 text-white font-bold text-[10px] px-2.5 py-1.5 rounded-lg transition-all">&#11015; Download</a>
            </td>
        </tr>"""
    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>GRN Downloads - EIPL</title>
<script src="https://cdn.tailwindcss.com"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
<style>body{{font-family:'Inter',sans-serif;}}.filter-input{{font-size:11px;padding:4px 8px;border:1px solid #e2e8f0;border-radius:6px;background:#f8fafc;width:100%;box-sizing:border-box;}}
.filter-input:focus{{outline:none;border-color:#6366f1;background:#fff;}}</style>
</head>
<body class="bg-slate-50 min-h-screen flex">
{build_sidebar(current_user, current_page='grn')}
<div class="flex-1 min-h-screen overflow-x-hidden p-6">
    <div class="max-w-7xl mx-auto">
        <div class="flex items-center justify-between mb-5">
            <div>
                <a href="/" class="text-indigo-600 text-xs font-bold hover:underline">&#8592; Back to Dashboard</a>
                <h1 class="text-xl font-black text-slate-900 mt-1">&#128196; GRN Download Centre</h1>
                <p class="text-xs text-slate-400 mt-0.5">Goods Received Notes — uploaded inward transaction records</p>
            </div>
            <div class="flex items-center gap-2">
                <button onclick="exportGRN()" class="bg-emerald-600 hover:bg-emerald-700 text-white font-bold text-xs px-4 py-2 rounded-lg transition-all flex items-center gap-1.5">
                    <i class="fa-solid fa-file-excel"></i> Export Excel
                </button>
                <a href="/mis/list" class="bg-indigo-50 hover:bg-indigo-100 text-indigo-700 font-bold text-xs px-4 py-2 rounded-lg transition-all border border-indigo-200">&#128203; MIS Centre</a>
            </div>
        </div>

        <!-- Filters Bar -->
        <div class="bg-white border border-slate-200 rounded-xl p-4 mb-4 shadow-sm">
            <div class="grid grid-cols-2 md:grid-cols-4 lg:grid-cols-7 gap-3 items-end">
                <div class="lg:col-span-2">
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">&#128269; Search</label>
                    <input type="text" id="grnSearch" placeholder="Item name, code, vendor..." oninput="applyGRNFilters()"
                        class="filter-input">
                </div>
                <div>
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">From Date</label>
                    <input type="date" id="grnFrom" oninput="applyGRNFilters()" class="filter-input">
                </div>
                <div>
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">To Date</label>
                    <input type="date" id="grnTo" oninput="applyGRNFilters()" class="filter-input">
                </div>
                <div>
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">Item Code</label>
                    <input type="text" id="grnCode" placeholder="Filter code..." oninput="applyGRNFilters()" class="filter-input">
                </div>
                <div>
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">Vendor</label>
                    <input type="text" id="grnVendor" placeholder="Filter vendor..." oninput="applyGRNFilters()" class="filter-input">
                </div>
                <div>
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">Uploaded By</label>
                    <input type="text" id="grnUploader" placeholder="Filter user..." oninput="applyGRNFilters()" class="filter-input">
                </div>
            </div>
            <div class="flex items-center justify-between mt-2.5">
                <span id="grnCount" class="text-[11px] text-slate-400 font-mono">Showing all records</span>
                <div class="flex items-center gap-3">
                    <button onclick="applyGRNFilters()" class="bg-indigo-600 hover:bg-indigo-700 text-white text-[10px] font-bold px-3 py-1.5 rounded-lg transition-all flex items-center gap-1"><i class="fa-solid fa-filter"></i> Apply Filters</button>
                    <button onclick="clearGRNFilters()" class="text-[10px] text-rose-500 font-bold hover:underline">Clear Filters</button>
                </div>
            </div>
        </div>

        <div class="bg-white border border-slate-200 rounded-2xl overflow-hidden shadow-sm">
            <div class="overflow-x-auto">
            <table class="w-full text-left border-collapse" id="grnTable">
                <thead>
                    <tr class="bg-slate-50 text-slate-500 font-semibold tracking-wider uppercase text-[10px] border-b border-slate-200">
                        <th class="p-3 whitespace-nowrap">Timestamp</th>
                        <th class="p-3">Item Name</th>
                        <th class="p-3">Item Code</th>
                        <th class="p-3 text-center">Qty / UOM</th>
                        <th class="p-3">Vendor</th>
                        <th class="p-3 text-center">GRN No.</th>
                        <th class="p-3 text-center">Challan No.</th>
                        <th class="p-3">Uploaded By</th>
                        <th class="p-3 text-right pr-4">Action</th>
                    </tr>
                </thead>
                <tbody class="divide-y divide-slate-100" id="grnTbody">{rows if rows else '<tr><td colspan="9" class="p-8 text-center text-slate-400 text-sm">No GRNs uploaded yet.</td></tr>'}</tbody>
            </table>
            </div>
            <!-- Pagination -->
            <div class="flex items-center justify-between px-4 py-3 border-t border-slate-100 bg-slate-50/50">
                <span class="text-[11px] text-slate-400" id="grnPageInfo"></span>
                <div class="flex items-center gap-1" id="grnPagination"></div>
            </div>
        </div>
    </div>

<script>
var GRN_PAGE = 1, GRN_PER_PAGE = 20, GRN_FILTERED = null;
function getAllGRNRows() {{ return Array.from(document.querySelectorAll('#grnTbody .grn-row')); }}
function applyGRNFilters() {{
    var q = (document.getElementById('grnSearch').value||'').toLowerCase().trim();
    var from = document.getElementById('grnFrom').value;
    var to = document.getElementById('grnTo').value;
    var code = (document.getElementById('grnCode').value||'').toLowerCase().trim();
    var vendor = (document.getElementById('grnVendor').value||'').toLowerCase().trim();
    var uploader = (document.getElementById('grnUploader').value||'').toLowerCase().trim();
    GRN_FILTERED = getAllGRNRows().filter(function(r) {{
        var d = r.dataset.date||'';
        if(from && d < from) return false;
        if(to && d > to) return false;
        if(q && !r.dataset.item.includes(q) && !r.dataset.code.includes(q) && !r.dataset.vendor.includes(q) && !r.dataset.grn.includes(q)) return false;
        if(code && !r.dataset.code.includes(code)) return false;
        if(vendor && !r.dataset.vendor.includes(vendor)) return false;
        if(uploader && !r.dataset.uploader.includes(uploader)) return false;
        return true;
    }});
    GRN_PAGE = 1;
    renderGRNPage();
}}
function renderGRNPage() {{
    var rows = GRN_FILTERED !== null ? GRN_FILTERED : getAllGRNRows();
    var total = rows.length;
    var pages = Math.ceil(total / GRN_PER_PAGE) || 1;
    if(GRN_PAGE > pages) GRN_PAGE = pages;
    var start = (GRN_PAGE-1)*GRN_PER_PAGE, end = start+GRN_PER_PAGE;
    getAllGRNRows().forEach(function(r){{r.style.display='none';}});
    rows.slice(start,end).forEach(function(r){{r.style.display='';}});
    document.getElementById('grnCount').textContent = 'Showing ' + Math.min(start+1,total) + '–' + Math.min(end,total) + ' of ' + total + ' records';
    document.getElementById('grnPageInfo').textContent = 'Page ' + GRN_PAGE + ' of ' + pages;
    var pgDiv = document.getElementById('grnPagination');
    pgDiv.innerHTML = '';
    if(pages <= 1) return;
    var mkBtn = function(label,pg,active) {{
        var b = document.createElement('button');
        b.textContent = label;
        b.className = 'px-2.5 py-1 rounded text-[11px] font-bold border ' + (active ? 'bg-indigo-600 text-white border-indigo-600' : 'bg-white text-slate-600 border-slate-200 hover:bg-slate-50');
        b.onclick = function(){{ GRN_PAGE=pg; renderGRNPage(); }};
        return b;
    }};
    pgDiv.appendChild(mkBtn('«',1,false));
    pgDiv.appendChild(mkBtn('‹',Math.max(1,GRN_PAGE-1),false));
    for(var p=Math.max(1,GRN_PAGE-2);p<=Math.min(pages,GRN_PAGE+2);p++) pgDiv.appendChild(mkBtn(p,p,p===GRN_PAGE));
    pgDiv.appendChild(mkBtn('›',Math.min(pages,GRN_PAGE+1),false));
    pgDiv.appendChild(mkBtn('»',pages,false));
}}
function clearGRNFilters() {{
    ['grnSearch','grnFrom','grnTo','grnCode','grnVendor','grnUploader'].forEach(function(id){{document.getElementById(id).value='';}});
    GRN_FILTERED=null; GRN_PAGE=1; renderGRNPage();
}}
function exportGRN() {{
    var from = document.getElementById('grnFrom').value;
    var to = document.getElementById('grnTo').value;
    var rows = GRN_FILTERED !== null ? GRN_FILTERED : getAllGRNRows();
    var data = [['Timestamp','Item Name','Item Code','Qty','UOM','Vendor','GRN No.','Challan No.','Uploaded By']];
    rows.forEach(function(r) {{
        var cells = r.querySelectorAll('td');
        data.push([
            cells[0].textContent.trim(), cells[1].textContent.trim(), cells[2].textContent.trim(),
            cells[3].textContent.trim().split(' ')[0], cells[3].textContent.trim().split(' ')[1]||'',
            cells[4].textContent.trim(), cells[5].textContent.trim(), cells[6].textContent.trim(), cells[7].textContent.trim()
        ]);
    }});
    var ws = XLSX.utils.aoa_to_sheet(data);
    ws['!cols'] = [{{wch:18}},{{wch:28}},{{wch:14}},{{wch:8}},{{wch:8}},{{wch:22}},{{wch:14}},{{wch:14}},{{wch:14}}];
    var wb = XLSX.utils.book_new();
    XLSX.utils.book_append_sheet(wb, ws, 'GRN Records');
    var fname = 'EIPL_GRN_Export' + (from?'_from_'+from:'') + (to?'_to_'+to:'') + '.xlsx';
    XLSX.writeFile(wb, fname);
}}
window.onload = function() {{ renderGRNPage(); }};
</script>
</div>
</body></html>"""
    return HTMLResponse(html)


@app.get("/grn/download/{grn_id}")
def grn_download(grn_id: int, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    import os
    from fastapi.responses import FileResponse, Response
    grn = db.query(models.GRNRecord).filter(models.GRNRecord.id == grn_id).first()
    if not grn:
        raise HTTPException(status_code=404, detail="GRN not found")

    # Preferred: bytes stored in the shared DB (works on every device)
    if getattr(grn, "grn_filedata", None):
        return Response(
            content=grn.grn_filedata,
            media_type="application/octet-stream",
            headers={"Content-Disposition": f'attachment; filename="{grn.grn_filename}"'}
        )

    # Fallback: legacy records whose bytes only live on this machine's disk
    grn_path = os.path.join("grn_uploads", grn.grn_filename or "")
    if grn.grn_filename and os.path.exists(grn_path):
        return FileResponse(path=grn_path, filename=grn.grn_filename, media_type="application/octet-stream")

    raise HTTPException(status_code=404, detail="GRN file missing from server")


@app.get("/mis/list", response_class=HTMLResponse)
def mis_list(current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    assignments = db.query(models.MaterialAssignment).filter(
        models.MaterialAssignment.mis_filename != None
    ).order_by(models.MaterialAssignment.timestamp.desc()).all()
    rows = ""
    for a in assignments:
        ts = fmt_ist(a.timestamp)
        ts_iso = fmt_ist(a.timestamp, "%Y-%m-%d")
        item_name = a.item.name if a.item else "Unknown Item"
        item_code = a.item.item_code if a.item else "—"
        dept = getattr(a, 'department', '—') or '—'
        issued_by = getattr(a, 'issued_by', '—') or '—'
        rows += f"""
        <tr class="border-b hover:bg-slate-50 text-xs mis-row"
            data-date="{ts_iso}" data-item="{item_name.lower()}" data-code="{item_code.lower()}"
            data-issued="{a.issued_to.lower() if a.issued_to else ''}" data-dept="{dept.lower()}" data-by="{issued_by.lower()}">
            <td class="p-3 text-slate-500 font-mono whitespace-nowrap">{ts}</td>
            <td class="p-3 font-semibold text-slate-800">{item_name}</td>
            <td class="p-3 font-mono text-slate-500">{item_code}</td>
            <td class="p-3 text-center font-mono font-bold text-slate-700">{a.quantity} {a.uom}</td>
            <td class="p-3 text-slate-700 font-medium">{a.issued_to}</td>
            <td class="p-3 text-slate-500">{dept}</td>
            <td class="p-3 text-slate-400 text-[11px]">{issued_by}</td>
            <td class="p-3 text-right">
                <a href="/mis/download/{a.id}" class="bg-indigo-600 hover:bg-indigo-700 text-white font-bold text-[10px] px-2.5 py-1.5 rounded-lg transition-all">&#11015; Download MIS</a>
            </td>
        </tr>"""
    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>MIS Download Centre - EIPL</title>
<script src="https://cdn.tailwindcss.com"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
<style>body{{font-family:'Inter',sans-serif;}}.filter-input{{font-size:11px;padding:4px 8px;border:1px solid #e2e8f0;border-radius:6px;background:#f8fafc;width:100%;box-sizing:border-box;}}
.filter-input:focus{{outline:none;border-color:#6366f1;background:#fff;}}</style>
</head>
<body class="bg-slate-50 min-h-screen flex">
{build_sidebar(current_user, current_page='mis')}
<div class="flex-1 min-h-screen overflow-x-hidden p-6">
    <div class="max-w-7xl mx-auto">
        <div class="flex items-center justify-between mb-5">
            <div>
                <a href="/" class="text-indigo-600 text-xs font-bold hover:underline">&#8592; Back to Dashboard</a>
                <h1 class="text-xl font-black text-slate-900 mt-1">&#128203; MIS Download Centre</h1>
                <p class="text-xs text-slate-400 mt-0.5">Material Issue Slips — uploaded issuance records</p>
            </div>
            <div class="flex items-center gap-2">
                <button onclick="exportMIS()" class="bg-emerald-600 hover:bg-emerald-700 text-white font-bold text-xs px-4 py-2 rounded-lg transition-all flex items-center gap-1.5">
                    <i class="fa-solid fa-file-excel"></i> Export Excel
                </button>
                <a href="/grn/list" class="bg-slate-100 hover:bg-slate-200 text-slate-700 font-bold text-xs px-4 py-2 rounded-lg transition-all border border-slate-200">&#128196; GRN Centre</a>
            </div>
        </div>

        <!-- Filters Bar -->
        <div class="bg-white border border-slate-200 rounded-xl p-4 mb-4 shadow-sm">
            <div class="grid grid-cols-2 md:grid-cols-4 lg:grid-cols-7 gap-3 items-end">
                <div class="lg:col-span-2">
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">&#128269; Search</label>
                    <input type="text" id="misSearch" placeholder="Item name, code, issued to..." oninput="applyMISFilters()"
                        class="filter-input">
                </div>
                <div>
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">From Date</label>
                    <input type="date" id="misFrom" oninput="applyMISFilters()" class="filter-input">
                </div>
                <div>
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">To Date</label>
                    <input type="date" id="misTo" oninput="applyMISFilters()" class="filter-input">
                </div>
                <div>
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">Item Code</label>
                    <input type="text" id="misCode" placeholder="Filter code..." oninput="applyMISFilters()" class="filter-input">
                </div>
                <div>
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">Issued To</label>
                    <input type="text" id="misIssued" placeholder="Filter employee..." oninput="applyMISFilters()" class="filter-input">
                </div>
                <div>
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">Department</label>
                    <input type="text" id="misDept" placeholder="Filter dept..." oninput="applyMISFilters()" class="filter-input">
                </div>
            </div>
            <div class="flex items-center justify-between mt-2.5">
                <span id="misCount" class="text-[11px] text-slate-400 font-mono">Showing all records</span>
                <div class="flex items-center gap-3">
                    <button onclick="applyMISFilters()" class="bg-indigo-600 hover:bg-indigo-700 text-white text-[10px] font-bold px-3 py-1.5 rounded-lg transition-all flex items-center gap-1"><i class="fa-solid fa-filter"></i> Apply Filters</button>
                    <button onclick="clearMISFilters()" class="text-[10px] text-rose-500 font-bold hover:underline">Clear Filters</button>
                </div>
            </div>
        </div>

        <div class="bg-white border border-slate-200 rounded-2xl overflow-hidden shadow-sm">
            <div class="overflow-x-auto">
            <table class="w-full text-left border-collapse" id="misTable">
                <thead>
                    <tr class="bg-slate-50 text-slate-500 font-semibold tracking-wider uppercase text-[10px] border-b border-slate-200">
                        <th class="p-3 whitespace-nowrap">Timestamp</th>
                        <th class="p-3">Item Name</th>
                        <th class="p-3">Item Code</th>
                        <th class="p-3 text-center">Qty / UOM</th>
                        <th class="p-3">Issued To</th>
                        <th class="p-3">Department</th>
                        <th class="p-3">Issued By</th>
                        <th class="p-3 text-right pr-4">Action</th>
                    </tr>
                </thead>
                <tbody class="divide-y divide-slate-100" id="misTbody">{rows if rows else '<tr><td colspan="8" class="p-8 text-center text-slate-400 text-sm">No MIS records uploaded yet.</td></tr>'}</tbody>
            </table>
            </div>
            <!-- Pagination -->
            <div class="flex items-center justify-between px-4 py-3 border-t border-slate-100 bg-slate-50/50">
                <span class="text-[11px] text-slate-400" id="misPageInfo"></span>
                <div class="flex items-center gap-1" id="misPagination"></div>
            </div>
        </div>
    </div>

<script>
var MIS_PAGE = 1, MIS_PER_PAGE = 20, MIS_FILTERED = null;
function getAllMISRows() {{ return Array.from(document.querySelectorAll('#misTbody .mis-row')); }}
function applyMISFilters() {{
    var q = (document.getElementById('misSearch').value||'').toLowerCase().trim();
    var from = document.getElementById('misFrom').value;
    var to = document.getElementById('misTo').value;
    var code = (document.getElementById('misCode').value||'').toLowerCase().trim();
    var issued = (document.getElementById('misIssued').value||'').toLowerCase().trim();
    var dept = (document.getElementById('misDept').value||'').toLowerCase().trim();
    MIS_FILTERED = getAllMISRows().filter(function(r) {{
        var d = r.dataset.date||'';
        if(from && d < from) return false;
        if(to && d > to) return false;
        if(q && !r.dataset.item.includes(q) && !r.dataset.code.includes(q) && !r.dataset.issued.includes(q)) return false;
        if(code && !r.dataset.code.includes(code)) return false;
        if(issued && !r.dataset.issued.includes(issued)) return false;
        if(dept && !r.dataset.dept.includes(dept)) return false;
        return true;
    }});
    MIS_PAGE = 1;
    renderMISPage();
}}
function renderMISPage() {{
    var rows = MIS_FILTERED !== null ? MIS_FILTERED : getAllMISRows();
    var total = rows.length;
    var pages = Math.ceil(total / MIS_PER_PAGE) || 1;
    if(MIS_PAGE > pages) MIS_PAGE = pages;
    var start = (MIS_PAGE-1)*MIS_PER_PAGE, end = start+MIS_PER_PAGE;
    getAllMISRows().forEach(function(r){{r.style.display='none';}});
    rows.slice(start,end).forEach(function(r){{r.style.display='';}});
    document.getElementById('misCount').textContent = 'Showing ' + Math.min(start+1,total) + '–' + Math.min(end,total) + ' of ' + total + ' records';
    document.getElementById('misPageInfo').textContent = 'Page ' + MIS_PAGE + ' of ' + pages;
    var pgDiv = document.getElementById('misPagination');
    pgDiv.innerHTML = '';
    if(pages <= 1) return;
    var mkBtn = function(label,pg,active) {{
        var b = document.createElement('button');
        b.textContent = label;
        b.className = 'px-2.5 py-1 rounded text-[11px] font-bold border ' + (active ? 'bg-indigo-600 text-white border-indigo-600' : 'bg-white text-slate-600 border-slate-200 hover:bg-slate-50');
        b.onclick = function(){{ MIS_PAGE=pg; renderMISPage(); }};
        return b;
    }};
    pgDiv.appendChild(mkBtn('«',1,false));
    pgDiv.appendChild(mkBtn('‹',Math.max(1,MIS_PAGE-1),false));
    for(var p=Math.max(1,MIS_PAGE-2);p<=Math.min(pages,MIS_PAGE+2);p++) pgDiv.appendChild(mkBtn(p,p,p===MIS_PAGE));
    pgDiv.appendChild(mkBtn('›',Math.min(pages,MIS_PAGE+1),false));
    pgDiv.appendChild(mkBtn('»',pages,false));
}}
function clearMISFilters() {{
    ['misSearch','misFrom','misTo','misCode','misIssued','misDept'].forEach(function(id){{document.getElementById(id).value='';}});
    MIS_FILTERED=null; MIS_PAGE=1; renderMISPage();
}}
function exportMIS() {{
    var from = document.getElementById('misFrom').value;
    var to = document.getElementById('misTo').value;
    var rows = MIS_FILTERED !== null ? MIS_FILTERED : getAllMISRows();
    var data = [['Timestamp','Item Name','Item Code','Qty','UOM','Issued To','Department','Issued By']];
    rows.forEach(function(r) {{
        var cells = r.querySelectorAll('td');
        var qtyuom = cells[3].textContent.trim().split(' ');
        data.push([
            cells[0].textContent.trim(), cells[1].textContent.trim(), cells[2].textContent.trim(),
            qtyuom[0]||'', qtyuom[1]||'',
            cells[4].textContent.trim(), cells[5].textContent.trim(), cells[6].textContent.trim()
        ]);
    }});
    var ws = XLSX.utils.aoa_to_sheet(data);
    ws['!cols'] = [{{wch:18}},{{wch:28}},{{wch:14}},{{wch:8}},{{wch:8}},{{wch:22}},{{wch:18}},{{wch:18}}];
    var wb = XLSX.utils.book_new();
    XLSX.utils.book_append_sheet(wb, ws, 'MIS Records');
    var fname = 'EIPL_MIS_Export' + (from?'_from_'+from:'') + (to?'_to_'+to:'') + '.xlsx';
    XLSX.writeFile(wb, fname);
}}
window.onload = function() {{ renderMISPage(); }};
</script>
</div>
</body></html>"""
    return HTMLResponse(html)


@app.get("/mis/download/{assignment_id}")
def mis_download(assignment_id: int, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    import os
    from fastapi.responses import FileResponse, Response
    a = db.query(models.MaterialAssignment).filter(models.MaterialAssignment.id == assignment_id).first()
    if not a or not a.mis_filename:
        raise HTTPException(status_code=404, detail="MIS not found")

    # Preferred: bytes stored in the shared DB (works on every device)
    if getattr(a, "mis_filedata", None):
        return Response(
            content=a.mis_filedata,
            media_type="application/octet-stream",
            headers={"Content-Disposition": f'attachment; filename="{a.mis_filename}"'}
        )

    # Fallback: legacy records whose bytes only live on this machine's disk
    mis_path = os.path.join("mis_uploads", a.mis_filename)
    if os.path.exists(mis_path):
        return FileResponse(path=mis_path, filename=a.mis_filename, media_type="application/octet-stream")

    raise HTTPException(status_code=404, detail="MIS file missing from server")


@app.get("/rts/list", response_class=HTMLResponse)
def rts_list(current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    """Return-to-Store Slip Download Centre — same UI as the GRN Centre."""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    returns = db.query(models.MaterialAssignment).filter(
        models.MaterialAssignment.is_return == True,
        models.MaterialAssignment.return_filename != None
    ).order_by(models.MaterialAssignment.timestamp.desc()).all()
    rows = ""
    for a in returns:
        ts = fmt_ist(a.timestamp)
        ts_iso = fmt_ist(a.timestamp, "%Y-%m-%d")
        item_name = a.item.name if a.item else "Unknown Item"
        item_code = a.item.item_code if a.item else "—"
        returned_by = a.issued_to or "—"
        recorded_by = a.mis_uploader.username if a.mis_uploader else (a.issued_by or "System")
        dept = a.department or "—"
        rows += f"""
        <tr class="border-b hover:bg-slate-50 text-xs rts-row"
            data-date="{ts_iso}" data-item="{item_name.lower()}" data-code="{item_code.lower()}"
            data-returnedby="{returned_by.lower()}" data-uploader="{recorded_by.lower()}">
            <td class="p-3 text-slate-500 font-mono whitespace-nowrap">{ts}</td>
            <td class="p-3 font-semibold text-slate-800">{item_name}</td>
            <td class="p-3 font-mono text-slate-500">{item_code}</td>
            <td class="p-3 text-center font-mono font-bold text-slate-700">{a.quantity} {a.uom or ''}</td>
            <td class="p-3 text-slate-600">{returned_by}</td>
            <td class="p-3 text-slate-500">{dept}</td>
            <td class="p-3 text-slate-500">{recorded_by}</td>
            <td class="p-3 text-right">
                <a href="/rts/download/{a.id}" class="bg-amber-500 hover:bg-amber-600 text-white font-bold text-[10px] px-2.5 py-1.5 rounded-lg transition-all">&#11015; Download</a>
            </td>
        </tr>"""
    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8"><title>RTS Downloads - EIPL</title>
<script src="https://cdn.tailwindcss.com"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
<script src="https://cdnjs.cloudflare.com/ajax/libs/xlsx/0.18.5/xlsx.full.min.js"></script>
<style>body{{font-family:'Inter',sans-serif;}}.filter-input{{font-size:11px;padding:4px 8px;border:1px solid #e2e8f0;border-radius:6px;background:#f8fafc;width:100%;box-sizing:border-box;}}
.filter-input:focus{{outline:none;border-color:#f59e0b;background:#fff;}}</style>
</head>
<body class="bg-slate-50 min-h-screen flex">
{build_sidebar(current_user, current_page='rts')}
<div class="flex-1 min-h-screen overflow-x-hidden p-6">
    <div class="max-w-7xl mx-auto">
        <div class="flex items-center justify-between mb-5">
            <div>
                <a href="/" class="text-amber-600 text-xs font-bold hover:underline">&#8592; Back to Dashboard</a>
                <h1 class="text-xl font-black text-slate-900 mt-1">&#128230; Return to Store Slip Download Centre</h1>
                <p class="text-xs text-slate-400 mt-0.5">Return-to-Store slips — uploaded material return records</p>
            </div>
            <div class="flex items-center gap-2">
                <button onclick="exportRTS()" class="bg-emerald-600 hover:bg-emerald-700 text-white font-bold text-xs px-4 py-2 rounded-lg transition-all flex items-center gap-1.5">
                    <i class="fa-solid fa-file-excel"></i> Export Excel
                </button>
                <a href="/grn/list" class="bg-indigo-50 hover:bg-indigo-100 text-indigo-700 font-bold text-xs px-4 py-2 rounded-lg transition-all border border-indigo-200">&#128196; GRN Centre</a>
                <a href="/mis/list" class="bg-rose-50 hover:bg-rose-100 text-rose-700 font-bold text-xs px-4 py-2 rounded-lg transition-all border border-rose-200">&#128203; MIS Centre</a>
            </div>
        </div>

        <!-- Filters Bar -->
        <div class="bg-white border border-slate-200 rounded-xl p-4 mb-4 shadow-sm">
            <div class="grid grid-cols-2 md:grid-cols-4 lg:grid-cols-6 gap-3 items-end">
                <div class="lg:col-span-2">
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">&#128269; Search</label>
                    <input type="text" id="rtsSearch" placeholder="Item, code, employee..." oninput="applyRTSFilters()" class="filter-input">
                </div>
                <div>
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">From Date</label>
                    <input type="date" id="rtsFrom" oninput="applyRTSFilters()" class="filter-input">
                </div>
                <div>
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">To Date</label>
                    <input type="date" id="rtsTo" oninput="applyRTSFilters()" class="filter-input">
                </div>
                <div>
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">Returned By</label>
                    <input type="text" id="rtsReturnedBy" placeholder="Filter employee..." oninput="applyRTSFilters()" class="filter-input">
                </div>
                <div>
                    <label class="block text-[10px] font-bold text-slate-500 uppercase tracking-wider mb-1">Recorded By</label>
                    <input type="text" id="rtsUploader" placeholder="Filter user..." oninput="applyRTSFilters()" class="filter-input">
                </div>
            </div>
            <div class="flex items-center justify-between mt-2.5">
                <span id="rtsCount" class="text-[11px] text-slate-400 font-mono">Showing all records</span>
                <div class="flex items-center gap-3">
                    <button onclick="applyRTSFilters()" class="bg-amber-500 hover:bg-amber-600 text-white text-[10px] font-bold px-3 py-1.5 rounded-lg transition-all flex items-center gap-1"><i class="fa-solid fa-filter"></i> Apply Filters</button>
                    <button onclick="clearRTSFilters()" class="text-[10px] text-rose-500 font-bold hover:underline">Clear Filters</button>
                </div>
            </div>
        </div>

        <div class="bg-white border border-slate-200 rounded-2xl overflow-hidden shadow-sm">
            <div class="overflow-x-auto">
            <table class="w-full text-left border-collapse" id="rtsTable">
                <thead>
                    <tr class="bg-slate-50 text-slate-500 font-semibold tracking-wider uppercase text-[10px] border-b border-slate-200">
                        <th class="p-3 whitespace-nowrap">Timestamp</th>
                        <th class="p-3">Item Name</th>
                        <th class="p-3">Item Code</th>
                        <th class="p-3 text-center">Qty / UOM</th>
                        <th class="p-3">Returned By</th>
                        <th class="p-3">Department</th>
                        <th class="p-3">Recorded By</th>
                        <th class="p-3 text-right pr-4">Action</th>
                    </tr>
                </thead>
                <tbody class="divide-y divide-slate-100" id="rtsTbody">{rows if rows else '<tr><td colspan="8" class="p-8 text-center text-slate-400 text-sm">No Return-to-Store slips uploaded yet.</td></tr>'}</tbody>
            </table>
            </div>
            <div class="flex items-center justify-between px-4 py-3 border-t border-slate-100 bg-slate-50/50">
                <span class="text-[11px] text-slate-400" id="rtsPageInfo"></span>
                <div class="flex items-center gap-1" id="rtsPagination"></div>
            </div>
        </div>
    </div>

<script>
var RTS_PAGE = 1, RTS_PER_PAGE = 20, RTS_FILTERED = null;
function getAllRTSRows() {{ return Array.from(document.querySelectorAll('#rtsTbody .rts-row')); }}
function applyRTSFilters() {{
    var q = (document.getElementById('rtsSearch').value||'').toLowerCase().trim();
    var from = document.getElementById('rtsFrom').value;
    var to = document.getElementById('rtsTo').value;
    var emp = (document.getElementById('rtsReturnedBy').value||'').toLowerCase().trim();
    var uploader = (document.getElementById('rtsUploader').value||'').toLowerCase().trim();
    RTS_FILTERED = getAllRTSRows().filter(function(r) {{
        var d = r.dataset.date||'';
        if(from && d < from) return false;
        if(to && d > to) return false;
        if(q && !r.dataset.item.includes(q) && !r.dataset.code.includes(q) && !r.dataset.returnedby.includes(q)) return false;
        if(emp && !r.dataset.returnedby.includes(emp)) return false;
        if(uploader && !r.dataset.uploader.includes(uploader)) return false;
        return true;
    }});
    RTS_PAGE = 1; renderRTSPage();
}}
function renderRTSPage() {{
    var rows = RTS_FILTERED !== null ? RTS_FILTERED : getAllRTSRows();
    var total = rows.length;
    var pages = Math.ceil(total / RTS_PER_PAGE) || 1;
    if(RTS_PAGE > pages) RTS_PAGE = pages;
    var start = (RTS_PAGE-1)*RTS_PER_PAGE, end = start+RTS_PER_PAGE;
    getAllRTSRows().forEach(function(r){{r.style.display='none';}});
    rows.slice(start,end).forEach(function(r){{r.style.display='';}});
    document.getElementById('rtsCount').textContent = 'Showing ' + Math.min(start+1,total) + '-' + Math.min(end,total) + ' of ' + total + ' records';
    document.getElementById('rtsPageInfo').textContent = 'Page ' + RTS_PAGE + ' of ' + pages;
    var pgDiv = document.getElementById('rtsPagination');
    pgDiv.innerHTML = '';
    if(pages <= 1) return;
    var mkBtn = function(label,pg,active) {{
        var b = document.createElement('button');
        b.textContent = label;
        b.className = 'px-2.5 py-1 rounded text-[11px] font-bold border ' + (active ? 'bg-amber-500 text-white border-amber-500' : 'bg-white text-slate-600 border-slate-200 hover:bg-slate-50');
        b.onclick = function(){{ RTS_PAGE=pg; renderRTSPage(); }};
        return b;
    }};
    pgDiv.appendChild(mkBtn('\u00ab',1,false));
    pgDiv.appendChild(mkBtn('\u2039',Math.max(1,RTS_PAGE-1),false));
    for(var p=Math.max(1,RTS_PAGE-2);p<=Math.min(pages,RTS_PAGE+2);p++) pgDiv.appendChild(mkBtn(p,p,p===RTS_PAGE));
    pgDiv.appendChild(mkBtn('\u203a',Math.min(pages,RTS_PAGE+1),false));
    pgDiv.appendChild(mkBtn('\u00bb',pages,false));
}}
function clearRTSFilters() {{
    ['rtsSearch','rtsFrom','rtsTo','rtsReturnedBy','rtsUploader'].forEach(function(id){{document.getElementById(id).value='';}});
    RTS_FILTERED=null; RTS_PAGE=1; renderRTSPage();
}}
function exportRTS() {{
    var from = document.getElementById('rtsFrom').value;
    var to = document.getElementById('rtsTo').value;
    var rows = RTS_FILTERED !== null ? RTS_FILTERED : getAllRTSRows();
    var data = [['Timestamp','Item Name','Item Code','Qty','UOM','Returned By','Department','Recorded By']];
    rows.forEach(function(r) {{
        var cells = r.querySelectorAll('td');
        var qty = cells[3].textContent.trim().split(' ');
        data.push([
            cells[0].textContent.trim(), cells[1].textContent.trim(), cells[2].textContent.trim(),
            qty[0]||'', qty.slice(1).join(' ')||'',
            cells[4].textContent.trim(), cells[5].textContent.trim(), cells[6].textContent.trim()
        ]);
    }});
    var ws = XLSX.utils.aoa_to_sheet(data);
    ws['!cols'] = [{{wch:18}},{{wch:28}},{{wch:14}},{{wch:8}},{{wch:8}},{{wch:22}},{{wch:18}},{{wch:18}}];
    var wb = XLSX.utils.book_new();
    XLSX.utils.book_append_sheet(wb, ws, 'RTS Records');
    var fname = 'EIPL_RTS_Export' + (from?'_from_'+from:'') + (to?'_to_'+to:'') + '.xlsx';
    XLSX.writeFile(wb, fname);
}}
window.onload = function() {{ renderRTSPage(); }};
</script>
</div>
</body></html>"""
    return HTMLResponse(html)


@app.get("/rts/download/{assignment_id}")
def rts_download(assignment_id: int, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    import os
    from fastapi.responses import FileResponse, Response
    a = db.query(models.MaterialAssignment).filter(models.MaterialAssignment.id == assignment_id).first()
    if not a or not a.return_filename:
        raise HTTPException(status_code=404, detail="Return-to-Store slip not found")

    # Preferred: bytes from shared DB
    if getattr(a, "return_filedata", None):
        return Response(
            content=a.return_filedata,
            media_type="application/octet-stream",
            headers={"Content-Disposition": f'attachment; filename="{a.return_filename}"'}
        )

    # Fallback: bytes on local disk
    rts_path = os.path.join("rts_uploads", a.return_filename)
    if os.path.exists(rts_path):
        return FileResponse(path=rts_path, filename=a.return_filename, media_type="application/octet-stream")

    raise HTTPException(status_code=404, detail="Return-to-Store slip file missing from server")


@app.post("/items/bulk-import")
async def items_bulk_import(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user or current_user.role != "Admin":
        return HTMLResponse("Access Denied", status_code=403)

    try:
        contents = await file.read()
        if not contents:
            return HTMLResponse("<script>alert('Error: Uploaded file is empty.'); window.location='/';</script>")
            
        try:
            decoded_content = contents.decode('utf-8-sig')
        except UnicodeDecodeError:
            try:
                decoded_content = contents.decode('cp1252')
            except UnicodeDecodeError:
                decoded_content = contents.decode('latin-1')

        clean_lines = decoded_content.splitlines()
        
        try:
            sample_data = "\n".join(clean_lines[:5])
            dialect = csv.Sniffer().sniff(sample_data)
            if dialect.delimiter not in [',', ';', '\t', '|']:
                dialect.delimiter = ','
            reader = csv.DictReader(clean_lines, dialect=dialect)
        except Exception:
            reader = csv.DictReader(clean_lines)

        added_count = 0
        updated_count = 0
        row_index = 0

        for row in reader:
            row_index += 1
            if not row:
                continue
                
            clean_row = {str(k).strip().lower(): str(v).strip() for k, v in row.items() if k is not None}

            name = clean_row.get("name") or clean_row.get("item name") or clean_row.get("item_name") or clean_row.get("product name")
            item_code = clean_row.get("item_code") or clean_row.get("code") or clean_row.get("item code") or clean_row.get("sku")

            if not name or not item_code:
                continue

            item_code = item_code.upper()

            try:
                stock_raw = clean_row.get("initial_stock") or clean_row.get("stock") or clean_row.get("current_stock") or clean_row.get("quantity") or "0"
                initial_stock = int(float(stock_raw))

                price_raw = clean_row.get("price") or clean_row.get("rate") or "0.0"
                price = float(price_raw)
            except ValueError:
                continue

            vendor = clean_row.get("vendor") or clean_row.get("supplier") or ""
            site = clean_row.get("site") or clean_row.get("storage_site") or ""

            exists = db.query(models.Item).filter(models.Item.item_code == item_code).first()
            if exists:
                exists.name = name
                exists.current_stock = initial_stock
                exists.price = price
                if hasattr(exists, 'supplier'): exists.supplier = vendor
                if hasattr(exists, 'storage_site'): exists.storage_site = site
                updated_count += 1
            else:
                new_item = models.Item(
                    name=name,
                    item_code=item_code,
                    current_stock=initial_stock,
                    price=price,
                    minimum_stock=5
                )
                if hasattr(new_item, 'supplier'): new_item.supplier = vendor
                if hasattr(new_item, 'storage_site'): new_item.storage_site = site
                
                db.add(new_item)
                added_count += 1

        db.commit()
        return HTMLResponse(f"<script>alert('Bulk Process Complete! Added: {added_count}, Updated: {updated_count}'); window.location='/';</script>")
        
    except Exception as e:
        db.rollback()
        print(f"[Bulk Import Critical Exception]: {str(e)}")
        return HTMLResponse(f"<script>alert('Import processing error: {str(e)}'); window.location='/';</script>")


@app.get("/items/bulk-template")
def items_bulk_template(current_user: models.User = Depends(get_current_user)):
    if not current_user or current_user.role != "Admin":
        return RedirectResponse(url="/login", status_code=303)
    from fastapi.responses import StreamingResponse
    import io
    csv_content = "name,item_code,initial_stock,price,vendor,site\n"
    csv_content += "Steel Pipe 25mm,EIPL-ST-01,50,350.00,Tata Steel,Udaipur Yard\n"
    csv_content += "Copper Cable 4mm,EIPL-CC-02,100,125.50,Polycab,Store Room A\n"
    return StreamingResponse(
        io.StringIO(csv_content),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=EIPL_Items_Bulk_Template.csv"}
    )


@app.post("/items/add")
def add_item(
    name: str = Form(...),
    item_code: str = Form(...),
    price: float = Form(0.0),
    initial_stock: int = Form(0),
    minimum_stock: int = Form(0),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    existing_item = db.query(models.Item).filter(models.Item.item_code == item_code).first()
    if existing_item:
        raise HTTPException(status_code=400, detail="Item code already exists!")

    db_item = models.Item(
        name=name,
        item_code=item_code,
        price=price,
        current_stock=initial_stock,
        minimum_stock=minimum_stock
    )
    db.add(db_item)
    db.flush()

    if initial_stock > 0:
        item_unit_price = price if price is not None else 0.0
        initial_transaction = models.Transaction(
            item_id=db_item.id,
            type="IN",
            quantity=initial_stock,
            user_id=current_user.id,
            total_value=float(initial_stock * item_unit_price)
        )
        db.add(initial_transaction)

    db.commit()
    return RedirectResponse(url="/", status_code=303)


@app.post("/items/edit/{item_id}")
def edit_item(
    item_id: int,
    name: str = Form(...),
    item_code: str = Form(...),
    current_stock: int = Form(...),
    minimum_stock: int = Form(...),
    price: float = Form(...),
    supplier: str = Form(""),
    storage_site: str = Form(""),
    category: str = Form(None),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if current_user.role != "Admin":
        return HTMLResponse("<html><body><h2>Access Denied: Admin privileges required</h2></body></html>", status_code=403)
        
    db_item = db.query(models.Item).filter(models.Item.id == item_id).first()
    if not db_item:
        raise HTTPException(status_code=404, detail="Item not found")

    # Check if the new item_code is already used by a DIFFERENT item
    duplicate = db.query(models.Item).filter(
        models.Item.item_code == item_code.strip().upper(),
        models.Item.id != item_id
    ).first()
    if duplicate:
        return HTMLResponse(
            f"<script>alert('Error: Item code \"{item_code.upper()}\" is already assigned to another item. Please use a unique code.'); window.history.back();</script>"
        )

    db_item.name = name
    db_item.item_code = item_code.strip().upper()
    stock_delta = current_stock - (db_item.current_stock or 0)
    db_item.minimum_stock = minimum_stock
    db_item.price = price
    if supplier.strip():
        db_item.supplier = supplier.strip()
    if storage_site.strip():
        db_item.storage_site = storage_site.strip()
    if category and category.strip() in ASSET_CLASSES:
        db_item.category = category.strip()
    # current_stock is updated by reconcile_stock_delta via the lot ledger
    reconcile_stock_delta(db, db_item, stock_delta)
    db.commit()
    return RedirectResponse(url="/", status_code=303)


@app.post("/items/delete/{item_id}")
def delete_item(item_id: int, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not current_user or current_user.role != "Admin":
        return HTMLResponse("Access Denied", status_code=403)
    try:
        # Nullify FK back-references on AssetUnits first
        db.query(models.AssetUnit).filter(models.AssetUnit.item_id == item_id).update(
            {"current_assignment_id": None}, synchronize_session=False
        )
        db.flush()
        # Delete all child rows in safe dependency order
        db.query(models.GRNRecord).filter(models.GRNRecord.item_id == item_id).delete(synchronize_session=False)
        db.query(models.MaterialAssignment).filter(models.MaterialAssignment.item_id == item_id).delete(synchronize_session=False)
        db.query(models.AssetUnit).filter(models.AssetUnit.item_id == item_id).delete(synchronize_session=False)
        db.query(models.ItemLot).filter(models.ItemLot.item_id == item_id).delete(synchronize_session=False)
        db.query(models.Transaction).filter(models.Transaction.item_id == item_id).delete(synchronize_session=False)
        # Procurement: delete messages first (FK -> procurement_requests.id), then the requests
        proc_ids = [r[0] for r in db.query(models.ProcurementRequest.id).filter(
            models.ProcurementRequest.item_id == item_id
        ).all()]
        if proc_ids:
            db.query(models.ProcurementMessage).filter(
                models.ProcurementMessage.request_id.in_(proc_ids)
            ).delete(synchronize_session=False)
        db.query(models.ProcurementRequest).filter(models.ProcurementRequest.item_id == item_id).delete(synchronize_session=False)
        db.query(models.MaterialRequest).filter(models.MaterialRequest.item_id == item_id).delete(synchronize_session=False)
        db.query(models.Item).filter(models.Item.id == item_id).delete(synchronize_session=False)
        db.commit()
    except Exception as exc:
        db.rollback()
        safe_msg = str(exc).replace("'", "").replace('"', '')[:200]
        return HTMLResponse(
            f"<script>alert('Delete failed: {safe_msg}'); window.history.back();</script>",
            status_code=200
        )
    return RedirectResponse(url="/", status_code=303)


@app.post("/items/update-stock-direct/{item_id}")
def update_stock_direct(
    item_id: int,
    new_stock: int = Form(...),
    session_user: str = Cookie(None),
    db: Session = Depends(get_db)
):
    if not session_user:
        return RedirectResponse(url="/login", status_code=303)

    item = db.query(models.Item).filter(models.Item.id == item_id).first()
    if item:
        delta = int(new_stock) - (item.current_stock or 0)
        reconcile_stock_delta(db, item, delta)
        db.commit()
        
    return RedirectResponse(url="/", status_code=303)


@app.post("/material/issue")
async def issue_materials(
    item_id: int = Form(...),
    quantity: int = Form(...),
    uom: str = Form(...),
    issued_to: str = Form(...),
    issued_by: str = Form(None),
    department: str = Form("General Operations"),
    remarks: str = Form(None),
    asset_unit_id: int = Form(None),     # required when issuing a Fixed Asset / Equipment unit
    mis_file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)

    if not issued_by:
        issued_by = current_user.full_name or current_user.username

    if not mis_file or not mis_file.filename:
        return HTMLResponse("<script>alert('MIS (Material Issue Slip) upload is mandatory before recording an issue.'); window.history.back();</script>")

    item = db.query(models.Item).filter(models.Item.id == item_id).first()
    if not item:
        return RedirectResponse(url="/?error=item_not_found", status_code=303)

    tracked = is_tracked(item)
    asset_unit = None
    if tracked:
        if not asset_unit_id:
            return HTMLResponse("<script>alert('Select a specific serial number to issue this Fixed Asset / Equipment.'); window.history.back();</script>")
        asset_unit = db.query(models.AssetUnit).filter(
            models.AssetUnit.id == asset_unit_id,
            models.AssetUnit.item_id == item_id
        ).first()
        if not asset_unit:
            return HTMLResponse("<script>alert('Selected asset unit not found.'); window.history.back();</script>")
        if asset_unit.status != "In Stock":
            return HTMLResponse(f"<script>alert('Serial {asset_unit.serial_no} is currently {asset_unit.status}, not In Stock. Refresh and re-select.'); window.history.back();</script>")
        quantity = 1
    else:
        if (item.current_stock or 0) < quantity:
            return RedirectResponse(url="/?error=insufficient_stock", status_code=303)

    import os, datetime as dt
    mis_dir = "mis_uploads"
    os.makedirs(mis_dir, exist_ok=True)
    timestamp_str = dt.datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_filename = f"MIS_{item.item_code}_{timestamp_str}_{mis_file.filename.replace(' ', '_')}"
    mis_path = os.path.join(mis_dir, safe_filename)
    contents = await mis_file.read()
    with open(mis_path, "wb") as f:
        f.write(contents)

    if tracked:
        fifo_cost = float(item.price or 0.0)
        item.current_stock = (item.current_stock or 0) - 1
    else:
        fifo_cost = consume_fifo(db, item, quantity)

    now = dt.datetime.utcnow()
    new_assignment = models.MaterialAssignment(
        item_id=item_id,
        quantity=quantity,
        uom=uom,
        issued_to=issued_to,
        issued_by=issued_by,
        department=department,
        remarks=remarks,
        custodian=issued_to,
        mis_filename=safe_filename,
        mis_filedata=contents,
        mis_uploaded_by_id=current_user.id,
        mis_upload_timestamp=now,
        timestamp=now,
        asset_unit_id=(asset_unit.id if tracked else None),
    )
    db.add(new_assignment)
    db.flush()

    if tracked:
        asset_unit.status = "Issued"
        asset_unit.current_holder = issued_to
        asset_unit.current_assignment_id = new_assignment.id

    db.add(models.Transaction(
        item_id=item_id,
        type="OUT",
        quantity=quantity,
        total_value=fifo_cost,
        user_id=current_user.id,
        timestamp=now
    ))

    db.commit()
    return RedirectResponse(url="/?tab=allocations", status_code=303)


@app.post("/material/return")
async def return_to_store(
    item_id: int = Form(...),
    quantity: int = Form(...),
    returned_by: str = Form(...),               # the employee returning the material
    department: str = Form("General Operations"),
    remarks: str = Form(None),
    asset_unit_id: int = Form(None),            # for tracked items: which serial is coming back
    return_file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user)
):
    """Record a return-to-store. For consumables/tools this adds quantity back via a
    return lot; for Fixed Assets and Equipments it flips the specific AssetUnit
    back to In Stock."""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)

    if not return_file or not return_file.filename:
        return HTMLResponse("<script>alert('Return-to-Store Slip is mandatory before recording a return.'); window.history.back();</script>")

    if quantity <= 0:
        return HTMLResponse("<script>alert('Return quantity must be greater than zero.'); window.history.back();</script>")

    item = db.query(models.Item).filter(models.Item.id == item_id).first()
    if not item:
        return HTMLResponse("<script>alert('Item not found.'); window.history.back();</script>")

    tracked = is_tracked(item)
    asset_unit = None

    if tracked:
        if not asset_unit_id:
            return HTMLResponse("<script>alert('Select the specific serial number being returned.'); window.history.back();</script>")
        asset_unit = db.query(models.AssetUnit).filter(
            models.AssetUnit.id == asset_unit_id,
            models.AssetUnit.item_id == item_id
        ).first()
        if not asset_unit:
            return HTMLResponse("<script>alert('Selected asset unit not found.'); window.history.back();</script>")
        if asset_unit.status != "Issued" or asset_unit.current_holder != returned_by:
            return HTMLResponse(f"<script>alert('Serial {asset_unit.serial_no} is not currently held by {returned_by}.'); window.history.back();</script>")
        quantity = 1
    else:
        # Net issued to this employee = sum issues - sum prior returns
        issued_total = db.query(models.MaterialAssignment).filter(
            models.MaterialAssignment.item_id == item_id,
            models.MaterialAssignment.issued_to == returned_by,
            (models.MaterialAssignment.is_return == False) | (models.MaterialAssignment.is_return.is_(None))
        ).all()
        returned_total = db.query(models.MaterialAssignment).filter(
            models.MaterialAssignment.item_id == item_id,
            models.MaterialAssignment.issued_to == returned_by,
            models.MaterialAssignment.is_return == True
        ).all()
        net_held = sum(a.quantity or 0 for a in issued_total) - sum(a.quantity or 0 for a in returned_total)
        if quantity > net_held:
            return HTMLResponse(f"<script>alert('Cannot return {quantity} units. {returned_by} currently holds only {net_held}.'); window.history.back();</script>")

    import os, datetime as dt
    rts_dir = "rts_uploads"
    os.makedirs(rts_dir, exist_ok=True)
    timestamp_str = dt.datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    safe_filename = f"RTS_{item.item_code}_{timestamp_str}_{return_file.filename.replace(' ', '_')}"
    rts_path = os.path.join(rts_dir, safe_filename)
    contents = await return_file.read()
    with open(rts_path, "wb") as f:
        f.write(contents)

    if tracked:
        asset_unit.status = "In Stock"
        asset_unit.current_holder = None
        asset_unit.current_assignment_id = None
        item.current_stock = (item.current_stock or 0) + 1
    else:
        return_vendor = f"Returned by {returned_by}"
        add_to_lot(db, item, return_vendor, 0.0, quantity)

    now = dt.datetime.utcnow()
    db.add(models.MaterialAssignment(
        item_id=item_id,
        quantity=quantity,
        uom=item.uom or "Nos",
        issued_to=returned_by,
        issued_by=current_user.full_name or current_user.username,
        department=department,
        remarks=remarks,
        custodian=returned_by,
        is_return=True,
        return_filename=safe_filename,
        return_filedata=contents,
        mis_uploaded_by_id=current_user.id,
        mis_upload_timestamp=now,
        timestamp=now,
        asset_unit_id=(asset_unit.id if tracked else None),
    ))

    db.add(models.Transaction(
        item_id=item_id,
        type="RETURN",
        quantity=quantity,
        total_value=0.0,
        user_id=current_user.id,
        timestamp=now
    ))

    db.commit()
    return RedirectResponse(url="/?tab=allocations", status_code=303)
@app.post("/procurement/request")
def create_procurement_request(
    item_id: str = Form(...),
    quantity: int = Form(...),
    uom: str = Form("Nos"),
    department: str = Form(...),
    new_item_name: str = Form(None),
    detailed_specification: str = Form(None),
    category: str = Form(None),
    db: Session = Depends(get_db),
    user_str: str = Cookie(None, alias="session_user")
):
    if not user_str:
        return RedirectResponse(url="/login", status_code=303)

    current_user = db.query(models.User).filter(models.User.username == user_str).first()

    # The category chosen on the form only applies to brand-new ad-hoc items.
    # For existing catalog items we must NEVER overwrite their asset class from
    # the indent form — the item keeps whatever class it already has (a
    # consumable stays a consumable) until it's changed manually in Edit Item.
    clean_category = category.strip() if category and category.strip() in ASSET_CLASSES else DEFAULT_ASSET_CLASS

    new_request = models.ProcurementRequest(
        quantity=quantity,
        uom=(uom or "Nos").strip(),
        department=department.strip(),
        requested_by_id=current_user.id,
        status="Pending",
    )

    # Vendor and unit price are NOT captured at indent time anymore.
    # The admin enters them later in the Procurement Pipeline
    # (Order Pending -> "Save" vendor & rate via /procurement/set-pricing).

    if item_id == "NEW_PROCUREMENT_AD_HOC":
        new_request.is_new_item = True
        new_request.item_id = None
        new_request.new_item_name = new_item_name.strip() if new_item_name else "Unlisted Item"
        new_request.detailed_specification = detailed_specification.strip() if detailed_specification else "No specs"
        # Brand-new item: honour the class the requester picked on the form.
        new_request.category = clean_category
        # No rate yet — priced in the pipeline before ordering.
        new_request.total_estimated_cost = 0.0
    else:
        item_ent = db.query(models.Item).filter(models.Item.id == int(item_id)).first()
        if not item_ent:
            return HTMLResponse("<h2>Error: Item not found in Catalog</h2>", status_code=400)

        new_request.is_new_item = False
        new_request.item_id = item_ent.id
        # Provisional estimate from the catalog's last known price; the final
        # rate (and vendor) is set by the admin in the pipeline before ordering.
        new_request.total_estimated_cost = float(item_ent.price or 0.0) * quantity
        # Inherit the item's EXISTING asset class — do not change the item itself.
        new_request.category = item_ent.category or DEFAULT_ASSET_CLASS

    db.add(new_request)
    db.commit()
    return RedirectResponse(url="/#requisitions-panel", status_code=303)


@app.get("/procurement/check-code")
def check_item_code(
    code: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Returns JSON indicating whether the given item code already exists in catalog."""
    from fastapi.responses import JSONResponse
    if not current_user or current_user.role != "Admin":
        return JSONResponse({"error": "Unauthorized"}, status_code=403)
    clean_code = code.strip().upper()
    duplicate = db.query(models.Item).filter(models.Item.item_code == clean_code).first()
    if duplicate:
        return JSONResponse({"exists": True, "item_name": duplicate.name, "item_code": duplicate.item_code})
    return JSONResponse({"exists": False})


@app.post("/procurement/assign-code/{req_id}")
def assign_item_code(
    req_id: int,
    new_item_code: str = Form(...),
    force: str = Form(default="false"),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user or current_user.role != "Admin":
        return HTMLResponse("Unauthorized", status_code=403)

    req = db.query(models.ProcurementRequest).filter(models.ProcurementRequest.id == req_id).first()
    if not req:
        return HTMLResponse("<script>alert('Request not found.'); window.history.back();</script>")

    clean_code = new_item_code.strip().upper()

    # Check uniqueness against existing catalog items — skip if admin confirmed via force flag
    if force.lower() != "true":
        duplicate = db.query(models.Item).filter(models.Item.item_code == clean_code).first()
        if duplicate:
            return HTMLResponse(
                f"<script>alert('Error: Code \"{clean_code}\" already exists in catalog for \"{duplicate.name}\". Please use a unique code.'); window.history.back();</script>"
            )

    # Store assigned code in the supplier field of the pending request (reuse unused nullable field)
    # We prefix it clearly so Accept logic can detect and extract it
    req.new_item_name = f"[CODE:{clean_code}]{req.new_item_name.split(']', 1)[-1] if ']' in (req.new_item_name or '') else (req.new_item_name or '')}"
    db.commit()
    return RedirectResponse(url="/#requisitions-panel", status_code=303)



@app.post("/procurement/assign-spec/{req_id}")
def assign_specification(
    req_id: int,
    specification: str = Form(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)

    req = db.query(models.ProcurementRequest).filter(models.ProcurementRequest.id == req_id).first()
    if not req:
        return HTMLResponse("<script>alert('Request not found.'); window.history.back();</script>")

    clean_spec = specification.strip()
    if not clean_spec:
        return HTMLResponse("<script>alert('Specification cannot be empty.'); window.history.back();</script>")

    req.detailed_specification = clean_spec
    db.commit()
    return RedirectResponse(url="/#requisitions-panel", status_code=303)



@app.post("/procurement/message/{req_id}")
def post_procurement_message(
    req_id: int,
    message: str = Form(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    req = db.query(models.ProcurementRequest).filter(models.ProcurementRequest.id == req_id).first()
    if not req:
        return HTMLResponse("<script>alert('Request not found.'); window.history.back();</script>")
    msg = models.ProcurementMessage(
        request_id=req_id,
        sender_id=current_user.id,
        message=message.strip()
    )
    db.add(msg)
    db.commit()
    return RedirectResponse(url="/#requisitions-panel", status_code=303)


@app.get("/procurement/bulk-template")
def procurement_bulk_template(current_user: models.User = Depends(get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    from fastapi.responses import StreamingResponse
    import io
    csv_content = "item_id_or_NEW,item_name,specification,quantity,uom,department\n"
    csv_content += "1,,(leave blank for existing items),5,Nos,Operations\n"
    csv_content += "NEW,Steel Pipe 25mm,Grade A - 25mm dia x 6m length - IS 1239,10,Mtr,Mechanical\n"
    return StreamingResponse(
        io.StringIO(csv_content),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=EIPL_Procurement_Indent_Template.csv"}
    )


@app.post("/procurement/bulk-import")
async def procurement_bulk_import(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    try:
        contents = await file.read()
        try:
            decoded = contents.decode('utf-8-sig')
        except UnicodeDecodeError:
            decoded = contents.decode('latin-1')
        lines = decoded.splitlines()
        reader = csv.DictReader(lines)
        added = 0
        for row in reader:
            clean = {str(k).strip().lower(): str(v).strip() for k, v in row.items() if k}
            item_ref = clean.get("item_id_or_new", "").upper()
            quantity_raw = clean.get("quantity", "1")
            department = clean.get("department", "General Operations")
            uom = clean.get("uom", "Nos").strip() or "Nos"
            try:
                quantity = int(float(quantity_raw))
            except ValueError:
                continue
            if quantity < 1:
                continue
            new_req = models.ProcurementRequest(
                quantity=quantity,
                uom=uom,
                department=department.strip(),
                requested_by_id=current_user.id,
                status="Pending",
                total_estimated_cost=0.0
            )
            if item_ref == "NEW":
                new_req.is_new_item = True
                new_req.item_id = None
                new_req.new_item_name = clean.get("item_name", "Unlisted Item").strip() or "Unlisted Item"
                new_req.detailed_specification = clean.get("specification", "No specs").strip() or "No specs"
            else:
                try:
                    item_id = int(item_ref)
                    item_ent = db.query(models.Item).filter(models.Item.id == item_id).first()
                    if not item_ent:
                        continue
                    new_req.is_new_item = False
                    new_req.item_id = item_ent.id
                    new_req.total_estimated_cost = float(item_ent.price * quantity)
                except ValueError:
                    continue
            db.add(new_req)
            added += 1
        db.commit()
        return HTMLResponse(f"<script>alert('Bulk Import Complete! {added} indent(s) created.'); window.location='/#requisitions-panel';</script>")
    except Exception as e:
        db.rollback()
        return HTMLResponse(f"<script>alert('Import error: {str(e)}'); window.history.back();</script>")


@app.get("/procurement/export-excel")
def procurement_export_excel(
    date_from: str = None,
    date_to: str = None,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Export procurement pipeline to Excel with optional date range filter."""
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    import io, datetime as _dt
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HTMLResponse("<script>alert('openpyxl not installed on server.'); window.history.back();</script>")

    reqs = db.query(models.ProcurementRequest).order_by(models.ProcurementRequest.timestamp.desc()).all()

    # Apply date filter
    if date_from:
        try:
            dt_from = _dt.datetime.strptime(date_from, "%Y-%m-%d")
            reqs = [r for r in reqs if r.timestamp and r.timestamp >= dt_from]
        except ValueError:
            pass
    if date_to:
        try:
            dt_to = _dt.datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59, second=59)
            reqs = [r for r in reqs if r.timestamp and r.timestamp <= dt_to]
        except ValueError:
            pass

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Procurement Pipeline"

    # Header styling
    hdr_fill = PatternFill("solid", fgColor="1E293B")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = ["#", "Date", "Item Name", "Item Code", "Qty", "Est. Value (₹)", "Vendor", "Unit Rate (₹)", "Department", "Requested By", "Status", "Specification"]
    col_widths = [4, 16, 30, 16, 8, 14, 22, 14, 18, 14, 14, 40]

    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.row_dimensions[1].height = 24

    # Status colors
    status_fills = {
        "Pending": PatternFill("solid", fgColor="FEF9C3"),
        "Order Pending": PatternFill("solid", fgColor="DBEAFE"),
        "Order Placed": PatternFill("solid", fgColor="E0E7FF"),
        "Delivered": PatternFill("solid", fgColor="D1FAE5"),
        "Rejected": PatternFill("solid", fgColor="FEE2E2"),
    }

    for ri, r in enumerate(reqs, 2):
        date_str = fmt_ist(r.timestamp)
        if getattr(r, 'is_new_item', False):
            raw_name = r.new_item_name or ""
            if raw_name.startswith("[CODE:"):
                item_code_str = raw_name[len("[CODE:"):raw_name.index("]")]
                item_name_str = raw_name[raw_name.index("]") + 1:].strip() or "New Item"
            else:
                item_name_str = raw_name or "Unlisted Item"
                item_code_str = "—"
        else:
            item_name_str = r.item.name if r.item else (r.new_item_name or "—")
            item_code_str = r.item.item_code if r.item else "—"

        row_data = [
            ri - 1,
            date_str,
            item_name_str,
            item_code_str,
            r.quantity or 0,
            round(r.total_estimated_cost or 0.0, 2),
            r.vendor or "—",
            round(r.unit_price or 0.0, 2) if r.unit_price else "—",
            r.department or "—",
            r.requester.username if r.requester else "—",
            r.status or "—",
            r.detailed_specification or "—"
        ]
        sfill = status_fills.get(r.status)
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="center" if ci in [1,4,5,6,8] else "left", vertical="center", wrap_text=(ci == 12))
            if sfill:
                cell.fill = sfill
        ws.row_dimensions[ri].height = 15

    # Freeze header row
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    suffix = ""
    if date_from: suffix += f"_from_{date_from}"
    if date_to: suffix += f"_to_{date_to}"
    filename = f"EIPL_Procurement_Pipeline{suffix}.xlsx"

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.post("/admin/verify-password")
def admin_verify_password(
    password: str = Form(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """Returns JSON 200 if the supplied password matches ANY admin account.
    Always returns JSON — never redirects — so the frontend fetch never
    receives an HTML response it can't parse."""
    from fastapi.responses import JSONResponse
    # Session may be missing if cookie wasn't sent — give a clear message
    if not current_user:
        return JSONResponse({"ok": False, "msg": "Session expired. Please re-login and try again."}, status_code=401)
    if current_user.role != "Admin":
        return JSONResponse({"ok": False, "msg": "Admin privileges required."}, status_code=403)
    hashed = hash_password(password)
    # Accept password of ANY admin user (not just the session owner)
    matching_admin = db.query(models.User).filter(
        models.User.role == "Admin",
        models.User.hashed_password == hashed
    ).first()
    if not matching_admin:
        return JSONResponse({"ok": False, "msg": "Incorrect admin password."}, status_code=403)
    return JSONResponse({"ok": True})


@app.post("/procurement/set-pricing/{req_id}")
def set_procurement_pricing(
    req_id: int,
    vendor: str = Form(...),
    unit_price: float = Form(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    # Only the admin may enter vendor + rate, and only before the order is placed
    if not current_user or current_user.role != "Admin":
        return HTMLResponse("Unauthorized", status_code=403)
    req = db.query(models.ProcurementRequest).filter(models.ProcurementRequest.id == req_id).first()
    if req and req.status == "Order Pending":
        v = (vendor or "").strip()
        if not v or unit_price is None or unit_price < 0:
            return HTMLResponse("<script>alert('Enter a valid vendor and rate.'); window.history.back();</script>")
        req.vendor = v
        req.unit_price = float(unit_price)
        req.total_estimated_cost = float((req.quantity or 0) * float(unit_price))
        # Mirror onto the catalog item so it shows as the latest known vendor/price
        if req.item:
            req.item.supplier = v
            req.item.price = float(unit_price)
        db.commit()
    return RedirectResponse(url="/?tab=requisitions", status_code=303)


@app.post("/procurement/order/{req_id}")
def place_order(req_id: int, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not current_user or current_user.role != "Admin":
        return HTMLResponse("Unauthorized", status_code=403)
    req = db.query(models.ProcurementRequest).filter(models.ProcurementRequest.id == req_id).first()
    if req and req.status == "Order Pending":
        # The order can only be placed once the admin has entered vendor + rate
        if not (req.vendor and req.vendor.strip()) or req.unit_price is None:
            return HTMLResponse("<script>alert('Enter vendor and rate before placing the order.'); window.history.back();</script>")
        req.status = "Order Placed"
        db.commit()
    return RedirectResponse(url="/?tab=requisitions", status_code=303)


@app.post("/procurement/action/{req_id}/{action_token}")
def handle_procurement_action(
    req_id: int,
    action_token: str,
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user or current_user.role != "Admin":
        return HTMLResponse("Unauthorized", status_code=403)

    req = db.query(models.ProcurementRequest).filter(models.ProcurementRequest.id == req_id).first()
    if req:
        if action_token.lower() == "accept":
            # Block Accept for new items that haven't had a code assigned yet
            if getattr(req, 'is_new_item', False) and req.item_id is None:
                raw_name = req.new_item_name or ""
                if not raw_name.startswith("[CODE:"):
                    return HTMLResponse(
                        "<script>alert('Cannot approve: Please assign a unique Item Code to this new item request before approving.'); window.history.back();</script>"
                    )
                # Check specification is filled
                raw_spec = req.detailed_specification or ""
                if not raw_spec.strip() or raw_spec.strip().lower() == "no specs":
                    return HTMLResponse(
                        "<script>alert('Cannot approve: Please add the item specification before approving.'); window.history.back();</script>"
                    )
                # Extract the assigned code and the real item name
                code_part = raw_name[len("[CODE:"):raw_name.index("]")]
                real_name = raw_name[raw_name.index("]") + 1:].strip() or "New Procured Item"
                assigned_code = code_part.strip().upper()
            
            req.status = "Order Pending"

            # For new ad-hoc items: create catalog entry with 0 stock (delivered via inward later)
            if getattr(req, 'is_new_item', False) and req.item_id is None:
                new_catalog_item = models.Item(
                    item_code=assigned_code,
                    name=real_name,
                    description=req.detailed_specification,
                    category=(req.category if getattr(req, 'category', None) in ASSET_CLASSES else DEFAULT_ASSET_CLASS),
                    supplier="Pending Selection",
                    storage_site=req.department,
                    price=0.0,
                    current_stock=0,
                    minimum_stock=0
                )
                db.add(new_catalog_item)
                db.flush()
                req.item_id = new_catalog_item.id
            # Stock NOT added here — added when inward transaction is recorded
        else:
            req.status = "Rejected"
        db.commit()
    return RedirectResponse(url="/", status_code=303)


@app.post("/procurement/edit/{request_id}")
def edit_procurement_request(
    request_id: int,
    item_id: int = Form(...),
    quantity: int = Form(...),
    uom: str = Form("Nos"),
    department: str = Form(...),
    session_user: str = Cookie(None),
    db: Session = Depends(get_db)
):
    if not session_user:
        return RedirectResponse(url="/login", status_code=303)
        
    req = db.query(models.ProcurementRequest).filter(models.ProcurementRequest.id == request_id).first()
    if not req:
        return RedirectResponse(url="/", status_code=303)
        
    item = db.query(models.Item).filter(models.Item.id == item_id).first()
    if item:
        req.item_id = item_id
        req.quantity = quantity
        req.uom = (uom or "Nos").strip()
        req.total_estimated_cost = float(quantity * item.price)
        req.department = department
        db.commit()
        
    return RedirectResponse(url="/", status_code=303)


@app.post("/procurement/delete/{request_id}")
def delete_procurement_request(
    request_id: int,
    session_user: str = Cookie(None),
    db: Session = Depends(get_db)
):
    if not session_user:
        return RedirectResponse(url="/login", status_code=303)
    # Admin-only: non-admins cannot delete procurement requests
    user = db.query(models.User).filter(models.User.username == session_user).first()
    if not user or user.role != "Admin":
        return HTMLResponse("Access Denied: Admin privileges required to delete records.", status_code=403)
    req = db.query(models.ProcurementRequest).filter(models.ProcurementRequest.id == request_id).first()
    if req:
        try:
            # Clear messages first (FK -> procurement_requests.id)
            db.query(models.ProcurementMessage).filter(
                models.ProcurementMessage.request_id == request_id
            ).delete(synchronize_session=False)
            db.delete(req)
            db.commit()
        except Exception as exc:
            db.rollback()
            safe_msg = str(exc).replace("'", "").replace('"', '')[:200]
            return HTMLResponse(
                f"<script>alert('Delete failed: {safe_msg}'); window.history.back();</script>",
                status_code=200
            )
    return RedirectResponse(url="/", status_code=303)


@app.post("/admin/users/create")
def admin_create_user(
    username: str = Form(...),
    password: str = Form(...),
    full_name: str = Form(...),
    designation: str = Form(...),
    workstation_location: str = Form(...),
    role: str = Form(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user or current_user.role != "Admin":
        return HTMLResponse("Unauthorized", status_code=403)

    if db.query(models.User).filter(models.User.username == username.strip().lower()).first():
        return HTMLResponse("<script>alert('Conflict: Identifier taken.'); window.location='/';</script>")

    db.add(models.User(
        username=username.strip().lower(),
        hashed_password=hash_password(password.strip()),
        role=role.strip(),
        full_name=full_name.strip(),
        designation=designation.strip(),
        workstation_location=workstation_location.strip()
    ))
    db.commit()
    return RedirectResponse(url="/", status_code=303)


@app.post("/admin/users/edit/{target_id}")
def admin_edit_user(
    target_id: int,
    full_name: str = Form(...),
    designation: str = Form(...),
    workstation_location: str = Form(...),
    role: str = Form(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user or current_user.role != "Admin":
        return HTMLResponse("Unauthorized", status_code=403)

    user = db.query(models.User).filter(models.User.id == target_id).first()
    if user:
        user.full_name = full_name.strip()
        user.designation = designation.strip()
        user.workstation_location = workstation_location.strip()
        user.role = role.strip()
        db.commit()
    return RedirectResponse(url="/", status_code=303)


@app.get("/account/password", response_class=HTMLResponse)
def account_password_page(current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    import html as _html

    admin_reset_panel = ""
    if current_user.role == "Admin":
        staff_users = db.query(models.User).filter(models.User.role != "Admin").order_by(models.User.username).all()
        rows = ""
        for u in staff_users:
            rows += f"""
            <div class="flex items-center justify-between bg-slate-50 p-2.5 rounded-lg border border-slate-200 text-[11px] mb-1.5">
                <div>
                    <p class="font-bold text-slate-800">{_html.escape(u.username)} ({_html.escape(u.role)})</p>
                    <p class="text-slate-500 text-[10px]">{_html.escape(u.full_name or '')}</p>
                </div>
                <button type="button" onclick="toggleResetForm({u.id})" class="text-indigo-600 font-bold hover:underline text-[10px]">set password</button>
            </div>
            <form id="resetForm{u.id}" action="/admin/users/set-password/{u.id}" method="POST" class="hidden mb-3 p-2.5 bg-indigo-50/50 border border-indigo-100 rounded-lg space-y-2">
                <input type="password" name="new_password" placeholder="New password for {_html.escape(u.username)}" required minlength="4"
                    class="w-full bg-white border border-slate-200 p-2 rounded-lg text-xs">
                <button type="submit" class="w-full bg-indigo-600 hover:bg-indigo-700 text-white p-2 rounded-lg font-bold text-xs">Update Password</button>
            </form>"""
        admin_reset_panel = f"""
        <div class="bg-white border border-slate-200 rounded-2xl shadow-sm p-5 mt-6">
            <h2 class="text-xs font-black tracking-wider text-slate-800 uppercase border-b border-slate-100 pb-3 mb-3">&#128272; Reset Staff Password</h2>
            <p class="text-[10px] text-slate-400 mb-3">As Admin, you can set a new password for any Staff account. Admin accounts cannot reset each other's passwords here.</p>
            {rows if rows else '<p class="text-[11px] text-slate-400">No staff accounts found.</p>'}
        </div>
        <script>
        function toggleResetForm(id) {{
            var f = document.getElementById('resetForm' + id);
            f.classList.toggle('hidden');
        }}
        </script>
        """

    html_page = f"""<!DOCTYPE html><html lang="en"><head>
<meta charset="UTF-8"><title>Change Password - EIPL</title>
<script src="https://cdn.tailwindcss.com"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
<style>body{{font-family:'Inter',sans-serif;}}</style>
</head><body class="bg-slate-50 min-h-screen flex">
{build_sidebar(current_user, current_page='password')}
<div class="flex-1 flex flex-col min-h-screen overflow-x-hidden">
<div class="bg-white border-b border-slate-200 h-16 flex items-center justify-between px-8 shrink-0 shadow-sm z-10">
    <div class="flex items-center gap-2 text-xs font-semibold text-slate-400">
        <span class="uppercase tracking-wider text-indigo-600 font-bold">EIPL Framework</span>
        <i class="fa-solid fa-chevron-right text-[9px] text-slate-300"></i>
        <span class="text-slate-700 font-medium">Change Password</span>
    </div>
    <span class="text-[11px] text-slate-400 font-mono">{_html.escape(current_user.username)} ({current_user.role})</span>
</div>
<div class="max-w-xl w-full mx-auto p-6">
    <div class="bg-white border border-slate-200 rounded-2xl shadow-sm p-5">
        <h2 class="text-xs font-black tracking-wider text-slate-800 uppercase border-b border-slate-100 pb-3 mb-3">&#128273; Change My Password</h2>
        <form action="/account/change-password" method="POST" class="space-y-3 text-xs text-slate-700">
            <div><label class="block font-semibold text-slate-500 mb-1">Current Password</label>
                <input type="password" name="current_password" required class="w-full bg-slate-50 border border-slate-200 p-2.5 rounded-xl text-xs"></div>
            <div><label class="block font-semibold text-slate-500 mb-1">New Password</label>
                <input type="password" name="new_password" required minlength="4" class="w-full bg-slate-50 border border-slate-200 p-2.5 rounded-xl text-xs"></div>
            <div><label class="block font-semibold text-slate-500 mb-1">Confirm New Password</label>
                <input type="password" name="confirm_password" required minlength="4" class="w-full bg-slate-50 border border-slate-200 p-2.5 rounded-xl text-xs"></div>
            <button type="submit" class="w-full bg-indigo-600 hover:bg-indigo-700 text-white p-2.5 rounded-xl font-bold tracking-wide transition-all text-xs">Update Password</button>
        </form>
        <div class="mt-3 pt-3 border-t border-slate-100">
            <p class="text-[10px] text-slate-400 leading-relaxed">
                <i class="fa-solid fa-circle-info text-slate-400 mr-1"></i>
                <b>Forgot your current password?</b> Ask your EIPL administrator to issue a temporary password
                from the Grant User Access panel; you can then sign in and change it from here.
            </p>
        </div>
    </div>
    {admin_reset_panel}
</div>
</div>
</body></html>"""
    return HTMLResponse(html_page)


@app.post("/account/change-password")
def account_change_password(
    current_password: str = Form(...),
    new_password: str = Form(...),
    confirm_password: str = Form(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    if current_user.hashed_password != hash_password(current_password):
        return HTMLResponse("<script>alert('Current password is incorrect.'); window.history.back();</script>")
    if new_password != confirm_password:
        return HTMLResponse("<script>alert('New password and confirmation do not match.'); window.history.back();</script>")
    if len(new_password) < 4:
        return HTMLResponse("<script>alert('New password must be at least 4 characters.'); window.history.back();</script>")
    current_user.hashed_password = hash_password(new_password)
    db.commit()
    return HTMLResponse("<script>alert('Password updated successfully.'); window.location='/account/password';</script>")


@app.post("/admin/users/set-password/{target_id}")
def admin_set_user_password(
    target_id: int,
    new_password: str = Form(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user or current_user.role != "Admin":
        return HTMLResponse("Unauthorized", status_code=403)
    target = db.query(models.User).filter(models.User.id == target_id).first()
    if not target:
        return HTMLResponse("<script>alert('User not found.'); window.location='/account/password';</script>")
    if target.role == "Admin":
        return HTMLResponse("<script>alert('Admins cannot reset another admin\\'s password.'); window.location='/account/password';</script>")
    if len(new_password) < 4:
        return HTMLResponse("<script>alert('New password must be at least 4 characters.'); window.history.back();</script>")
    target.hashed_password = hash_password(new_password)
    db.commit()
    return HTMLResponse(f"<script>alert('Password updated for {target.username}.'); window.location='/account/password';</script>")


@app.post("/admin/users/delete/{target_id}")
def admin_delete_user(target_id: int, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not current_user or current_user.role != "Admin":
        return HTMLResponse("Unauthorized", status_code=403)
    if current_user.id == target_id:
        return HTMLResponse("<script>alert('Constraint blocked.'); window.location='/';</script>")
    try:
        # Preserve audit history: null FK references on dependent rows instead
        # of cascading deletes (we want to keep the transactions / GRNs / MIS records).
        db.query(models.ProcurementRequest).filter(
            models.ProcurementRequest.requested_by_id == target_id
        ).update({"requested_by_id": None}, synchronize_session=False)
        db.query(models.MaterialAssignment).filter(
            models.MaterialAssignment.mis_uploaded_by_id == target_id
        ).update({"mis_uploaded_by_id": None}, synchronize_session=False)
        db.query(models.Transaction).filter(
            models.Transaction.user_id == target_id
        ).update({"user_id": None}, synchronize_session=False)
        db.query(models.GRNRecord).filter(
            models.GRNRecord.uploaded_by_id == target_id
        ).update({"uploaded_by_id": None}, synchronize_session=False)
        db.query(models.MaterialRequest).filter(
            models.MaterialRequest.requested_by_id == target_id
        ).update({"requested_by_id": None}, synchronize_session=False)
        db.query(models.ProcurementMessage).filter(
            models.ProcurementMessage.sender_id == target_id
        ).update({"sender_id": None}, synchronize_session=False)
        db.query(models.User).filter(models.User.id == target_id).delete(synchronize_session=False)
        db.commit()
    except Exception as exc:
        db.rollback()
        safe_msg = str(exc).replace("'", "").replace('"', '')[:200]
        return HTMLResponse(
            f"<script>alert('Delete failed: {safe_msg}'); window.history.back();</script>",
            status_code=200
        )
    return RedirectResponse(url="/", status_code=303)


@app.post("/employees/add")
def add_employee(
    name: str = Form(...),
    role_title: str = Form(...),
    location: str = Form("Not Specified"),
    contact: str = Form("Not Specified"),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user or current_user.role != "Admin":
        return HTMLResponse("Access Denied", status_code=403)

    db.add(models.Employee(
        name=name.strip(),
        role_title=role_title.strip(),
        location=location.strip(),
        contact=contact.strip()
    ))
    db.commit()
    return RedirectResponse(url="/", status_code=303)


@app.post("/employees/edit/{emp_id}")
def edit_employee(
    emp_id: int,
    name: str = Form(...),
    role_title: str = Form(...),
    location: str = Form(...),
    contact: str = Form(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user or current_user.role != "Admin":
        return HTMLResponse("Access Denied", status_code=403)

    emp = db.query(models.Employee).filter(models.Employee.id == emp_id).first()
    if emp:
        emp.name = name.strip()
        emp.role_title = role_title.strip()
        emp.location = location.strip()
        emp.contact = contact.strip()
        db.commit()
    return RedirectResponse(url="/", status_code=303)


@app.post("/employees/delete/{emp_id}")
def delete_employee(emp_id: int, current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not current_user or current_user.role != "Admin":
        return HTMLResponse("Access Denied", status_code=403)

    db.query(models.Employee).filter(models.Employee.id == emp_id).delete()
    db.commit()
    return RedirectResponse(url="/", status_code=303)


def build_sidebar(current_user, current_page: str = "") -> str:
    """Returns the persistent left sidebar HTML for standalone pages.
    `current_page` keys: material-movement, grn, mis, rts, password."""
    import html as _html
    role_initial = (current_user.role or "S")[0].upper()
    # Style helpers for active vs inactive nav items
    nav_active = "w-full flex items-center gap-3 px-3 py-2.5 text-indigo-700 bg-indigo-50 border border-indigo-100/50 rounded-xl text-sm font-bold transition-all text-left"
    nav_idle   = "w-full flex items-center gap-3 px-3 py-2.5 text-slate-600 hover:bg-slate-50 hover:text-slate-900 rounded-xl text-sm font-medium transition-all group"
    sub_active = "w-full flex items-center gap-3 px-3 py-2 text-indigo-700 bg-indigo-50 border border-indigo-100/50 rounded-lg text-xs font-bold transition-all"
    sub_idle   = "w-full flex items-center gap-3 px-3 py-2 text-slate-600 hover:bg-slate-50 rounded-lg text-xs font-medium transition-all group"
    def cls(key, active_cls, idle_cls):
        return active_cls if current_page == key else idle_cls
    def icon_cls(key, active_color, idle_size="w-5"):
        # active variant uses solid color, idle uses muted slate with hover tint
        if current_page == key:
            return f"fa-solid {idle_size} text-center {active_color}"
        return f"fa-solid {idle_size} text-center text-slate-400 group-hover:{active_color}"
    return f"""
    <aside class="w-64 bg-white border-r border-slate-200 flex flex-col shrink-0 shadow-sm z-20 sticky top-0 h-screen">
        <div class="p-5 border-b border-slate-100 text-center">
            <img src="{LOGO_DATA_URL}" alt="EIPL Logo" class="h-16 mx-auto mb-2 object-contain"
                onerror="this.style.display='none';document.getElementById('sbLogoFallback').style.display='inline-block';">
            <div id="sbLogoFallback" style="display:none;" class="bg-indigo-600 px-3 py-2 rounded-xl text-white font-black text-sm mb-2">EIPL</div>
            <h2 class="font-black text-slate-900 tracking-tight text-[13px] leading-tight uppercase">Electra Infracon Pvt Ltd</h2>
            <p class="text-[10px] text-slate-400 font-mono mt-1">{_html.escape(current_user.username)}</p>
        </div>
        <nav class="flex-1 p-4 space-y-1 overflow-y-auto">
            <p class="text-[10px] font-bold text-slate-400 uppercase tracking-widest px-3 mb-2">Core Dashboard</p>
            <a href="/?tab=requisitions" class="{nav_idle}">
                <i class="fa-solid fa-file-invoice-dollar w-5 text-center text-slate-400 group-hover:text-indigo-600"></i> Requisitions
            </a>
            <a href="/material-movement" class="{cls('material-movement', nav_active, nav_idle)}">
                <i class="{icon_cls('material-movement', 'text-emerald-600')} fa-truck-ramp-box"></i> Material Movement
            </a>
            <a href="/" class="{nav_idle}">
                <i class="fa-solid fa-boxes-stacked w-5 text-center text-slate-400 group-hover:text-indigo-600"></i> Inventory Log
            </a>
            <a href="/?tab=allocations" class="{nav_idle}">
                <i class="fa-solid fa-list-check w-5 text-center text-slate-400 group-hover:text-indigo-600"></i> Allocation Log
            </a>
            <div class="pt-4 mt-4 border-t border-slate-100 space-y-1">
                <p class="text-[10px] font-bold text-slate-400 uppercase tracking-widest px-3 mb-2">Administration</p>
                <a href="/?open=employee" class="{sub_idle}">
                    <i class="fa-solid fa-user-plus w-4 text-center text-slate-400 group-hover:text-indigo-600"></i> Employee Registry
                </a>
                <a href="/?open=access" class="{sub_idle}">
                    <i class="fa-solid fa-key w-4 text-center text-slate-400 group-hover:text-emerald-600"></i> Grant User Access
                </a>
                <a href="/grn/list" class="{cls('grn', sub_active, sub_idle)}">
                    <i class="{icon_cls('grn', 'text-indigo-600', 'w-4')} fa-download"></i> GRN Download Centre
                </a>
                <a href="/mis/list" class="{cls('mis', sub_active, sub_idle)}">
                    <i class="{icon_cls('mis', 'text-indigo-600', 'w-4')} fa-file-arrow-down"></i> MIS Download Centre
                </a>
                <a href="/rts/list" class="{cls('rts', sub_active, sub_idle)}">
                    <i class="{icon_cls('rts', 'text-amber-600', 'w-4')} fa-arrow-rotate-left"></i> RTS Download Centre
                </a>
                <a href="/account/password" class="{cls('password', sub_active, sub_idle)}">
                    <i class="{icon_cls('password', 'text-indigo-600', 'w-4')} fa-lock"></i> Change Password
                </a>
            </div>
        </nav>
        <div class="p-4 border-t border-slate-100 bg-slate-50/50">
            <div class="flex items-center justify-between mb-2">
                <div class="flex items-center gap-2.5 min-w-0">
                    <div class="w-8 h-8 rounded-lg bg-indigo-100 flex items-center justify-center font-bold text-[11px] text-indigo-700 shrink-0">{role_initial}</div>
                    <div class="min-w-0">
                        <h4 class="text-xs font-bold text-slate-900 leading-none truncate">{_html.escape(current_user.full_name or current_user.username)}</h4>
                        <span class="text-[10px] text-slate-400 mt-0.5 inline-block truncate max-w-[120px]">{_html.escape(current_user.designation or '')}</span>
                    </div>
                </div>
                <a href="/logout" class="text-slate-400 hover:text-rose-600 transition-colors p-1.5 rounded-lg hover:bg-rose-50" title="Sign Out">
                    <i class="fa-solid fa-right-from-bracket text-xs"></i>
                </a>
            </div>
        </div>
    </aside>"""



@app.get("/rts/bulk-template")
def rts_bulk_template(current_user: models.User = Depends(get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    from fastapi.responses import StreamingResponse
    import io
    csv_content = "item_code,quantity,uom,returned_by,department,remarks\n"
    csv_content += "EIPL-ST-01,2,Nos,Piyush Bhatia,Udaipur Project,Unused — returned to store\n"
    csv_content += "EIPL-CC-02,5,Mtr,Biswajit Pradhan,Electrical Works,Excess cable returned\n"
    return StreamingResponse(
        io.StringIO(csv_content),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=EIPL_RTS_Bulk_Template.csv"}
    )


@app.get("/mis/bulk-template")
def mis_bulk_template(current_user: models.User = Depends(get_current_user)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    from fastapi.responses import StreamingResponse
    import io
    csv_content = "item_code,quantity,uom,issued_to,department,remarks\n"
    csv_content += "EIPL-ST-01,10,Nos,Piyush Bhatia,Udaipur Project,For site installation\n"
    csv_content += "EIPL-CC-02,50,Mtr,Biswajit Pradhan,Electrical Works,Cable run phase 2\n"
    return StreamingResponse(
        io.StringIO(csv_content),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=EIPL_MIS_Bulk_Template.csv"}
    )


@app.post("/mis/bulk-import")
async def mis_bulk_import(
    file: UploadFile = File(...),
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    try:
        contents = await file.read()
        try:
            decoded = contents.decode('utf-8-sig')
        except UnicodeDecodeError:
            decoded = contents.decode('latin-1')
        lines = [l for l in decoded.splitlines() if l.strip() and not l.strip().startswith('#')]
        reader = csv.DictReader(lines)
        added = 0
        errors = []
        import datetime as _dt
        for row in reader:
            clean = {str(k).strip().lower(): str(v).strip() for k, v in row.items() if k}
            item_code = clean.get("item_code", "").upper()
            quantity_raw = clean.get("quantity", "0")
            uom = clean.get("uom", "Nos")
            issued_to = clean.get("issued_to", "").strip()
            department = clean.get("department", "General Operations").strip()
            remarks = clean.get("remarks", "").strip()

            if not item_code or not issued_to:
                errors.append(f"Skipped row — missing item_code or issued_to")
                continue
            try:
                quantity = int(float(quantity_raw))
            except ValueError:
                errors.append(f"Skipped {item_code} — invalid quantity")
                continue
            if quantity < 1:
                errors.append(f"Skipped {item_code} — quantity must be >= 1")
                continue

            item = db.query(models.Item).filter(models.Item.item_code == item_code).first()
            if not item:
                errors.append(f"Skipped {item_code} — item not found in catalog")
                continue
            if item.current_stock < quantity:
                errors.append(f"Skipped {item_code} — insufficient stock ({item.current_stock} available)")
                continue

            bulk_fifo_cost = consume_fifo(db, item, quantity)
            now_ts = _dt.datetime.utcnow()
            db.add(models.MaterialAssignment(
                item_id=item.id,
                quantity=quantity,
                uom=uom or (item.uom or "Nos"),
                issued_to=issued_to,
                issued_by=current_user.full_name or current_user.username,
                department=department,
                remarks=remarks,
                custodian=issued_to,
                timestamp=now_ts
            ))
            # Record outward transaction for the flow dashboard
            db.add(models.Transaction(
                item_id=item.id,
                type="OUT",
                quantity=quantity,
                total_value=bulk_fifo_cost,
                user_id=current_user.id,
                timestamp=now_ts
            ))
            added += 1

        db.commit()
        msg = f"MIS Bulk Import: {added} issue(s) recorded."
        if errors:
            msg += " Errors: " + " | ".join(errors[:5])
        return HTMLResponse(f"<script>alert('{msg}'); window.location='/material-movement';</script>")
    except Exception as e:
        db.rollback()
        return HTMLResponse(f"<script>alert('MIS Bulk Import Error: {str(e)}'); window.history.back();</script>")


@app.get("/material-movement", response_class=HTMLResponse)
def material_movement_page(current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)
    items = db.query(models.Item).order_by(models.Item.name).all()
    employees = db.query(models.Employee).order_by(models.Employee.name).all()
    import json as _json, html as _html
    inv_json = _json.dumps([{
        "id": i.id, "name": i.name, "code": i.item_code, "stock": i.current_stock,
        "uom": getattr(i,'uom',None) or "", "price": i.price or 0.0,
        "vendor": getattr(i,'supplier','') or "",
        "category": getattr(i, 'category', None) or DEFAULT_ASSET_CLASS,
        "tracked": is_tracked(i),
    } for i in items])

    item_by_id = {i.id: i for i in items}

    # Tracked items: in-stock asset units for the Outward serial picker
    in_stock_units = {}
    for u in db.query(models.AssetUnit).filter(models.AssetUnit.status == "In Stock").all():
        in_stock_units.setdefault(u.item_id, []).append({"id": u.id, "serial_no": u.serial_no})
    units_by_item_json = _json.dumps(in_stock_units)

    # Held units per employee (tracked items only) — for Return picker
    held_units_map = {}
    for u in db.query(models.AssetUnit).filter(models.AssetUnit.status == "Issued").all():
        if not u.current_holder:
            continue
        it = item_by_id.get(u.item_id)
        if not it:
            continue
        held_units_map.setdefault(u.current_holder, []).append({
            "asset_unit_id": u.id, "serial_no": u.serial_no,
            "item_id": u.item_id, "item_name": it.name, "item_code": it.item_code,
            "uom": (it.uom or "Nos"),
            "category": getattr(it, 'category', None) or DEFAULT_ASSET_CLASS,
        })
    held_units_json = _json.dumps(held_units_map)

    # Consumable/tools net-held per (employee, item) — skip tracked items (handled above)
    held = {}
    all_assigns = db.query(models.MaterialAssignment).all()
    for a in all_assigns:
        if not a.issued_to or not a.item_id:
            continue
        it = item_by_id.get(a.item_id)
        if it and is_tracked(it):
            continue
        key = (a.issued_to, a.item_id)
        sign = -1 if getattr(a, "is_return", False) else +1
        held[key] = held.get(key, 0) + sign * (a.quantity or 0)
    held_map = {}
    for (emp, iid), qty in held.items():
        if qty <= 0:
            continue
        it = item_by_id.get(iid)
        if not it:
            continue
        held_map.setdefault(emp, []).append({
            "item_id": iid, "name": it.name, "code": it.item_code,
            "uom": (it.uom or "Nos"), "qty_held": qty,
            "category": getattr(it, 'category', None) or DEFAULT_ASSET_CLASS,
        })
    held_json = _json.dumps(held_map)

    emp_opts = "".join([f'<option value="{_html.escape(e.name)}">{_html.escape(e.name)} ({e.role_title})</option>' for e in employees])
    recv = _html.escape(current_user.full_name or current_user.username)
    html_page = f"""<!DOCTYPE html><html lang="en"><head>
<meta charset="UTF-8"><title>Material Movement - EIPL</title>
<script src="https://cdn.tailwindcss.com"></script>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
<link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
<style>body{{font-family:'Inter',sans-serif;}}</style>
</head><body class="bg-slate-50 min-h-screen flex">
{build_sidebar(current_user, current_page='material-movement')}
<div class="flex-1 flex flex-col min-h-screen overflow-x-hidden">
<div class="bg-white border-b border-slate-200 h-16 flex items-center justify-between px-8 shrink-0 shadow-sm z-10">
    <div class="flex items-center gap-2 text-xs font-semibold text-slate-400">
        <span class="uppercase tracking-wider text-indigo-600 font-bold">EIPL Framework</span>
        <i class="fa-solid fa-chevron-right text-[9px] text-slate-300"></i>
        <span class="text-slate-700 font-medium">Material Movement</span>
    </div>
    <span class="text-[11px] text-slate-400 font-mono">{_html.escape(current_user.username)} ({current_user.role})</span>
</div>
<div class="max-w-[1600px] mx-auto p-6">
  <div class="grid grid-cols-1 xl:grid-cols-3 gap-6 items-start">

  <!-- LEFT: INWARD -->
  <div class="bg-white border border-slate-200 rounded-2xl shadow-sm overflow-hidden">
    <div class="flex items-center justify-between px-5 py-3 border-b-2 border-emerald-400 bg-emerald-50/60">
      <div>
        <div class="flex items-center gap-2 mb-0.5">
          <span class="bg-emerald-500 text-white text-[10px] font-black px-2 py-0.5 rounded tracking-widest uppercase">Inward</span>
          <h2 class="text-xs font-black tracking-wider text-slate-800 uppercase">&#8595; Material Inward</h2>
        </div>
        <p class="text-[10px] text-slate-400">Search existing item or register new — GRN mandatory</p>
      </div>
      <a href="/inward/bulk-template" class="bg-emerald-50 hover:bg-emerald-100 text-emerald-700 border border-emerald-300 px-2 py-1 rounded-lg text-[10px] font-bold flex items-center gap-1 whitespace-nowrap"><i class="fa-solid fa-download"></i> Template</a>
    </div>
    <form action="/transaction" method="POST" enctype="multipart/form-data" class="p-5 space-y-3 text-xs text-slate-700">
      <div>
        <label class="block font-semibold text-slate-600 mb-1">Search &amp; Select Item</label>
        <div class="relative">
          <input type="text" id="inward_item_search" placeholder="Type item name or code..." autocomplete="off"
            class="w-full bg-slate-50 border border-slate-200 text-slate-900 p-2.5 rounded-xl focus:outline-none focus:border-emerald-500 focus:bg-white shadow-sm text-xs"
            oninput="filterInwardItems(this.value)" onfocus="showInwardDropdown()">
          <div id="inward_item_dropdown" class="absolute z-30 w-full bg-white border border-slate-200 rounded-xl shadow-xl mt-1 max-h-48 overflow-y-auto hidden"></div>
        </div>
        <input type="hidden" id="inward_item_id" name="item_id" value="" required>
        <p id="inward_selected_label" class="text-[10px] text-slate-400 mt-1 italic">No item selected</p>
      </div>
      <div>
        <label class="block font-semibold text-slate-500 mb-1">Category</label>
        <input type="text" id="inward_category_display" readonly placeholder="\u2014 (auto-filled)"
          class="w-full bg-slate-100 border border-slate-200 p-2 rounded-lg text-xs text-slate-500 cursor-not-allowed">
      </div>
      <div class="grid grid-cols-2 gap-2">
        <div><label class="block font-semibold text-slate-500 mb-1">Quantity Received</label>
          <input type="number" name="quantity" min="1" value="1" required class="w-full bg-slate-50 border border-slate-200 p-2 rounded-lg font-mono font-semibold text-xs"></div>
        <div><label class="block font-semibold text-slate-500 mb-1">Unit of Measure</label>
          <select id="inward_uom_select" name="uom" required class="w-full bg-slate-50 border border-slate-200 p-2 rounded-lg text-xs focus:outline-none focus:border-indigo-500">
            <option value="Nos">Nos</option><option value="Mtr">Mtr</option><option value="Kg">Kg</option>
            <option value="Ltr">Ltr</option><option value="Set">Set</option><option value="Pair">Pair</option>
            <option value="Box">Box</option><option value="Roll">Roll</option><option value="Bag">Bag</option>
            <option value="Ton">Ton</option><option value="Sqm">Sqm</option><option value="Rmt">Rmt</option><option value="Lot">Lot</option>
          </select></div>
      </div>
      <div class="grid grid-cols-2 gap-2">
        <div><label class="block font-semibold text-slate-500 mb-1">GRN No.</label>
          <input type="text" name="grn_no" placeholder="e.g. GRN-2026-001" class="w-full bg-slate-50 border border-slate-200 p-2 rounded-lg text-xs focus:outline-none focus:border-emerald-500"></div>
        <div><label class="block font-semibold text-slate-500 mb-1">Challan / Invoice No.</label>
          <input type="text" name="challan_no" placeholder="e.g. INV-4521" class="w-full bg-slate-50 border border-slate-200 p-2 rounded-lg text-xs focus:outline-none focus:border-emerald-500"></div>
      </div>
      <div class="bg-indigo-50/60 border border-indigo-100 rounded-lg p-2 text-[11px] text-indigo-700 font-medium">
        <i class="fa-solid fa-user-check mr-1"></i> Received By: <span class="font-black">{recv} ({_html.escape(current_user.username)})</span>
      </div>
      <div id="inward_serial_block" class="hidden bg-amber-50/40 border border-amber-200/60 rounded-xl p-3">
        <div class="flex items-center justify-between mb-1">
          <label class="block font-bold text-amber-700 text-[10px] uppercase tracking-wider">
            <i class="fa-solid fa-barcode mr-1"></i> Serial Numbers <span class="text-rose-500">*</span>
          </label>
          <button type="button" onclick="autoFillSerials()" class="text-[10px] text-amber-700 hover:underline font-bold">Auto-fill</button>
        </div>
        <textarea id="inward_serial_numbers" name="serial_numbers" rows="3" placeholder="EIPL-LAP-001&#10;EIPL-LAP-002&#10;..."
          class="w-full bg-white border border-amber-200 p-2 rounded-lg font-mono text-[11px] focus:outline-none focus:border-amber-500"></textarea>
        <p class="text-[10px] text-amber-700 mt-1 leading-snug">
          One serial per line. Required because this item is a <b>Fixed Asset / Equipment</b>.
          The number of serials must match the Quantity Received.
        </p>
      </div>
      <div><label class="block font-semibold text-slate-500 mb-1">Upload GRN <span class="text-rose-500 font-black">*</span></label>
        <input type="file" name="grn_file" required accept=".pdf,.jpg,.jpeg,.png,.xlsx,.xls,.doc,.docx"
          class="w-full bg-rose-50 border border-rose-200 text-slate-700 text-[10px] p-2 rounded-lg">
        <p class="text-[10px] text-rose-500 mt-0.5">&#9888; GRN must be uploaded to proceed</p>
      </div>
      <button type="submit" class="w-full bg-emerald-600 hover:bg-emerald-700 text-white p-2.5 rounded-xl font-bold tracking-wide transition-all text-xs">&#10003; Record Inward &amp; Upload GRN</button>
    </form>
    <div class="px-5 pb-5 border-t border-dashed border-slate-200 pt-3">
      <label class="block font-black text-[10px] uppercase text-emerald-700 mb-2">&#128202; Bulk Inward Import (CSV)</label>
      <form action="/inward/bulk-import" method="POST" enctype="multipart/form-data" class="flex gap-2">
        <input type="file" name="file" accept=".csv" required class="w-full bg-slate-50 border border-slate-200 text-slate-600 text-[10px] p-2 rounded-xl">
        <button type="submit" class="bg-emerald-600 text-white px-3 py-2 rounded-xl font-bold text-[10px] shrink-0">Import</button>
      </form>
    </div>
  </div>

  <!-- RIGHT: OUTWARD -->
  <div class="bg-white border border-slate-200 rounded-2xl shadow-sm overflow-hidden">
    <div class="px-5 py-3 border-b-2 border-rose-400 bg-rose-50/60">
      <div class="flex items-center justify-between mb-0.5">
        <div class="flex items-center gap-2">
          <span class="bg-rose-500 text-white text-[10px] font-black px-2 py-0.5 rounded tracking-widest uppercase">Outward</span>
          <h2 class="text-xs font-black tracking-wider text-slate-800 uppercase">&#8593; Material Issue</h2>
        </div>
        <a href="/mis/bulk-template" class="bg-rose-50 hover:bg-rose-100 text-rose-700 border border-rose-300 px-2 py-1 rounded-lg text-[10px] font-bold flex items-center gap-1 whitespace-nowrap"><i class="fa-solid fa-download"></i> Template</a>
      </div>
      <p class="text-[10px] text-slate-400">Issue materials from store — MIS upload mandatory</p>
    </div>
    <form action="/material/issue" method="POST" enctype="multipart/form-data" class="p-5 space-y-3 text-xs text-slate-700">
      <div>
        <label class="block font-semibold text-slate-600 mb-1">Search &amp; Select Item</label>
        <div class="relative">
          <input type="text" id="mis_item_search" placeholder="Type item name or code..." autocomplete="off"
            class="w-full bg-slate-50 border border-slate-200 text-slate-900 p-2.5 rounded-xl focus:outline-none focus:border-rose-500 focus:bg-white shadow-sm text-xs"
            oninput="filterMISItems(this.value)" onfocus="showMISDropdown()">
          <div id="mis_item_dropdown" class="absolute z-30 w-full bg-white border border-slate-200 rounded-xl shadow-xl mt-1 max-h-48 overflow-y-auto hidden"></div>
        </div>
        <input type="hidden" id="mis_item_id" name="item_id" required>
        <p id="mis_selected_label" class="text-[10px] text-slate-400 mt-1 italic">No item selected</p>
      </div>
      <div>
        <label class="block font-semibold text-slate-500 mb-1">Category</label>
        <input type="text" id="mis_category_display" readonly placeholder="\u2014 (auto-filled)"
          class="w-full bg-slate-100 border border-slate-200 p-2 rounded-lg text-xs text-slate-500 cursor-not-allowed">
      </div>
      <div id="mis_serial_block" class="hidden bg-amber-50/40 border border-amber-200/60 rounded-xl p-3">
        <label class="block font-bold text-amber-700 text-[10px] uppercase tracking-wider mb-1">
          <i class="fa-solid fa-barcode mr-1"></i> Select Serial <span class="text-rose-500">*</span>
        </label>
        <select id="mis_asset_unit" name="asset_unit_id" class="w-full bg-white border border-amber-200 p-2 rounded-lg font-mono text-[11px] focus:outline-none focus:border-amber-500">
          <option value="">-- Select serial in stock --</option>
        </select>
        <p class="text-[10px] text-amber-700 mt-1">One physical unit per issue (Fixed Asset / Equipment).</p>
      </div>
      <div class="grid grid-cols-2 gap-2">
        <div><label class="block font-semibold text-slate-500 mb-1">Quantity to Issue</label>
          <input type="number" name="quantity" min="1" value="1" required class="w-full bg-slate-50 border border-slate-200 p-2 rounded-lg font-mono font-semibold text-xs"></div>
        <div><label class="block font-semibold text-slate-500 mb-1">UOM</label>
          <input type="text" id="mis_uom_display" name="uom" placeholder="UOM (auto-filled)" readonly class="w-full bg-indigo-50 border border-indigo-200 p-2 rounded-lg text-xs font-bold text-indigo-700"></div>
      </div>
      <div><label class="block font-semibold text-slate-500 mb-1">Issue To (Employee)</label>
        <select name="issued_to" required class="w-full bg-slate-50 border border-slate-200 p-2.5 rounded-xl text-xs focus:outline-none focus:border-indigo-500">
          <option value="">&#8212; Select Employee &#8212;</option>{emp_opts}
        </select></div>
      <div><label class="block font-semibold text-slate-500 mb-1">Department</label>
        <input type="text" name="department" placeholder="e.g. Udaipur Project" class="w-full bg-slate-50 border border-slate-200 p-2 rounded-lg text-xs"></div>
      <div><label class="block font-semibold text-slate-500 mb-1">Remarks (optional)</label>
        <input type="text" name="remarks" placeholder="Purpose / Location" class="w-full bg-slate-50 border border-slate-200 p-2 rounded-lg text-xs"></div>
      <div><label class="block font-semibold text-slate-500 mb-1">Upload MIS <span class="text-rose-500 font-black">*</span></label>
        <input type="file" name="mis_file" required accept=".pdf,.jpg,.jpeg,.png,.xlsx,.xls,.doc,.docx"
          class="w-full bg-rose-50 border border-rose-200 text-slate-700 text-[10px] p-2 rounded-lg"></div>
      <button type="submit" class="w-full bg-rose-600 hover:bg-rose-700 text-white p-2.5 rounded-xl font-bold tracking-wide transition-all text-xs">&#8593; Issue Material &amp; Upload MIS</button>
    </form>
    <div class="px-5 pb-5 border-t border-dashed border-slate-200 pt-3">
      <label class="block font-black text-[10px] uppercase text-rose-700 tracking-wider mb-2">&#128202; Bulk MIS Import (CSV)</label>
      <form action="/mis/bulk-import" method="POST" enctype="multipart/form-data" class="flex gap-2">
        <input type="file" name="file" accept=".csv" required class="w-full bg-slate-50 border border-slate-200 text-slate-600 text-[10px] p-2 rounded-xl">
        <button type="submit" class="bg-rose-600 text-white px-3 py-2 rounded-xl font-bold text-[10px] shrink-0">Import</button>
      </form>
      <p class="text-[10px] text-slate-400 mt-1">Columns: item_code, quantity, uom, issued_to, department, remarks</p>
    </div>
  </div>

  <!-- RIGHTMOST: RETURN TO STORE -->
  <div class="bg-white border border-slate-200 rounded-2xl shadow-sm overflow-hidden">
    <div class="px-5 py-3 border-b-2 border-amber-400 bg-amber-50/60">
      <div class="flex items-center justify-between mb-0.5">
        <div class="flex items-center gap-2">
          <span class="bg-amber-500 text-white text-[10px] font-black px-2 py-0.5 rounded tracking-widest uppercase">Return</span>
          <h2 class="text-xs font-black tracking-wider text-slate-800 uppercase">&#8634; Return to Store</h2>
        </div>
        <a href="/rts/bulk-template" class="bg-amber-50 hover:bg-amber-100 text-amber-700 border border-amber-300 px-2 py-1 rounded-lg text-[10px] font-bold flex items-center gap-1 whitespace-nowrap"><i class="fa-solid fa-download"></i> Template</a>
      </div>
      <p class="text-[10px] text-slate-400">Take returned materials back into stock &mdash; slip mandatory</p>
    </div>
    <form action="/material/return" method="POST" enctype="multipart/form-data" class="p-5 space-y-3 text-xs text-slate-700">
      <div>
        <label class="block font-semibold text-slate-600 mb-1">Returning Employee</label>
        <select id="rts_employee" name="returned_by" required onchange="onRtsEmployeeChange()"
          class="w-full bg-slate-50 border border-slate-200 p-2.5 rounded-xl focus:outline-none focus:border-amber-500 focus:bg-white text-xs">
          <option value="">-- Select Employee --</option>
          {emp_opts}
        </select>
      </div>

      <div id="rts_held_panel" class="hidden bg-amber-50/40 border border-amber-200/60 rounded-xl p-3">
        <p class="text-[10px] font-bold text-amber-700 uppercase tracking-wider mb-1.5">Items currently held by this employee</p>
        <div id="rts_held_list" class="space-y-1 text-[11px] text-slate-700 max-h-32 overflow-y-auto"></div>
      </div>
      <div id="rts_no_items" class="hidden bg-slate-50 border border-slate-200 rounded-xl p-3 text-center">
        <p class="text-[11px] text-slate-500">No outstanding items for this employee.</p>
      </div>

      <div>
        <label class="block font-semibold text-slate-600 mb-1">Item Being Returned</label>
        <select id="rts_item" name="item_id" required disabled onchange="onRtsItemChange()"
          class="w-full bg-slate-50 border border-slate-200 p-2.5 rounded-xl focus:outline-none focus:border-amber-500 focus:bg-white text-xs disabled:opacity-50">
          <option value="">-- Select employee first --</option>
        </select>
      </div>

      <div id="rts_serial_block" class="hidden bg-amber-50/40 border border-amber-200/60 rounded-xl p-3">
        <label class="block font-bold text-amber-700 text-[10px] uppercase tracking-wider mb-1">
          <i class="fa-solid fa-barcode mr-1"></i> Serial Being Returned <span class="text-rose-500">*</span>
        </label>
        <select id="rts_asset_unit" name="asset_unit_id" class="w-full bg-white border border-amber-200 p-2 rounded-lg font-mono text-[11px] focus:outline-none focus:border-amber-500">
          <option value="">-- Select serial being returned --</option>
        </select>
        <p class="text-[10px] text-amber-700 mt-1">One physical unit per return (Fixed Asset / Equipment).</p>
      </div>

      <div class="grid grid-cols-2 gap-2">
        <div>
          <label class="block font-semibold text-slate-500 mb-1">Quantity Returned</label>
          <input type="number" id="rts_qty" name="quantity" min="1" value="1" required
            class="w-full bg-slate-50 border border-slate-200 p-2 rounded-lg font-mono font-semibold text-xs">
          <p id="rts_qty_help" class="text-[10px] text-slate-400 mt-0.5">Max: \u2014</p>
        </div>
        <div>
          <label class="block font-semibold text-slate-500 mb-1">UOM (auto)</label>
          <input type="text" id="rts_uom_display" readonly placeholder="\u2014"
            class="w-full bg-slate-100 border border-slate-200 p-2 rounded-lg text-xs text-slate-500 cursor-not-allowed">
        </div>
      </div>

      <div>
        <label class="block font-semibold text-slate-500 mb-1">Category</label>
        <input type="text" id="rts_category_display" readonly placeholder="\u2014 (auto-filled)"
          class="w-full bg-slate-100 border border-slate-200 p-2 rounded-lg text-xs text-slate-500 cursor-not-allowed">
      </div>

      <div>
        <label class="block font-semibold text-slate-500 mb-1">Department</label>
        <input type="text" name="department" value="General Operations"
          class="w-full bg-slate-50 border border-slate-200 p-2 rounded-lg text-xs">
      </div>

      <div>
        <label class="block font-semibold text-slate-500 mb-1">Remarks <span class="text-slate-400 font-normal">(optional)</span></label>
        <textarea name="remarks" rows="2" placeholder="Reason for return, condition, etc."
          class="w-full bg-slate-50 border border-slate-200 p-2 rounded-lg text-xs resize-none"></textarea>
      </div>

      <div>
        <label class="block font-semibold text-slate-500 mb-1">Upload Return-to-Store Slip <span class="text-rose-500 font-black">*</span></label>
        <input type="file" name="return_file" required accept=".pdf,.jpg,.jpeg,.png,.xlsx,.xls,.doc,.docx"
          class="w-full bg-amber-50 border border-amber-200 text-slate-700 text-[10px] p-2 rounded-lg">
        <p class="text-[10px] text-amber-700 mt-0.5">&#9888; Return-to-Store Slip is mandatory</p>
      </div>

      <button type="submit" class="w-full bg-amber-500 hover:bg-amber-600 text-white p-2.5 rounded-xl font-bold tracking-wide transition-all text-xs">&#8634; Record Return to Store</button>
    </form>
  </div>

  </div><!-- end grid -->
</div>

<script id="mm-held" type="application/json">{held_json}</script>
<script id="mm-held-units" type="application/json">{held_units_json}</script>
<script>
var MMHELD = JSON.parse(document.getElementById('mm-held').textContent);
var MMHELD_UNITS = JSON.parse(document.getElementById('mm-held-units').textContent);
function onRtsEmployeeChange() {{
    var emp = document.getElementById('rts_employee').value;
    var itemSel = document.getElementById('rts_item');
    var panel = document.getElementById('rts_held_panel');
    var noBox = document.getElementById('rts_no_items');
    var listDiv = document.getElementById('rts_held_list');
    var qty = document.getElementById('rts_qty');
    var qtyHelp = document.getElementById('rts_qty_help');
    var uomDisp = document.getElementById('rts_uom_display');
    var catDisp = document.getElementById('rts_category_display');
    var serialBlock = document.getElementById('rts_serial_block');
    var serialSel = document.getElementById('rts_asset_unit');
    itemSel.innerHTML = '<option value="">-- Select item --</option>';
    uomDisp.value = ''; qty.value = 1; qty.max = ''; qtyHelp.textContent = 'Max: \u2014';
    if (catDisp) catDisp.value = '';
    if (serialBlock) {{ serialBlock.classList.add('hidden'); serialSel.required=false; serialSel.value=''; }}

    if (!emp) {{
        itemSel.disabled = true; panel.classList.add('hidden'); noBox.classList.add('hidden'); return;
    }}
    var consumRows = MMHELD[emp] || [];
    var unitRows = MMHELD_UNITS[emp] || [];
    if (!consumRows.length && !unitRows.length) {{
        itemSel.disabled = true; panel.classList.add('hidden'); noBox.classList.remove('hidden'); return;
    }}
    noBox.classList.add('hidden');
    panel.classList.remove('hidden');

    var consumDisplay = consumRows.map(function(r) {{
        return '<div class="flex justify-between gap-2 border-b border-amber-100 last:border-0 pb-1">' +
            '<span class="truncate"><b>' + r.name + '</b> <span class="text-slate-400 text-[10px]">(' + r.code + ')</span></span>' +
            '<span class="font-mono text-amber-700 font-bold shrink-0">' + r.qty_held + ' ' + r.uom + '</span></div>';
    }}).join('');
    // Group tracked rows by item_id for the held panel + dropdown
    var byItem = {{}};
    unitRows.forEach(function(u) {{
        if (!byItem[u.item_id]) byItem[u.item_id] = {{ name: u.item_name, code: u.item_code, uom: u.uom, category: u.category, serials: [] }};
        byItem[u.item_id].serials.push({{ id: u.asset_unit_id, sn: u.serial_no }});
    }});
    var unitDisplay = Object.keys(byItem).map(function(iid) {{
        var g = byItem[iid];
        var serialList = g.serials.map(function(s){{return s.sn;}}).join(', ');
        return '<div class="border-b border-amber-100 last:border-0 pb-1">' +
            '<div class="flex justify-between gap-2"><span class="truncate"><b>' + g.name + '</b> <span class="text-slate-400 text-[10px]">(' + g.code + ')</span></span>' +
            '<span class="font-mono text-amber-700 font-bold shrink-0">' + g.serials.length + ' ' + g.uom + '</span></div>' +
            '<div class="text-[9px] text-slate-500 font-mono mt-0.5">' + serialList + '</div></div>';
    }}).join('');
    listDiv.innerHTML = consumDisplay + unitDisplay;

    consumRows.forEach(function(r) {{
        var opt = document.createElement('option');
        opt.value = r.item_id;
        opt.textContent = r.name + ' (' + r.code + ') \u2014 holding ' + r.qty_held + ' ' + r.uom;
        opt.dataset.uom = r.uom; opt.dataset.max = r.qty_held;
        opt.dataset.category = r.category || ''; opt.dataset.tracked = '0';
        itemSel.appendChild(opt);
    }});
    Object.keys(byItem).forEach(function(iid) {{
        var g = byItem[iid];
        var opt = document.createElement('option');
        opt.value = iid;
        opt.textContent = g.name + ' (' + g.code + ') \u2014 ' + g.serials.length + ' serial(s) held';
        opt.dataset.uom = g.uom; opt.dataset.max = g.serials.length;
        opt.dataset.category = g.category || ''; opt.dataset.tracked = '1';
        opt.dataset.serials = JSON.stringify(g.serials);
        itemSel.appendChild(opt);
    }});
    itemSel.disabled = false;
}}
function onRtsItemChange() {{
    var sel = document.getElementById('rts_item');
    var opt = sel.options[sel.selectedIndex];
    var qty = document.getElementById('rts_qty');
    var qtyHelp = document.getElementById('rts_qty_help');
    var uomDisp = document.getElementById('rts_uom_display');
    var catDisp = document.getElementById('rts_category_display');
    var serialBlock = document.getElementById('rts_serial_block');
    var serialSel = document.getElementById('rts_asset_unit');
    if (!opt || !opt.value) {{
        uomDisp.value = ''; qty.max = ''; qtyHelp.textContent = 'Max: \u2014';
        if (catDisp) catDisp.value = '';
        if (serialBlock) {{ serialBlock.classList.add('hidden'); serialSel.required=false; serialSel.value=''; }}
        return;
    }}
    uomDisp.value = opt.dataset.uom || '';
    if (catDisp) catDisp.value = opt.dataset.category || '';
    qty.max = opt.dataset.max || '';
    qty.value = 1;
    qtyHelp.textContent = 'Max: ' + (opt.dataset.max || '\u2014') + ' ' + (opt.dataset.uom || '');
    if (opt.dataset.tracked === '1' && serialBlock) {{
        var serials = JSON.parse(opt.dataset.serials || '[]');
        serialSel.innerHTML = '<option value="">-- Select serial being returned --</option>';
        serials.forEach(function(s) {{
            var o = document.createElement('option');
            o.value = s.id; o.textContent = s.sn;
            serialSel.appendChild(o);
        }});
        serialBlock.classList.remove('hidden');
        serialSel.required = true;
        qty.value = 1; qty.readOnly = true;
    }} else if (serialBlock) {{
        serialBlock.classList.add('hidden');
        serialSel.required = false; serialSel.value = '';
        qty.readOnly = false;
    }}
}}
</script>

<script id="mm-inv" type="application/json">{inv_json}</script>
<script id="mm-units-by-item" type="application/json">{units_by_item_json}</script>
<script>
var MMINV = JSON.parse(document.getElementById('mm-inv').textContent);
var MM_UNITS_BY_ITEM = JSON.parse(document.getElementById('mm-units-by-item').textContent);
function mkDd(items, cls) {{
    return items.slice(0,10).map(function(i) {{
        return '<div class="p-2.5 hover:bg-indigo-50 cursor-pointer border-b border-slate-100 last:border-0 ' + cls + '"'
            + ' data-id="'+i.id+'"'
            + ' data-name="'+i.name.replace(/"/g,"&quot;")+'"'
            + ' data-code="'+i.code+'"'
            + ' data-stock="'+i.stock+'"'
            + ' data-uom="'+i.uom+'"'
            + ' data-price="'+i.price+'"'
            + ' data-vendor="'+i.vendor+'"'
            + ' data-category="'+(i.category||'')+'"'
            + ' data-tracked="'+(i.tracked?'1':'0')+'">'
            + '<div class="font-semibold text-xs pointer-events-none">'+i.name+(i.tracked?'<span class="ml-1 bg-amber-100 text-amber-700 text-[9px] font-bold px-1.5 py-0.5 rounded">TRACKED</span>':'')+'</div>'
            + '<div class="text-[10px] text-slate-400 pointer-events-none">'+i.code+' | Stock: '+i.stock+'</div></div>';
    }}).join('') || '<div class="p-3 text-slate-400 text-xs">No items found</div>';
}}

// Auto-fill serial textarea: prompts for prefix + start number, generates N serials.
function autoFillSerials() {{
    var qtyInput = document.querySelector('form[action="/transaction"] [name=quantity]');
    var qty = parseInt((qtyInput && qtyInput.value) || '0', 10);
    if (qty <= 0) {{ alert('Set Quantity Received first.'); return; }}
    var codeFallback = document.getElementById('inward_item_search').value.match(/\\(([^)]+)\\)/);
    codeFallback = codeFallback ? codeFallback[1] : 'EIPL';
    var prefix = window.prompt('Enter serial prefix (without trailing dash):', codeFallback);
    if (!prefix) return;
    var startStr = window.prompt('Starting number (e.g. 1):', '1');
    var start = parseInt(startStr, 10);
    if (isNaN(start) || start < 0) return;
    var lines = [];
    for (var i = 0; i < qty; i++) {{
        lines.push(prefix + '-' + String(start + i).padStart(3, '0'));
    }}
    document.getElementById('inward_serial_numbers').value = lines.join('\\n');
}}

function filterInwardItems(q) {{
    var dd=document.getElementById('inward_item_dropdown');
    var m=q ? MMINV.filter(function(i){{return i.name.toLowerCase().includes(q.toLowerCase())||i.code.toLowerCase().includes(q.toLowerCase());}}) : [];
    if(!q){{dd.classList.add('hidden');return;}}
    dd.innerHTML=mkDd(m,'inw-item'); dd.classList.remove('hidden');
}}
function showInwardDropdown(){{filterInwardItems(document.getElementById('inward_item_search').value);}}
function filterMISItems(q) {{
    var dd=document.getElementById('mis_item_dropdown');
    var m=q ? MMINV.filter(function(i){{return i.name.toLowerCase().includes(q.toLowerCase())||i.code.toLowerCase().includes(q.toLowerCase());}}) : [];
    if(!q){{dd.classList.add('hidden');return;}}
    dd.innerHTML=mkDd(m,'mis-item'); dd.classList.remove('hidden');
}}
function showMISDropdown(){{filterMISItems(document.getElementById('mis_item_search').value);}}
document.addEventListener('click',function(e){{
    var el=e.target.closest('.inw-item');
    if(el){{
        var tracked = el.dataset.tracked === '1';
        document.getElementById('inward_item_id').value=el.dataset.id;
        document.getElementById('inward_item_search').value=el.dataset.name+' ('+el.dataset.code+')';
        document.getElementById('inward_selected_label').textContent='Selected: '+el.dataset.name+' | Stock: '+el.dataset.stock+(tracked?' | TRACKED':'');
        document.getElementById('inward_selected_label').className='text-[10px] text-indigo-600 font-semibold mt-1';
        document.getElementById('inward_item_dropdown').classList.add('hidden');
        if(el.dataset.uom)document.getElementById('inward_uom_select').value=el.dataset.uom;
        if(document.getElementById('inward_category_display'))document.getElementById('inward_category_display').value=el.dataset.category||'';
        // Show/hide Serial Numbers block based on tracking
        var sb = document.getElementById('inward_serial_block');
        var ta = document.getElementById('inward_serial_numbers');
        if (tracked) {{
            sb.classList.remove('hidden');
            ta.required = true;
        }} else {{
            sb.classList.add('hidden');
            ta.value = '';
            ta.required = false;
        }}
        return;
    }}
    var el2=e.target.closest('.mis-item');
    if(el2){{
        var tracked2 = el2.dataset.tracked === '1';
        document.getElementById('mis_item_id').value=el2.dataset.id;
        document.getElementById('mis_item_search').value=el2.dataset.name+' ('+el2.dataset.code+')';
        document.getElementById('mis_selected_label').textContent='Selected: '+el2.dataset.name+' | Stock: '+el2.dataset.stock+(tracked2?' | TRACKED':'');
        document.getElementById('mis_uom_display').value=el2.dataset.uom||'';
        if(document.getElementById('mis_category_display'))document.getElementById('mis_category_display').value=el2.dataset.category||'';
        document.getElementById('mis_item_dropdown').classList.add('hidden');
        // Show/hide MIS serial picker
        var sblk = document.getElementById('mis_serial_block');
        var sel = document.getElementById('mis_asset_unit');
        var misQty = document.querySelector('form[action="/material/issue"] [name=quantity]');
        if (tracked2) {{
            sblk.classList.remove('hidden');
            sel.required = true;
            sel.innerHTML = '<option value="">-- Select serial in stock --</option>';
            var units = MM_UNITS_BY_ITEM[el2.dataset.id] || [];
            units.forEach(function(u) {{
                var opt = document.createElement('option');
                opt.value = u.id; opt.textContent = u.serial_no;
                sel.appendChild(opt);
            }});
            if (misQty) {{ misQty.value = 1; misQty.readOnly = true; }}
        }} else {{
            sblk.classList.add('hidden');
            sel.required = false; sel.value = '';
            if (misQty) {{ misQty.readOnly = false; }}
        }}
        return;
    }}
    if(document.getElementById('inward_item_dropdown')&&!document.getElementById('inward_item_dropdown').contains(e.target)&&e.target.id!=='inward_item_search')document.getElementById('inward_item_dropdown').classList.add('hidden');
    if(document.getElementById('mis_item_dropdown')&&!document.getElementById('mis_item_dropdown').contains(e.target)&&e.target.id!=='mis_item_search')document.getElementById('mis_item_dropdown').classList.add('hidden');
}});
</script>
</div><!-- end flex-1 -->
</body></html>"""
    return HTMLResponse(html_page)


@app.get("/", response_class=HTMLResponse)
def root_dashboard(current_user: models.User = Depends(get_current_user), db: Session = Depends(get_db)):
    if not current_user:
        return RedirectResponse(url="/login", status_code=303)

    items = db.query(models.Item).all()
    assignments = db.query(models.MaterialAssignment).all()
    employees = db.query(models.Employee).all()
    users = db.query(models.User).all()

    # Per-item transaction logs for the merged log column / modal
    import json as _json_txn
    txn_logs = build_txn_logs(db, items)

    # Per-item asset register for the Serials modal (tracked items)
    asset_register = {}
    for u in db.query(models.AssetUnit).order_by(models.AssetUnit.acquired_at.asc()).all():
        asset_register.setdefault(u.item_id, []).append({
            "serial_no": u.serial_no or "—",
            "status": u.status or "In Stock",
            "current_holder": u.current_holder or "",
            "acquired_at": fmt_ist(u.acquired_at, "%d-%m-%Y"),
            "notes": u.notes or "",
        })

    # Per-item GRN receipt history — used as a fallback for the vendor-lot panel
    # when an item has stock but no live ItemLot rows (legacy stock recorded
    # before the lot ledger existed, or lots fully consumed). Lets the panel
    # still surface the real vendors the item was received from.
    grn_hist_by_item = {}
    for g in db.query(models.GRNRecord).all():
        if (g.quantity or 0) > 0:
            grn_hist_by_item.setdefault(g.item_id, []).append(g)

    item_options = "".join([f'<option value="{i.id}">{i.name} ({i.item_code})</option>' for i in items])
    employee_options = "".join([f'<option value="{e.name}">{e.name} - {e.role_title}</option>' for e in employees])

    # 1. FIXED MODAL SCRIPT ENGINE BLOCK WITH ADDED DEPARTMENT INPUT
    inventory_rows = """
    <script>
    // 1. THE CORRECTED PROCUREMENT REQUEST MODAL ENGINE
    window.openProcurementModal = function(id, name, item_code) {
        let modal = document.getElementById('dynamicProcurementModal');
        if (!modal) {
            modal = document.createElement('div');
            modal.id = 'dynamicProcurementModal';
            modal.className = 'fixed inset-0 bg-slate-900/50 backdrop-blur-sm flex items-center justify-center z-50';
            document.body.appendChild(modal);
        }
        modal.innerHTML = `
            <div class="bg-white rounded-xl border border-slate-200 shadow-2xl w-full max-w-md p-6 text-left relative">
                <h3 class="text-lg font-bold text-slate-900 mb-4 flex items-center gap-2" style="font-size: 1.125rem; font-weight: 700; color: #0f172a; margin-bottom: 1rem;">📋 Create Procurement Request</h3>
                <form action="/procurement/request" method="POST" style="display: flex; flex-direction: column; gap: 1rem;">
                    <input type="hidden" name="item_id" value="` + id + `">
                    <div>
                        <label style="display: block; font-size: 0.75rem; font-weight: 600; color: #64748b; text-transform: uppercase; margin-bottom: 0.25rem;">Item Name</label>
                        <input type="text" value="` + name + `" readonly style="width: 100%; padding: 0.5rem 0.75rem; border: 1px solid #cbd5e1; border-radius: 0.375rem; font-size: 0.875rem; color: #64748b; background-color: #f8fafc;">
                    </div>
                    <div>
                        <label style="display: block; font-size: 0.75rem; font-weight: 600; color: #64748b; text-transform: uppercase; margin-bottom: 0.25rem;">Item Code</label>
                        <input type="text" value="` + item_code + `" readonly style="width: 100%; padding: 0.5rem 0.75rem; border: 1px solid #cbd5e1; border-radius: 0.375rem; font-size: 0.875rem; font-family: monospace; color: #64748b; background-color: #f8fafc;">
                    </div>
                    <div>
                        <label style="display: block; font-size: 0.75rem; font-weight: 600; color: #64748b; text-transform: uppercase; margin-bottom: 0.25rem;">Target Department</label>
                        <select name="department" required style="width: 100%; padding: 0.5rem 0.75rem; border: 1px solid #cbd5e1; border-radius: 0.375rem; font-size: 0.875rem; color: #0f172a; background-color: #ffffff;">
                            <option value="Operations">Operations</option>
                            <option value="Mechanical">Mechanical</option>
                            <option value="Electrical">Electrical</option>
                            <option value="Commercial">Commercial</option>
                            <option value="Stores">Stores</option>
                        </select>
                    </div>
                    <div>
                        <label style="display: block; font-size: 0.75rem; font-weight: 600; color: #64748b; text-transform: uppercase; margin-bottom: 0.25rem;">Quantity Required</label>
                        <input type="number" name="quantity" min="1" value="1" required style="width: 100%; padding: 0.5rem 0.75rem; border: 1px solid #cbd5e1; border-radius: 0.375rem; font-size: 0.875rem;">
                    </div>
                    <div style="display: flex; justify-content: flex-end; gap: 0.5rem; padding-top: 0.75rem; border-top: 1px solid #e2e8f0;">
                        <button type="button" onclick="document.getElementById('dynamicProcurementModal').remove()" style="padding: 0.5rem 1rem; font-size: 0.75rem; font-weight: 700; color: #475569; background-color: #f1f5f9; border: none; border-radius: 0.375rem; cursor: pointer;">Cancel</button>
                        <button type="submit" style="padding: 0.5rem 1rem; font-size: 0.75rem; font-weight: 700; color: white; background-color: #4338ca; border: none; border-radius: 0.375rem; cursor: pointer;">Submit Request</button>
                    </div>
                </form>
            </div>
        `;
    };

    // 2. THE EDIT ITEM MENU — styled to match Procure widget panel
    window.openEditItemModal = function(id, name, Unique_code, price, current_stock, minimum_stock, supplier, site) {
        let modal = document.getElementById('dynamicEditItemModal');
        if (!modal) {
            modal = document.createElement('div');
            modal.id = 'dynamicEditItemModal';
            modal.className = 'fixed inset-0 bg-slate-900/50 backdrop-blur-sm flex items-center justify-center z-50';
            modal.onclick = function(e) { if (e.target === modal) modal.remove(); };
            document.body.appendChild(modal);
        }
        modal.innerHTML = `
            <div style="background:#fff;border-radius:1rem;border:1px solid #e2e8f0;box-shadow:0 25px 50px -12px rgba(0,0,0,0.25);width:100%;max-width:28rem;padding:1.5rem;text-align:left;position:relative;">
                <div style="display:flex;align-items:center;justify-content:space-between;border-bottom:1px solid #f1f5f9;padding-bottom:0.75rem;margin-bottom:1.25rem;">
                    <div>
                        <h3 style="font-size:0.75rem;font-weight:800;color:#0f172a;text-transform:uppercase;letter-spacing:0.05em;">&#9998; Edit Inventory Item</h3>
                        <p style="font-size:0.625rem;color:#94a3b8;margin-top:0.125rem;">Update item details and stock levels</p>
                    </div>
                    <button type="button" onclick="document.getElementById('dynamicEditItemModal').remove()"
                        style="color:#94a3b8;background:#f8fafc;border:none;border-radius:0.5rem;padding:0.375rem 0.5rem;cursor:pointer;font-size:0.75rem;">&#10005;</button>
                </div>
                <form action="/items/edit/` + id + `" method="POST" style="display:flex;flex-direction:column;gap:0.875rem;">
                    <div>
                        <label style="display:block;font-size:0.65rem;font-weight:700;color:#4338ca;text-transform:uppercase;letter-spacing:0.05em;margin-bottom:0.25rem;">&#9670; Unique Item Code</label>
                        <input type="text" name="item_code" value="` + Unique_code + `" required
                            style="width:100%;padding:0.625rem 0.75rem;border:2px solid #6366f1;border-radius:0.625rem;font-size:0.8rem;font-family:monospace;color:#0f172a;background:#fafafa;box-sizing:border-box;text-transform:uppercase;"
                            oninput="this.value=this.value.toUpperCase()">
                    </div>
                    <div>
                        <label style="display:block;font-size:0.65rem;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:0.05em;margin-bottom:0.25rem;">Item Name</label>
                        <input type="text" name="name" value="` + name + `" required
                            style="width:100%;padding:0.625rem 0.75rem;border:1px solid #e2e8f0;border-radius:0.625rem;font-size:0.8rem;color:#0f172a;background:#f8fafc;box-sizing:border-box;">
                    </div>
                    <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.75rem;">
                        <div>
                            <label style="display:block;font-size:0.65rem;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:0.05em;margin-bottom:0.25rem;">Price (&#8377;)</label>
                            <input type="number" step="0.01" name="price" value="` + price + `" required
                                style="width:100%;padding:0.625rem 0.75rem;border:1px solid #e2e8f0;border-radius:0.625rem;font-size:0.8rem;font-family:monospace;color:#1e293b;box-sizing:border-box;">
                        </div>
                        <div>
                            <label style="display:block;font-size:0.65rem;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:0.05em;margin-bottom:0.25rem;">Safety Stock</label>
                            <input type="number" name="minimum_stock" value="` + minimum_stock + `" required
                                style="width:100%;padding:0.625rem 0.75rem;border:1px solid #e2e8f0;border-radius:0.625rem;font-size:0.8rem;font-family:monospace;color:#1e293b;box-sizing:border-box;">
                        </div>
                    </div>
                    <div>
                        <label style="display:block;font-size:0.65rem;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:0.05em;margin-bottom:0.25rem;">Current Stock Level</label>
                        <input type="number" name="current_stock" value="` + current_stock + `" required
                            style="width:100%;padding:0.625rem 0.75rem;border:1px solid #e2e8f0;border-radius:0.625rem;font-size:0.8rem;font-family:monospace;color:#1e293b;box-sizing:border-box;">
                    </div>
                    <div style="display:grid;grid-template-columns:1fr 1fr;gap:0.75rem;">
                        <div>
                            <label style="display:block;font-size:0.65rem;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:0.05em;margin-bottom:0.25rem;">Supplier / Vendor</label>
                            <input type="text" name="supplier" value="` + (supplier||'') + `"
                                style="width:100%;padding:0.625rem 0.75rem;border:1px solid #e2e8f0;border-radius:0.625rem;font-size:0.8rem;color:#1e293b;background:#f8fafc;box-sizing:border-box;" placeholder="Vendor name">
                        </div>
                        <div>
                            <label style="display:block;font-size:0.65rem;font-weight:700;color:#64748b;text-transform:uppercase;letter-spacing:0.05em;margin-bottom:0.25rem;">Storage Site</label>
                            <input type="text" name="storage_site" value="` + (site||'') + `"
                                style="width:100%;padding:0.625rem 0.75rem;border:1px solid #e2e8f0;border-radius:0.625rem;font-size:0.8rem;color:#1e293b;background:#f8fafc;box-sizing:border-box;" placeholder="e.g. Store Yard">
                        </div>
                    </div>
                    <div style="display:flex;justify-content:flex-end;gap:0.5rem;padding-top:0.75rem;border-top:1px solid #f1f5f9;margin-top:0.25rem;">
                        <button type="button" onclick="document.getElementById('dynamicEditItemModal').remove()"
                            style="padding:0.5rem 1.125rem;font-size:0.7rem;font-weight:700;color:#475569;background:#f1f5f9;border:none;border-radius:0.5rem;cursor:pointer;">Cancel</button>
                        <button type="submit"
                            style="padding:0.5rem 1.125rem;font-size:0.7rem;font-weight:700;color:white;background:#d97706;border:none;border-radius:0.5rem;cursor:pointer;">&#10003; Save Changes</button>
                    </div>
                </form>
            </div>
        `;
    };
    </script>
    """

    for it in items:
        alert_status = ""
        if it.current_stock <= it.minimum_stock:
            alert_status = ' <span class="text-rose-600 font-black animate-pulse">&#9888;&#65039; LOW</span>'
        
        item_cat = getattr(it, 'category', None) or DEFAULT_ASSET_CLASS
        item_sup = getattr(it, 'supplier', 'EIPL Approved Vendor')
        item_site = getattr(it, 'storage_site', 'Store Yard')

        import html as _html
        h_name = _html.escape(it.name, quote=True)
        h_code = _html.escape(it.item_code, quote=True)
        h_vendor = _html.escape(item_sup, quote=True)
        h_uom = _html.escape(getattr(it, 'uom', '') or '', quote=True)
        h_cat = _html.escape(item_cat, quote=True)

        if current_user.role == "Admin":
            admin_only_cells = f"""
            <td class="p-3 text-slate-600 text-[11px] font-medium text-center">{item_sup}</td>
            <td class="p-3 font-mono text-slate-600 text-center">&#8377;{it.price:,.2f}</td>"""
            action_cell = f"""
            <td class="p-3">
                <div class="flex items-center gap-2">
                    <button type="button"
                        data-action="create-indent"
                        data-id="{it.id}" data-name="{h_name}" data-code="{h_code}"
                        data-uom="{h_uom}" data-dept="{_html.escape(item_site, quote=True)}"
                        class="text-indigo-600 hover:underline font-bold text-[11px]">+ Indent</button>
                    <button type="button"
                        data-action="edit-item"
                        data-id="{it.id}" data-name="{h_name}" data-code="{h_code}"
                        data-price="{it.price}" data-stock="{it.current_stock}" data-minstock="{it.minimum_stock}"
                        data-supplier="{h_vendor}" data-site="{_html.escape(item_site, quote=True)}"
                        data-category="{h_cat}"
                        class="text-amber-600 hover:underline font-bold text-[11px]">Edit</button>
                    <form action="/items/delete/{it.id}" method="POST" class="inline m-0 delete-protected-form" onsubmit="event.preventDefault(); openAdminDeleteModal(this, 'Delete item: {it.name} ({it.item_code})?');">
                        <button type="submit" class="text-rose-600 hover:underline font-bold text-[11px] bg-transparent border-none p-0 cursor-pointer inline">Delete</button>
                    </form>
                </div>
            </td>"""
        else:
            admin_only_cells = ""
            action_cell = f"""
            <td class="p-3">
                <button type="button"
                    data-action="create-indent"
                    data-id="{it.id}" data-name="{h_name}" data-code="{h_code}"
                    data-uom="{h_uom}" data-dept="{_html.escape(item_site, quote=True)}"
                    class="text-indigo-600 hover:underline font-bold text-[11px]">+ Indent</button>
            </td>"""

        inventory_rows += f"""
        <tr class="border-b border-slate-100 hover:bg-slate-50/50" data-row="1">
            <td class="p-3 font-semibold text-slate-900">
                <span>{it.name}</span>{alert_status}
                <div class="text-[10px] text-slate-400 mt-0.5">
                    <span class="bg-blue-50 px-1 py-0.5 rounded text-blue-600">{item_site}</span>
                </div>
            </td>
            <td class="p-3 font-mono text-[11px] text-slate-500">{it.item_code}</td>
            <td class="p-3 text-center">
                <span class="bg-slate-100 px-2 py-0.5 rounded text-slate-600 text-[10px] font-semibold whitespace-nowrap">{item_cat}</span>
            </td>
            <td class="p-3 text-center font-mono text-[11px] text-slate-500">{h_uom or '—'}</td>
            {admin_only_cells}
            <td class="p-3 font-mono font-bold text-slate-900" data-sort="{it.current_stock}">
                {it.current_stock}
                <span class="lot-toggle text-[10px] text-indigo-500 font-semibold cursor-pointer underline ml-1" data-target="lot-row-{it.id}">vendors</span>
            </td>
            <td class="p-3 font-mono text-slate-400" data-sort="{it.minimum_stock}">{it.minimum_stock}</td>
            {action_cell}
            <td class="p-3 text-center">
                <div class="flex items-center justify-center gap-1 flex-wrap">
                    <button type="button"
                        onclick="openTxnLog({it.id}, '{h_name}', '{h_code}')"
                        class="bg-indigo-50 hover:bg-indigo-100 text-indigo-700 border border-indigo-200 font-bold text-[10px] px-2.5 py-1.5 rounded-lg transition-all inline-flex items-center gap-1">
                        <i class="fa-solid fa-book-open text-[9px]"></i> Log
                    </button>
                    {f'''<button type="button"
                        onclick="openAssetRegister({it.id}, '{h_name}', '{h_code}')"
                        class="bg-amber-50 hover:bg-amber-100 text-amber-700 border border-amber-200 font-bold text-[10px] px-2.5 py-1.5 rounded-lg transition-all inline-flex items-center gap-1">
                        <i class="fa-solid fa-barcode text-[9px]"></i> Serials
                    </button>''' if is_tracked(it) else ''}
                </div>
            </td>
        </tr>
        <script>window.__txnData = window.__txnData||{{}};window.__txnData[{it.id}]={_json_txn.dumps(txn_logs.get(it.id, []))};window.__assetData = window.__assetData||{{}};window.__assetData[{it.id}]={_json_txn.dumps(asset_register.get(it.id, []))};</script>
        """

        # Multi-vendor / multi-price lot breakdown — hidden by default, toggled via "vendors" link.
        lots = db.query(models.ItemLot).filter(
            models.ItemLot.item_id == it.id,
            models.ItemLot.quantity > 0
        ).order_by(models.ItemLot.received_at.asc()).all()

        n_cols = 9 if current_user.role == "Admin" else 7
        lot_caption = f"Vendor / Price Lots for {h_name} (FIFO order — oldest first)"
        lot_footer = f'Total Quantity: <span class="font-bold text-slate-600">{it.current_stock} {h_uom or ""}</span> &bull; Outward issues consume the oldest lot first.'

        if lots:
            lot_lines = "".join(
                f"""<div class="flex items-center justify-between gap-3 py-1 px-2 odd:bg-white even:bg-slate-50 rounded">
                        <span class="font-medium text-slate-600">{_html.escape(lot.vendor, quote=True)}</span>
                        <span class="font-mono text-slate-400">&#8377;{lot.price:,.2f}</span>
                        <span class="font-mono font-bold text-slate-800">{lot.quantity} {h_uom or ''}</span>
                    </div>"""
                for lot in lots
            )
        else:
            # No live lot ledger — reconstruct a per-vendor view from this item's GRN
            # receipt history so the real vendors it was received from still appear.
            # This runs for EVERY asset class: consumables/tools whose lots are
            # consumed or pre-date the ledger, AND serial-tracked assets (which keep
            # their vendor/price only in GRN, never in the lot ledger).
            grn_groups = {}
            for g in grn_hist_by_item.get(it.id, []):
                v = (g.vendor or it.supplier or "Approved Vendor").strip() or "Approved Vendor"
                p = round(float(g.unit_price if g.unit_price is not None else (it.price or 0.0)), 2)
                grn_groups[(v, p)] = grn_groups.get((v, p), 0) + (g.quantity or 0)

            if grn_groups:
                lot_caption = f"Vendor / Acquisition for {h_name} (from receipt history)"
                footer_core = ('Quantities above are total received per vendor, not the live '
                               'remaining balance.')
                if is_tracked(it):
                    footer_core += ' Open the <b>Serials</b> register for unit-level detail.'
                lot_footer = footer_core
                lot_lines = "".join(
                    f"""<div class="flex items-center justify-between gap-3 py-1 px-2 odd:bg-white even:bg-slate-50 rounded">
                            <span class="font-medium text-slate-600">{_html.escape(v, quote=True)}</span>
                            <span class="font-mono text-slate-400">&#8377;{p:,.2f}</span>
                            <span class="font-mono font-bold text-slate-800">{qty} {h_uom or ''}</span>
                        </div>"""
                    for (v, p), qty in grn_groups.items()
                )
            else:
                lot_caption = f"Vendor / Acquisition for {h_name}"
                lot_lines = '<div class="text-slate-400 italic px-2 py-1">No vendor records yet — recorded as a single combined balance.</div>'

        inventory_rows += f"""
        <tr id="lot-row-{it.id}" class="lot-detail-row hidden bg-slate-50/70 border-b border-slate-100">
            <td colspan="{n_cols}" class="p-3">
                <div class="text-[10px] uppercase font-bold text-slate-400 tracking-wider mb-1.5">{lot_caption}</div>
                <div class="grid grid-cols-1 sm:grid-cols-2 lg:grid-cols-3 gap-1 text-[11px]">
                    {lot_lines}
                </div>
                <div class="text-[10px] text-slate-400 mt-1.5">{lot_footer}</div>
            </td>
        </tr>
        """

    req_rows = ""
    reqs = db.query(models.ProcurementRequest).order_by(models.ProcurementRequest.timestamp.desc()).all()
    is_admin = current_user.role == "Admin"

    for r in reqs:
        date_str = fmt_ist(r.timestamp, "%d-%m %H:%M")

        # Status badge
        status_badge_map = {
            "Pending":      '<span class="bg-amber-100 text-amber-700 font-bold px-2 py-0.5 rounded text-[10px]">Pending</span>',
            "Order Pending":'<span class="bg-blue-100 text-blue-700 font-bold px-2 py-0.5 rounded text-[10px]">Order Pending</span>',
            "Order Placed": '<span class="bg-indigo-100 text-indigo-700 font-bold px-2 py-0.5 rounded text-[10px]">Order Placed</span>',
            "Delivered":    '<span class="bg-emerald-100 text-emerald-700 font-bold px-2 py-0.5 rounded text-[10px]">&#10003; Delivered</span>',
            "Rejected":     '<span class="bg-rose-100 text-rose-700 font-bold px-2 py-0.5 rounded text-[10px]">Rejected</span>',
        }
        badge = status_badge_map.get(r.status, f'<span class="bg-slate-100 text-slate-600 font-bold px-2 py-0.5 rounded text-[10px]">{r.status}</span>')

        # Item description and code/spec cells
        if getattr(r, 'is_new_item', False) and not r.item:
            raw_name = r.new_item_name or ""
            if raw_name.startswith("[CODE:"):
                assigned_code = raw_name[len("[CODE:"):raw_name.index("]")]
                real_display_name = raw_name[raw_name.index("]") + 1:].strip() or "New Item"
                code_cell = f"<span class='bg-emerald-100 text-emerald-700 font-black font-mono px-2 py-0.5 rounded text-[11px] border border-emerald-200'>&#10003; {assigned_code}</span>"
                code_assigned = True
            else:
                real_display_name = raw_name or "Unlisted Item"
                code_cell = f"""<form id='assign-code-form-{r.id}' action='/procurement/assign-code/{r.id}' method='POST' class='flex gap-1 items-center justify-center'
                        onsubmit='return handleAssignCodeSubmit(event, {r.id})'>
                        <input type='hidden' name='force' value='false' id='assign-force-{r.id}'>
                        <input type='text' name='new_item_code' id='assign-input-{r.id}' placeholder='e.g. EIPL-ST-09' required
                            class='border border-amber-300 bg-amber-50 text-slate-900 font-mono text-[11px] px-2 py-1 rounded-lg w-28 uppercase focus:outline-none focus:border-indigo-500'
                            style='text-transform:uppercase'
                            oninput='this.value=this.value.toUpperCase()'>
                        <button type='submit' class='bg-amber-500 hover:bg-amber-600 text-white font-bold text-[10px] px-2 py-1 rounded-lg'>Assign</button>
                    </form>"""
                code_assigned = False

            raw_spec = r.detailed_specification or ""
            spec_missing = (not raw_spec.strip() or raw_spec.strip().lower() == "no specs")
            if spec_missing:
                spec_cell = f"""<form action='/procurement/assign-spec/{r.id}' method='POST' class='flex flex-col gap-1 items-center'>
                        <textarea name='specification' rows='2' placeholder='Enter specs...' required
                            class='border border-rose-300 bg-rose-50 text-slate-900 text-[11px] px-2 py-1 rounded-lg w-36 focus:outline-none focus:border-indigo-500 resize-none'></textarea>
                        <button type='submit' class='bg-rose-500 hover:bg-rose-600 text-white font-bold text-[10px] px-2 py-1 rounded-lg'>Submit</button>
                    </form>"""
                spec_filled = False
            else:
                spec_cell = f"<span class='bg-emerald-100 text-emerald-700 font-bold px-2 py-0.5 rounded text-[11px] border border-emerald-200' title='{raw_spec}'>&#10003; Spec Added</span>"
                spec_filled = True

            item_desc_display = f"<span class='text-amber-600 font-bold'>[NEW]</span> <b>{real_display_name}</b><br><span class='text-[10px] text-slate-400 italic font-normal'>Specs: {r.detailed_specification or 'No Specs'}</span>"
            est_value_text = "&#8377; TBD"
        else:
            item_name_str = r.item.name if r.item else (r.new_item_name or "Catalog Item Request")
            item_code_str = f" ({r.item.item_code})" if r.item else ""
            item_desc_display = f"<b>{item_name_str}</b>{item_code_str}"
            est_value_text = f"&#8377;{r.total_estimated_cost:,.2f}" if r.total_estimated_cost else "&#8377;0.00"
            code_cell = "<span class='text-slate-400 text-[10px]'>&#8212;</span>"
            spec_cell = "<span class='text-slate-400 text-[10px]'>&#8212;</span>"
            code_assigned = True
            spec_filled = True

        if getattr(r, 'department', None):
            item_desc_display += f" <span class='text-[10px] text-indigo-600 font-semibold bg-indigo-50 px-1.5 py-0.5 rounded'>[{r.department}]</span>"

        # Workflow execution column
        import html as _html
        flow = ""
        if r.status == "Pending":
            edit_btn = ""
            if is_admin or r.requested_by_id == current_user.id:
                h_dept = _html.escape(r.department or "", quote=True)
                cur_item_id = r.item_id if r.item_id else 0
                edit_btn = f"""<button data-action="edit-request" data-id="{r.id}" data-qty="{r.quantity}" data-dept="{h_dept}" data-itemid="{cur_item_id}" class="text-indigo-600 font-bold hover:underline text-[10px]">Edit</button>"""
            if is_admin:
                if not code_assigned:
                    approve_btn = "<span class='text-[10px] text-slate-400 italic'>Assign code first &#8593;</span>"
                elif not spec_filled:
                    approve_btn = "<span class='text-[10px] text-slate-400 italic'>Add spec first &#8593;</span>"
                else:
                    approve_btn = f"""<form action="/procurement/action/{r.id}/accept" method="POST" class="inline"><input type="submit" value="Approve" class="bg-emerald-600 hover:bg-emerald-700 text-white font-bold px-1.5 py-0.5 rounded cursor-pointer text-[10px]"></form>"""
                flow = f"""<div class="flex items-center gap-1 justify-center">{edit_btn} {approve_btn}
                    <form action="/procurement/action/{r.id}/reject" method="POST" class="inline"><input type="submit" value="Reject" class="bg-rose-600 hover:bg-rose-700 text-white font-bold px-1.5 py-0.5 rounded cursor-pointer text-[10px]"></form>
                    </div>"""
            else:
                flow = f"""<div class="flex items-center gap-1 justify-center">{edit_btn}<span class="text-slate-400 italic text-[10px]">Awaiting Review</span></div>"""

        elif r.status == "Order Pending":
            po_name = _html.escape(r.item.name if r.item else (r.new_item_name or "Item"), quote=True)
            po_code = _html.escape(r.item.item_code if r.item else "N/A", quote=True)
            print_btn = f"""<button data-action="print-po" data-id="{r.id}" data-name="{po_name}" data-code="{po_code}" data-qty="{r.quantity}" class="bg-slate-700 hover:bg-slate-800 text-white text-[10px] font-bold px-2 py-0.5 rounded">&#128438; Print PO</button>"""
            has_pricing = bool(r.vendor and r.vendor.strip()) and (r.unit_price is not None)
            if is_admin:
                cur_vendor = _html.escape(r.vendor or "", quote=True)
                cur_rate = f"{r.unit_price:.2f}" if r.unit_price is not None else ""
                # Admin enters vendor + rate here; the Order button unlocks only after both are saved
                pricing_form = f"""<form action="/procurement/set-pricing/{r.id}" method="POST" class="flex items-center gap-1 justify-center flex-wrap">
                        <input type="text" name="vendor" value="{cur_vendor}" placeholder="Vendor" required
                            class="border border-slate-300 bg-white text-slate-900 text-[10px] px-1.5 py-1 rounded-lg w-24 focus:outline-none focus:border-indigo-500">
                        <input type="number" step="0.01" min="0" name="unit_price" value="{cur_rate}" placeholder="Rate &#8377;" required
                            class="border border-slate-300 bg-white text-slate-900 font-mono text-[10px] px-1.5 py-1 rounded-lg w-20 focus:outline-none focus:border-indigo-500">
                        <button type="submit" class="bg-amber-500 hover:bg-amber-600 text-white font-bold text-[10px] px-2 py-1 rounded-lg">Save</button>
                    </form>"""
                if has_pricing:
                    order_btn = f"""<form action="/procurement/order/{r.id}" method="POST" class="inline"><input type="submit" value="Order" class="bg-indigo-600 hover:bg-indigo-700 text-white font-bold px-2 py-0.5 rounded cursor-pointer text-[10px]"></form>"""
                else:
                    order_btn = """<span class="bg-slate-200 text-slate-400 font-bold px-2 py-0.5 rounded text-[10px] cursor-not-allowed" title="Enter vendor & rate to enable">Order</span>"""
                flow = f"""<div class="flex flex-col items-center gap-1.5">{pricing_form}<div class="flex items-center gap-1 justify-center">{print_btn} {order_btn}</div></div>"""
            else:
                vendor_note = f"<span class='text-[10px] text-slate-500'>Vendor: <b>{_html.escape(r.vendor)}</b></span>" if has_pricing else "<span class='text-[10px] text-slate-400 italic'>Awaiting vendor &amp; rate</span>"
                flow = f"""<div class="flex flex-col items-center gap-1">{print_btn}{vendor_note}</div>"""

        elif r.status == "Order Placed":
            po_name = _html.escape(r.item.name if r.item else (r.new_item_name or "Item"), quote=True)
            po_code = _html.escape(r.item.item_code if r.item else "N/A", quote=True)
            flow = f"""<div class="flex justify-center"><button data-action="print-po" data-id="{r.id}" data-name="{po_name}" data-code="{po_code}" data-qty="{r.quantity}" class="bg-slate-700 hover:bg-slate-800 text-white text-[10px] font-bold px-2 py-0.5 rounded">&#128438; Print PO</button></div>"""

        elif r.status == "Delivered":
            flow = """<div class="flex justify-center"><span class="text-emerald-600 font-bold text-[10px]">&#10003; Delivered</span></div>"""

        else:
            flow = """<div class="flex justify-center"><span class="text-slate-400 line-through text-[11px]">Closed</span></div>"""

        # Est Value: only show for admin
        est_td = f'<td class="p-3 font-mono text-slate-600 text-center whitespace-nowrap" data-sort="{r.total_estimated_cost or 0}">{est_value_text}</td>' if is_admin else ""

        # Build message thread
        messages = db.query(models.ProcurementMessage).filter(
            models.ProcurementMessage.request_id == r.id
        ).order_by(models.ProcurementMessage.timestamp.asc()).all()

        msg_bubbles = ""
        for msg in messages:
            sender_role = msg.sender.role if msg.sender else "Staff"
            sender_label = "Admin" if sender_role == "Admin" else (msg.sender.username if msg.sender else "Staff")
            bubble_color = "bg-indigo-50 border-indigo-100 text-indigo-900" if sender_role == "Admin" else "bg-slate-50 border-slate-200 text-slate-800"
            label_color = "text-indigo-600" if sender_role == "Admin" else "text-slate-500"
            msg_time = fmt_ist(msg.timestamp, "%d-%m %H:%M")
            safe_msg = _html.escape(msg.message)
            msg_bubbles += f"""<div class='border rounded-lg p-2 {bubble_color} mb-1'>
                <div class='flex justify-between items-center mb-0.5'>
                    <span class='font-bold text-[10px] {label_color}'>{sender_label}</span>
                    <span class='text-[9px] text-slate-400 font-mono'>{msg_time}</span>
                </div>
                <p class='text-[11px] leading-snug'>{safe_msg}</p>
            </div>"""

        msg_thread_cell = f"""<div class='w-52'>
            <div class='max-h-24 overflow-y-auto mb-1.5 space-y-1 pr-0.5'>
                {msg_bubbles if msg_bubbles else "<p class='text-[10px] text-slate-400 italic'>No messages yet.</p>"}
            </div>
            <form action='/procurement/message/{r.id}' method='POST' class='flex gap-1'>
                <input type='text' name='message' placeholder='Type message...' required
                    class='flex-1 border border-slate-200 bg-slate-50 text-[11px] px-2 py-1 rounded-lg focus:outline-none focus:border-indigo-400 min-w-0'>
                <button type='submit' class='bg-indigo-600 hover:bg-indigo-700 text-white font-bold text-[10px] px-2 py-1 rounded-lg shrink-0'>Send</button>
            </form>
        </div>"""

        # Category: falls back to the linked item's category, then default
        req_category = getattr(r, 'category', None) or (getattr(r.item, 'category', None) if r.item else None) or DEFAULT_ASSET_CLASS

        req_rows += f"""<tr class="border-b hover:bg-slate-50 text-xs align-middle" data-row="1" data-date="{fmt_ist(r.timestamp, '%Y-%m-%d')}" data-text="{date_str} {r.requester.username if r.requester else ''} {r.status} {r.department or ''}">
            <td class="p-3 text-slate-500 font-mono whitespace-nowrap text-center">{date_str}</td>
            <td class="p-3 text-slate-800 text-center">{item_desc_display}</td>
            <td class="p-3 font-mono font-semibold text-center">{r.quantity}</td>
            <td class="p-3 font-mono text-slate-500 text-center">{getattr(r, 'uom', None) or 'Nos'}</td>
            {est_td}
            <td class="p-3 text-center">{code_cell}</td>
            <td class="p-3 text-center">{spec_cell}</td>
            <td class="p-3 text-center"><span class="bg-slate-100 px-2 py-0.5 rounded text-slate-600 text-[10px] font-semibold whitespace-nowrap">{req_category}</span></td>
            <td class="p-3 text-slate-500 whitespace-nowrap text-center">{r.requester.username if r.requester else 'System'}</td>
            <td class="p-3 text-center" data-status="{r.status}">{badge}</td>
            <td class="p-3">{msg_thread_cell}</td>
            <td class="p-3 text-center">{flow}</td>
        </tr>"""


    assigned_rows = ""
    for a in assignments:
        ts = fmt_ist(a.timestamp) or "—"
        item_name = a.item.name if a.item else 'Archived Asset'
        item_category = getattr(a.item, 'category', None) or DEFAULT_ASSET_CLASS if a.item else DEFAULT_ASSET_CLASS
        dept = getattr(a, 'department', '—') or '—'
        is_ret = bool(getattr(a, 'is_return', False))
        if is_ret:
            qty_html = f'<span class="font-mono font-bold text-amber-700">+{a.quantity}</span> <span class="bg-amber-100 text-amber-700 text-[9px] font-black px-1.5 py-0.5 rounded ml-1">RETURN</span>'
            person_label = "Returned by"
        else:
            qty_html = f'<span class="font-mono font-bold text-blue-700">{a.quantity}</span>'
            person_label = "Issued to"
        assigned_rows += f"""<tr class="border-b hover:bg-slate-50 text-xs" data-row="1">
            <td class="p-3 font-mono text-slate-400 text-[11px]">{ts}</td>
            <td class="p-3 font-bold text-slate-800">{item_name}</td>
            <td class="p-3 text-center" data-sort="{a.quantity}">{qty_html}</td>
            <td class="p-3 font-mono text-slate-500 text-center">{a.uom}</td>
            <td class="p-3 text-center"><span class="bg-slate-100 px-2 py-0.5 rounded text-slate-600 text-[10px] font-semibold whitespace-nowrap">{item_category}</span></td>
            <td class="p-3 text-slate-700 font-medium"><span class="text-[9px] text-slate-400 block leading-tight">{person_label}</span>{a.issued_to}</td>
            <td class="p-3 text-slate-500">{dept}</td>
            <td class="p-3 text-slate-400 italic text-[11px]">{a.remarks or '—'}</td>
        </tr>"""

    employee_control_panel = ""
    admin_panel = ""
    user_directory_control_panel = ""

    if current_user.role == "Admin":
        employee_list_markup = ""
        for e in employees:
            import html as _html
            h_emp_name    = _html.escape(e.name, quote=True)
            h_emp_role    = _html.escape(e.role_title, quote=True)
            h_emp_loc     = _html.escape(e.location, quote=True)
            h_emp_contact = _html.escape(e.contact, quote=True)
            employee_list_markup += f"""
            <div class='flex items-center justify-between bg-slate-50 p-2.5 rounded-lg border border-slate-200 text-[11px]'>
                <div class='space-y-0.5'>
                    <p class='font-bold text-slate-800'>{e.name}</p>
                    <p class='text-slate-500 text-[10px]'>{e.role_title} | Location: {e.location}</p>
                    <p class='text-slate-400 font-mono text-[9px]'>Ph: {e.contact}</p>
                </div>
                <div class='flex gap-1.5 ml-2'>
                    <button type='button'
                        data-action="edit-employee"
                        data-id="{e.id}" data-name="{h_emp_name}" data-role="{h_emp_role}"
                        data-loc="{h_emp_loc}" data-contact="{h_emp_contact}"
                        class='text-indigo-600 font-bold hover:underline text-[10px]'>edit</button>
                    <form action='/employees/delete/{e.id}' method='POST' class='inline delete-protected-form' onsubmit="event.preventDefault(); openAdminDeleteModal(this, 'Delete employee: {e.name}?');"><input type='submit' value='delete' class='text-rose-500 font-bold cursor-pointer hover:underline bg-transparent border-0 text-[10px]'></form>
                </div>
            </div>"""

        employee_control_panel = f"""
        <div class="bg-white p-6 rounded-xl border border-slate-200 shadow-xl mb-6">
            <h2 class="text-xs font-black text-slate-500 uppercase tracking-wider border-b pb-2 mb-4">👥 Employee Registration</h2>
            <form action="/employees/add" method="POST" class="space-y-2.5 text-xs mb-4">
                <input type="text" name="name" placeholder="Employee Full Name" required class="w-full bg-slate-50 border p-2.5 rounded-lg">
                <input type="text" name="role_title" placeholder="Designation / Role" required class="w-full bg-slate-50 border p-2.5 rounded-lg">
                <div class="grid grid-cols-2 gap-2">
                    <input type="text" name="location" placeholder="Work Station Loc" required class="w-full bg-slate-50 border p-2.5 rounded-lg">
                    <input type="text" name="contact" placeholder="Contact Mobile" required class="w-full bg-slate-50 border p-2.5 rounded-lg">
                </div>
                <button type="submit" class="w-full bg-slate-800 text-white font-bold p-2 rounded-lg">Register Personnel</button>
            </form>
            <div class="space-y-1.5 max-h-[220px] overflow-y-auto pr-1">{employee_list_markup}</div>
        </div>
        """

        admin_panel = ""

        user_list_markup = ""
        for u in users:
            import html as _html
            h_uname = _html.escape(u.full_name, quote=True)
            h_desig = _html.escape(u.designation, quote=True)
            h_loc   = _html.escape(u.workstation_location, quote=True)
            user_list_markup += f"""
            <div class='flex items-center justify-between bg-slate-50 p-2.5 rounded-lg border border-slate-200 text-[11px]'>
                <div>
                    <p class='font-bold text-slate-800'>{u.username} ({u.role})</p>
                    <p class='text-slate-500 text-[10px]'>{u.full_name} - {u.designation}</p>
                </div>
                <div class='flex gap-1.5'>
                    <button type='button'
                        data-action="edit-user"
                        data-id="{u.id}" data-name="{h_uname}" data-desig="{h_desig}"
                        data-loc="{h_loc}" data-role="{u.role}"
                        class='text-indigo-600 font-bold hover:underline text-[10px]'>edit</button>
                    <form action='/admin/users/delete/{u.id}' method='POST' class='inline delete-protected-form' onsubmit="event.preventDefault(); openAdminDeleteModal(this, 'Delete user account: {u.username}?');"><input type='submit' value='delete' class='text-rose-500 font-bold cursor-pointer hover:underline bg-transparent border-0 text-[10px]'></form>
                </div>
            </div>"""

        user_directory_control_panel = f"""
        <div class="bg-white p-6 rounded-xl border border-slate-200 shadow-xl mb-6">
            <h2 class="text-xs font-black tracking-wider text-slate-500 uppercase border-b pb-2 mb-4">❖ Grant user Access</h2>
            <form action="/admin/users/create" method="POST" class="space-y-2 text-xs mb-4">
                <div class="grid grid-cols-2 gap-2">
                    <input type="text" name="username" placeholder="Login User Id" required class="w-full bg-slate-50 border p-2 rounded-lg">
                    <input type="password" name="password" placeholder="Account Password" required class="w-full bg-slate-50 border p-2 rounded-lg">
                </div>
                <input type="text" name="full_name" placeholder="Full Employee Legal Name" required class="w-full bg-slate-50 border p-2 rounded-lg">
                <div class="grid grid-cols-2 gap-2">
                    <input type="text" name="designation" placeholder="Corporate Title / Designation" required class="w-full bg-slate-50 border p-2 rounded-lg">
                    <input type="text" name="workstation_location" placeholder="Workstation Base Location" required class="w-full bg-slate-50 border p-2 rounded-lg">
                </div>
                <select name="role" class="w-full bg-white font-bold border p-2 rounded-lg"><option value="Staff">Grant Staff Level</option><option value="Admin">Grant Full Admin</option></select>
                <button type="submit" class="w-full bg-indigo-600 text-white p-2 rounded-lg font-bold">Provision Account</button>
            </form>
            <div class="space-y-1.5 max-h-[200px] overflow-y-auto pr-1">{user_list_markup}</div>
        </div>
        """

    import json as _json
    mis_inventory_json_safe = _json.dumps([
        {"id": i.id, "name": i.name, "code": i.item_code, "stock": i.current_stock,
         "uom": getattr(i, 'uom', None) or "", "price": i.price or 0.0, "vendor": getattr(i, 'supplier', '') or ""}
        for i in items
    ])

    est_value_header = '<th class="p-3 text-center whitespace-nowrap cursor-pointer hover:text-indigo-600" onclick="sortTable(\'req\',4)">Est. Value <i class="fa-solid fa-sort text-[9px]"></i></th>' if current_user.role == "Admin" else ""

    return templates.LAYOUT_HTML\
        .replace("__LOGO_DATA_URL__", LOGO_DATA_URL)\
        .replace("__USER__", current_user.username)\
        .replace("__ROLE__", current_user.role)\
        .replace("__USER_FULL_NAME__", current_user.full_name)\
        .replace("__USER_DESIGNATION__", current_user.designation)\
        .replace("__USER_LOCATION__", current_user.workstation_location)\
        .replace("__ADMIN_PANEL__", admin_panel)\
        .replace("__EMPLOYEE_CONTROL_PANEL__", employee_control_panel)\
        .replace("__USER_DIRECTORY_CONTROL_PANEL__", user_directory_control_panel)\
        .replace("__OPTIONS__", item_options)\
        .replace("__EMPLOYEE_OPTIONS__", employee_options)\
        .replace("__INVENTORY_ROWS__", inventory_rows)\
        .replace("__REG_ROWS__", req_rows)\
        .replace("__ASSIGNED_ROWS__", assigned_rows)\
        .replace("__MIS_INVENTORY_JSON__", mis_inventory_json_safe)\
        .replace("__EST_VALUE_HEADER__", est_value_header)\
        .replace("__ASSET_CLASS_OPTIONS_JSON__", _json.dumps(ASSET_CLASSES))\
        .replace("__ASSET_CLASS_OPTIONS__", category_options_html())\
        .replace("__IS_ADMIN__", "true" if current_user.role == "Admin" else "false")